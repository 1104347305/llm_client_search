"""
导出请求日志到 Excel

用法：
    python export_logs.py                        # 导出全部，文件名自动生成
    python export_logs.py -o my_logs.xlsx        # 指定输出文件名
    python export_logs.py --agent A001           # 只导出某代理人
    python export_logs.py --since 2026-03-01     # 只导出指定日期之后的记录
"""

import argparse
import json
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DB_PATH = Path(__file__).parent / "logs" / "requests.db"

HEADERS = [
    ("id",              "ID"),
    ("agent_id",        "代理人号"),
    ("query",           "问题"),
    ("request_payload", "请求入参"),
    ("result_data",     "返回客户数据（最多3条）"),
    ("matched_level",   "命中层级"),
    ("confidence",      "置信度"),
    ("request_time",    "请求时间"),
]

LEVEL_LABELS = {1: "L1-规则", 2: "L2-模板", 3: "L3-缓存", 4: "L4-LLM", 0: "未知"}


def fetch_rows(agent_id: str = None, since: str = None) -> list:
    if not DB_PATH.exists():
        print(f"数据库文件不存在（尚未产生任何请求记录）：{DB_PATH}")
        return []

    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row

    clauses, params = [], []
    if agent_id:
        clauses.append("agent_id = ?")
        params.append(agent_id)
    if since:
        clauses.append("request_time >= ?")
        params.append(since)

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    rows = conn.execute(
        f"SELECT * FROM search_requests {where} ORDER BY id ASC",
        params,
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def format_json(raw: str, max_len: int = 500) -> str:
    """将 JSON 字符串格式化为可读文本，超长截断。"""
    try:
        obj = json.loads(raw)
        text = json.dumps(obj, ensure_ascii=False, indent=2)
    except Exception:
        text = raw or ""
    return text[:max_len] + ("…" if len(text) > max_len else "")


def format_result_data(raw: str) -> str:
    """将 result_data JSON 转为简洁的多行文本。"""
    try:
        items = json.loads(raw)
    except Exception:
        return raw or ""
    if not items:
        return "（无数据）"
    lines = []
    for i, item in enumerate(items[:3], 1):
        if isinstance(item, dict):
            name = item.get("name") or item.get("customer_name") or ""
            cid  = item.get("customer_id") or item.get("id") or ""
            lines.append(f"{i}. {name}（{cid}）" if name or cid else f"{i}. {json.dumps(item, ensure_ascii=False)[:80]}")
        else:
            lines.append(f"{i}. {str(item)[:80]}")
    return "\n".join(lines)


def write_excel(rows: list, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "请求日志"

    # ── 表头样式 ──
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F75B6")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap   = Alignment(vertical="top", wrap_text=True)

    col_widths = [6, 14, 30, 45, 35, 12, 10, 22]

    for col_idx, (_, label) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

    ws.row_dimensions[1].height = 22

    # ── 数据行 ──
    for row_idx, row in enumerate(rows, 2):
        values = [
            row["id"],
            row["agent_id"],
            row["query"],
            format_json(row["request_payload"]),
            format_result_data(row["result_data"]),
            LEVEL_LABELS.get(row["matched_level"], str(row["matched_level"])),
            round(row["confidence"], 4),
            row["request_time"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap

        # 斑马纹
        if row_idx % 2 == 0:
            fill = PatternFill("solid", fgColor="EEF4FB")
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # ── 冻结首行 ──
    ws.freeze_panes = "A2"

    # ── 汇总 sheet ──
    ws2 = wb.create_sheet("汇总")
    ws2.append(["统计项", "值"])
    ws2.append(["总记录数", len(rows)])
    level_counts = {}
    for r in rows:
        lbl = LEVEL_LABELS.get(r["matched_level"], str(r["matched_level"]))
        level_counts[lbl] = level_counts.get(lbl, 0) + 1
    for lbl, cnt in sorted(level_counts.items()):
        ws2.append([f"命中层级-{lbl}", cnt])
    if rows:
        ws2.append(["最早记录时间", rows[0]["request_time"]])
        ws2.append(["最晚记录时间", rows[-1]["request_time"]])

    for cell in ws2["A"]:
        cell.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 18

    wb.save(str(output_path))


def main():
    parser = argparse.ArgumentParser(description="导出请求日志到 Excel")
    parser.add_argument("-o", "--output", default=None, help="输出文件路径（默认自动生成）")
    parser.add_argument("--agent", default=None, help="只导出指定代理人号")
    parser.add_argument("--since", default=None, help="只导出该日期之后的记录，格式 YYYY-MM-DD")
    args = parser.parse_args()

    output_path = Path(args.output) if args.output else Path(
        f"logs/request_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"读取数据库：{DB_PATH}")
    rows = fetch_rows(agent_id=args.agent, since=args.since)
    print(f"共 {len(rows)} 条记录")

    if not rows:
        print("没有符合条件的数据，未生成文件。")
        return

    write_excel(rows, output_path)
    print(f"已导出至：{output_path.resolve()}")


if __name__ == "__main__":
    main()
