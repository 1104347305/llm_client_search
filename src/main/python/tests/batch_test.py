"""
批量测试脚本 - 测试自然语言搜索接口并导出结果到Excel

功能：
1. 从 test_queries.txt 读取测试问题
2. 调用自然语言搜索接口
3. 将查询条件和客户列表保存到 Excel 文件

依赖：
pip install openpyxl httpx
"""
import asyncio
import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any
import json


class BatchTester:
    """批量测试器"""

    def __init__(self, base_url: str = "http://localhost:8000"):
        self.base_url = base_url
        self.results = []

    async def test_query(self, query: str, agent_id: str = "A000000") -> Dict[str, Any]:
        """测试单个查询"""
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                response = await client.post(
                    f"{self.base_url}/api/v1/client_search_query_parse_no_encipher",
                    json= {
                        "user_text": query,
                        "user_id": "A000001",
                        "trace_id": "0090920020921920021e4u9ej9u9239",
                        "session_id": "1298318939931983198",
                        "source": "askbob"
                    }

                )
                response.raise_for_status()
                return response.json()['data']
        except Exception as e:
            return {
                "success": False,
                "message": str(e),
                "data": {},
                "matched_level": 0,
                "confidence": 0.0,
                "conditions": [],
                "query_logic": "AND"
            }

    async def run_batch_test(self, queries_file: str):
        """批量测试"""
        # 读取测试问题
        queries = []
        count = 0
        with open(queries_file, 'r', encoding='utf-8') as f:
            for line in f:
                # if count >= 2:
                #     continue
                line = line.strip()
                if line and not line.startswith('#'):
                    queries.append(line)
                    count += 1

        print(f"共加载 {len(queries)} 个测试问题")

        # 逐个测试
        for i, query in enumerate(queries, 1):
            print(f"\n[{i}/{len(queries)}] 测试: {query}")
            result = await self.test_query(query)
            elapsed = result['extra_output_params'].get('cost_times', 0)
            query_logic = result['extra_output_params'].get('query_logic', [])
            conditions = result['extra_output_params'].get('conditions', [])
            matched_level = result['extra_output_params'].get('matched_level', [])
            intent = result.get('robot_text', [])

            parseResult = {}
            parseResult['matched_level'] = matched_level
            parseResult['query_logic'] = query_logic
            parseResult['conditions'] = conditions
            parseResult['intent'] = intent
            parseResult['elapsed'] = elapsed

            # 保存结果
            self.results.append({
                "query": query,
                "parseResult": parseResult
            })

            # 打印简要信息
            print(f"  ✓ Level {matched_level}, "
                  f"条件数 {len(conditions)}, "
                  f"耗时 {elapsed}ms",
                  f"{query_logic}"
                  f"{conditions}",
                  f"{intent}"
                  )
            # else:
            #     print(f"  ✗ 失败: {result.get('message')}")

            # 避免请求过快
            await asyncio.sleep(0.1)

        print(f"\n测试完成！共 {len(self.results)} 个查询")

    def export_to_excel(self, output_file: str):
        """导出结果到Excel"""
        wb = Workbook()

        # 创建汇总表
        ws_summary = wb.active
        ws_summary.title = "测试结果"

        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # 表头
        summary_headers = ["序号", "查询问题", "Level", "意图", "耗时(ms)", "query_logic", "查询条件"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 填充数据
        for i, item in enumerate(self.results, 2):
            query = item["query"]
            parseResult = item["parseResult"]

            # 序号
            ws_summary.cell(i, 1, i - 1)

            # 查询问题
            ws_summary.cell(i, 2, query)

            # Level
            ws_summary.cell(i, 3, parseResult.get("matched_level", 0))

            # 置信度
            ws_summary.cell(i, 4, parseResult.get("intent", 0))

            # 耗时
            elapsed_ms = round(parseResult.get("elapsed", 0), 0)
            ws_summary.cell(i, 5, elapsed_ms)

            # query_logic
            ws_summary.cell(i, 6, parseResult.get("query_logic", "AND"))

            # 查询条件（单行JSON格式）
            conditions = parseResult.get("conditions") or []
            ws_summary.cell(i, 7, json.dumps(conditions, ensure_ascii=False))

        # 调整列宽
        ws_summary.column_dimensions['A'].width = 8
        ws_summary.column_dimensions['B'].width = 50
        ws_summary.column_dimensions['C'].width = 10
        ws_summary.column_dimensions['D'].width = 12
        ws_summary.column_dimensions['E'].width = 12
        ws_summary.column_dimensions['F'].width = 12
        ws_summary.column_dimensions['G'].width = 80

        # 保存文件
        wb.save(output_file)
        print(f"\n结果已导出到: {output_file}")


async def main():
    """主函数"""
    # 配置
    queries_file = "/Users/mickey/project/PA-ALG/llm_client_search/src/main/python/docs/字段验证集问题列表.md"
    # queries_file = "/Users/mickey/project/PA-ALG/llm_client_search/src/main/python/docs/test_questions.txt"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"test_questions_{timestamp}.xlsx"

    # 检查文件是否存在
    if not Path(queries_file).exists():
        print(f"错误: 找不到测试问题文件 {queries_file}")
        return

    # 创建测试器
    tester = BatchTester(base_url="http://localhost:8000")

    # 运行批量测试
    await tester.run_batch_test(queries_file)

    # 导出结果
    tester.export_to_excel(output_file)

    print("\n测试统计:")
    total = len(tester.results)
    # success = sum(1 for r in tester.results if r["code"]==0)
    print(f"  总数: {total}")
    # print(f"  成功: {success}")
    # print(f"  失败: {total - success}")

    # 按层级统计
    level_stats = {}
    for r in tester.results:
        level = r["parseResult"].get("matched_level", 0)
        level_stats[level] = level_stats.get(level, 0) + 1

    print("\n层级分布:")
    for level in sorted(level_stats.keys()):
        print(f"  Level {level}: {level_stats[level]} 个")


if __name__ == "__main__":
    asyncio.run(main())
