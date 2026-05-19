import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parents[4]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from src.main.python.tools.iteration_pipeline.evaluator import (
    build_expected_candidates,
    build_expected_candidates_from_config_examples,
    build_intent_gold_from_batch_excel,
    build_intent_label_candidates,
    build_label_candidates,
    build_skill_eval_from_batch_excel,
    compare_result,
    evaluate_intent_cases,
    evaluate_question_batch,
    expected_intent_lines_from_case,
    expected_intent_summary_from_case,
    extract_intent_lines,
    load_question_batch,
    prepare_intent_review_workbook,
    compare_intent_lines,
    write_expected_candidates_excel,
    write_intent_gold_jsonl,
    write_intent_label_candidates_excel,
    write_label_candidates_excel,
)
from src.main.python.tools.iteration_pipeline.report_writer import (
    write_batch_eval_artifacts,
    write_intent_eval_artifacts,
    write_skill_eval_artifacts,
)


def test_load_question_batch_from_text(tmp_path: Path):
    input_path = tmp_path / "questions.md"
    input_path.write_text(
        "\n".join(
            [
                "# 标题",
                "1. 查询高温客户",
                "- 找一下北京客户",
                "",
            ]
        ),
        encoding="utf-8",
    )

    cases = load_question_batch(input_path)

    assert cases == [
        {"id": "q_0001", "query": "查询高温客户", "tags": ["ungraded"]},
        {"id": "q_0002", "query": "找一下北京客户", "tags": ["ungraded"]},
    ]


def test_load_question_batch_from_jsonl_with_expected(tmp_path: Path):
    input_path = tmp_path / "questions.jsonl"
    input_path.write_text(
        (
            '{"id":"case_1","query":"高温客户","expected":{"query_logic":"AND",'
            '"conditions":[{"field":"clientTemperature","operator":"MATCH","value":"高温"}]},'
            '"tags":["positive"]}\n'
        ),
        encoding="utf-8",
    )

    cases = load_question_batch(input_path)

    assert cases[0]["id"] == "case_1"
    assert cases[0]["query"] == "高温客户"
    assert cases[0]["expected"]["conditions"][0]["field"] == "clientTemperature"
    assert cases[0]["tags"] == ["positive"]


def test_ungraded_client_search_eval_set_has_no_expected_answers():
    input_path = REPO_ROOT / "src/main/python/docs/eval_sets/ungraded_client_search_queries.jsonl"

    cases = load_question_batch(input_path)

    assert len(cases) == 427
    assert cases[0] == {"id": "q_0001", "query": "张三", "tags": ["ungraded"]}
    assert all(case["tags"] == ["ungraded"] for case in cases)
    assert all("expected" not in case for case in cases)


def test_compare_result_exact_match():
    expected = {
        "query_logic": "AND",
        "conditions": [{"field": "clientTemperature", "operator": "MATCH", "value": "高温"}],
    }
    actual = {
        "query_logic": "AND",
        "conditions": [{"field": "clientTemperature", "operator": "MATCH", "value": "高温"}],
    }

    assert compare_result(expected, actual)["exact_match"] is True


def test_expected_intent_lines_from_expected_uses_intent_summary_operator_labels():
    lines = expected_intent_lines_from_case(
        {
            "expected": {
                "query_logic": "AND",
                "conditions": [{"field": "clientAge", "operator": "GTE", "value": 45}],
            }
        }
    )

    assert lines == ["客户年龄≥45的客户"]


def test_expected_intent_summary_uses_no_conditions_message_from_labels():
    summary = expected_intent_summary_from_case(
        {"expected": {"query_logic": "AND", "conditions": []}}
    )

    assert summary == "未识别到明确查询条件"


def test_expected_intent_summary_uses_unsupported_only_message_from_labels():
    summary = expected_intent_summary_from_case(
        {
            "expected": {
                "query_logic": "AND",
                "conditions": [
                    {
                        "field": "policies_cooling_off",
                        "operator": "RANGE",
                        "value": {"min": "2026-01-01", "max": "2026-01-31"},
                    }
                ],
            }
        }
    )

    assert summary == "提示：犹豫期时间暂不支持搜索，无法进行查询。"


def test_expected_intent_summary_uses_supported_and_unsupported_message_from_labels():
    summary = expected_intent_summary_from_case(
        {
            "expected": {
                "query_logic": "AND",
                "conditions": [
                    {"field": "clientSex", "operator": "MATCH", "value": "男"},
                    {
                        "field": "policies_cooling_off",
                        "operator": "RANGE",
                        "value": {"min": "2026-01-01", "max": "2026-01-31"},
                    },
                ],
            }
        }
    )

    assert summary == "客户性别为男的客户\n提示：犹豫期时间暂不支持搜索，系统将按可支持字段搜索。"


def test_intent_summary_exact_match_catches_query_logic_connector():
    expected_summary = expected_intent_summary_from_case(
        {
            "expected": {
                "query_logic": "AND",
                "conditions": [
                    {"field": "clientSex", "operator": "MATCH", "value": "男"},
                    {"field": "clientTemperature", "operator": "MATCH", "value": "高温"},
                ],
            }
        }
    )
    actual_summary = expected_summary.replace("\n并且", "\n或者")

    comparison = compare_intent_lines(
        extract_intent_lines(expected_summary),
        actual_summary,
        expected_summary,
    )

    assert comparison["intent_line_exact_match"] is True
    assert comparison["intent_summary_exact_match"] is False
    assert comparison["intent_exact_match"] is False


def test_extract_intent_lines_ignores_prefix_and_connectors():
    lines = extract_intent_lines("系统识别查询条件：\n客户年龄≥45\n并且\n客户性别为女的客户")

    assert lines == ["客户年龄≥45", "客户性别为女的客户"]


@pytest.mark.asyncio
async def test_intent_eval_compares_summary_lines(monkeypatch):
    async def fake_call_parse_api(client, base_url, query):
        return (
            {
                "query_logic": "AND",
                "conditions": [{"field": "clientAge", "operator": "GTE", "value": 45}],
                "intent_summary": "系统识别查询条件：\n客户年龄≥45的客户",
                "matched_level": 2,
            },
            10.0,
            None,
        )

    monkeypatch.setattr(
        "src.main.python.tools.iteration_pipeline.evaluator._call_parse_api",
        fake_call_parse_api,
    )

    result = await evaluate_intent_cases(
        [
            {
                "id": "intent_1",
                "query": "45岁以上客户",
                "expected": {
                    "query_logic": "AND",
                    "conditions": [{"field": "clientAge", "operator": "GTE", "value": 45}],
                },
            }
        ],
        options=type("Options", (), {"base_url": "http://example.test", "timeout_seconds": 1, "concurrency": 1})(),
    )

    assert result["summary"]["intent_exact_match_rate"] == 1.0
    assert result["summary"]["intent_line_recall"] == 1.0
    assert result["cases"][0]["expected_intent_lines"] == ["客户年龄≥45的客户"]


@pytest.mark.asyncio
async def test_batch_eval_summary_exposes_total_accuracy(monkeypatch):
    async def fake_call_parse_api(client, base_url, query):
        return (
            {
                "query_logic": "AND",
                "conditions": [{"field": "clientTemperature", "operator": "MATCH", "value": "高温"}],
                "matched_level": 2,
            },
            10.0,
            None,
        )

    monkeypatch.setattr(
        "src.main.python.tools.iteration_pipeline.evaluator._call_parse_api",
        fake_call_parse_api,
    )

    result = await evaluate_question_batch(
        [
            {
                "id": "case_1",
                "query": "高温客户",
                "expected": {
                    "query_logic": "AND",
                    "conditions": [{"field": "clientTemperature", "operator": "MATCH", "value": "高温"}],
                },
            }
        ],
        options=type("Options", (), {"base_url": "http://example.test", "timeout_seconds": 1, "concurrency": 1})(),
    )

    assert result["summary"]["total_accuracy"] == 1.0


@pytest.mark.asyncio
async def test_batch_eval_progress_callback(monkeypatch):
    async def fake_call_parse_api(client, base_url, query):
        return (
            {
                "query_logic": "AND",
                "conditions": [],
                "matched_level": 1,
            },
            5.0,
            None,
        )

    monkeypatch.setattr(
        "src.main.python.tools.iteration_pipeline.evaluator._call_parse_api",
        fake_call_parse_api,
    )
    events = []

    await evaluate_question_batch(
        [
            {"id": "case_1", "query": "问题1"},
            {"id": "case_2", "query": "问题2"},
        ],
        options=type("Options", (), {"base_url": "http://example.test", "timeout_seconds": 1, "concurrency": 2})(),
        progress_callback=events.append,
    )

    assert [event["completed"] for event in events] == [1, 2]
    assert events[-1]["total"] == 2
    assert events[-1]["errors"] == 0


@pytest.mark.asyncio
async def test_ungraded_batch_eval_exposes_observable_metrics(monkeypatch):
    async def fake_call_parse_api(client, base_url, query):
        return (
            {
                "query_logic": "AND",
                "conditions": [{"field": "city", "operator": "MATCH", "value": "北京"}],
                "matched_level": 2,
            },
            5.0,
            None,
        )

    monkeypatch.setattr(
        "src.main.python.tools.iteration_pipeline.evaluator._call_parse_api",
        fake_call_parse_api,
    )

    result = await evaluate_question_batch(
        [{"id": "case_1", "query": "北京客户"}],
        options=type("Options", (), {"base_url": "http://example.test", "timeout_seconds": 1, "concurrency": 1})(),
    )

    summary = result["summary"]
    assert summary["graded_total"] == 0
    assert summary["total_accuracy"] is None
    assert summary["api_success_rate"] == 1.0
    assert summary["condition_non_empty_rate"] == 1.0
    assert summary["known_level_rate"] == 1.0


def test_build_expected_candidates_uses_actual_parse_result():
    candidates = build_expected_candidates(
        {
            "cases": [
                {
                    "id": "case_1",
                    "query": "高温客户",
                    "tags": ["ungraded"],
                    "actual": {
                        "query_logic": "AND",
                        "conditions": [{"field": "clientTemperature", "operator": "MATCH", "value": "高温"}],
                        "matched_level": 2,
                    },
                    "elapsed_ms": 10.0,
                    "error": None,
                }
            ]
        }
    )

    assert candidates[0]["label_status"] == "candidate"
    assert candidates[0]["expected"]["conditions"][0]["field"] == "clientTemperature"
    assert candidates[0]["label_meta"]["needs_review"] is True


def test_build_expected_candidates_records_label_source():
    candidates = build_expected_candidates(
        {
            "cases": [
                {
                    "id": "case_1",
                    "query": "高温客户",
                    "tags": [],
                    "actual": {
                        "query_logic": "AND",
                        "conditions": [{"field": "clientTemperature", "operator": "MATCH", "value": "高温"}],
                        "matched_level": 2,
                    },
                    "elapsed_ms": 10.0,
                    "error": None,
                }
            ]
        },
        source="current_config_inprocess",
        extra_label_meta={"labeling_mode": "config_rules", "allow_l4": False},
    )

    assert candidates[0]["label_meta"]["source"] == "current_config_inprocess"
    assert candidates[0]["label_meta"]["labeling_mode"] == "config_rules"
    assert candidates[0]["label_meta"]["allow_l4"] is False


def test_build_expected_candidates_keeps_empty_results_for_manual_labeling():
    candidates = build_expected_candidates(
        {
            "cases": [
                {
                    "id": "case_1",
                    "query": "未知问题",
                    "tags": [],
                    "actual": {"query_logic": "AND", "conditions": [], "matched_level": None},
                    "elapsed_ms": 10.0,
                    "error": None,
                }
            ]
        }
    )

    assert candidates[0]["label_status"] == "manual_required"
    assert "expected" not in candidates[0]
    assert candidates[0]["label_meta"]["reason"] == "empty_parse_result"


def test_build_expected_candidates_from_config_examples(tmp_path: Path):
    field_definitions_path = tmp_path / "field_definitions_args.yaml"
    field_definitions_path.write_text(
        """
intents:
  - id: gender
    field: clientSex
    operator: MATCH
    examples:
      - query: 男客户
        output: {field: clientSex, operator: MATCH, value: 男}
    negative_examples:
      - query: 子女是男性的客户
        reason: 家庭成员性别不是客户本人性别
""",
        encoding="utf-8",
    )

    candidates = build_expected_candidates_from_config_examples(
        [
            {"id": "case_1", "query": "男客户"},
            {"id": "case_2", "query": "子女是男性的客户"},
            {"id": "case_3", "query": "未覆盖问题"},
        ],
        field_definitions_path=field_definitions_path,
    )

    assert candidates[0]["label_status"] == "candidate"
    assert candidates[0]["expected"]["conditions"][0]["field"] == "clientSex"
    assert candidates[0]["robot_text"]
    assert candidates[0]["label_meta"]["source"] == "field_definitions_examples"
    assert candidates[0]["label_meta"]["needs_review"] is False
    assert candidates[1]["expected"]["conditions"] == []
    assert candidates[2]["label_status"] == "manual_required"
    assert "expected" not in candidates[2]


def test_write_batch_eval_artifacts_writes_excel(tmp_path: Path):
    eval_result = {
        "summary": {
            "total": 1,
            "total_accuracy": 1.0,
            "exact_match_rate": 1.0,
            "error_count": 0,
            "level_distribution": {"2": 1},
        },
        "cases": [
            {
                "id": "case_1",
                "query": "男客户",
                "tags": ["positive"],
                "graded": True,
                "expected": {
                    "query_logic": "AND",
                    "conditions": [{"field": "clientSex", "operator": "MATCH", "value": "男"}],
                },
                "actual": {
                    "query_logic": "AND",
                    "conditions": [{"field": "clientSex", "operator": "MATCH", "value": "男"}],
                    "matched_level": 2,
                },
                "comparison": {
                    "exact_match": True,
                    "field_match": True,
                    "operator_match": True,
                    "query_logic_match": True,
                    "missing_conditions": [],
                    "unexpected_conditions": [],
                },
                "elapsed_ms": 10.0,
                "error": None,
            }
        ],
        "failed_cases": [],
    }

    artifacts = write_batch_eval_artifacts(tmp_path / "questions.jsonl", tmp_path, eval_result)

    assert artifacts["excel"].exists()
    workbook = load_workbook(artifacts["excel"])
    assert workbook.sheetnames == ["summary", "cases", "failed_cases"]
    assert workbook["summary"]["A1"].value == "metric"
    assert workbook["cases"]["A2"].value == "case_1"


def test_write_intent_eval_artifacts_writes_excel(tmp_path: Path):
    eval_result = {
        "summary": {
            "total": 1,
            "graded_total": 1,
            "ungraded_total": 0,
            "intent_exact_match_rate": 1.0,
            "intent_line_recall": 1.0,
            "intent_line_precision": 1.0,
            "error_count": 0,
            "level_distribution": {"2": 1},
        },
        "cases": [
            {
                "id": "intent_1",
                "query": "45岁以上客户",
                "tags": ["age"],
                "graded": True,
                "expected_intent_lines": ["客户年龄≥45的客户"],
                "actual_intent_summary": "系统识别查询条件：\n客户年龄≥45的客户",
                "actual_intent_lines": ["客户年龄≥45的客户"],
                "actual": {
                    "conditions": [{"field": "clientAge", "operator": "GTE", "value": 45}],
                    "matched_level": 2,
                },
                "intent_comparison": {
                    "intent_exact_match": True,
                    "intent_line_recall": 1.0,
                    "intent_line_precision": 1.0,
                    "missing_intent_lines": [],
                    "unexpected_intent_lines": [],
                },
                "elapsed_ms": 10.0,
                "error": None,
            }
        ],
        "failed_cases": [],
    }

    artifacts = write_intent_eval_artifacts(tmp_path / "intent.jsonl", tmp_path, eval_result)

    assert artifacts["excel"].exists()
    workbook = load_workbook(artifacts["excel"])
    assert workbook.sheetnames == ["summary", "cases", "failed_cases"]
    assert workbook["cases"]["A2"].value == "intent_1"


def test_write_expected_candidates_excel(tmp_path: Path):
    output_path = tmp_path / "candidate_expected.xlsx"
    candidates = [
        {
            "id": "case_1",
            "query": "男客户",
            "robot_text": "系统识别查询条件：\n客户性别为男的客户",
            "tags": ["auto_expected"],
            "label_status": "candidate",
            "expected": {
                "query_logic": "AND",
                "conditions": [{"field": "clientSex", "operator": "MATCH", "value": "男"}],
            },
            "label_meta": {
                "source": "field_definitions_examples",
                "labeling_mode": "config_examples_exact_match",
                "needs_review": False,
            },
        }
    ]

    write_expected_candidates_excel(candidates, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook["expected_candidates"]
    assert sheet["A1"].value == "id"
    assert sheet["A2"].value == "case_1"
    assert sheet["C1"].value == "robot_text"
    assert sheet["C2"].value == "系统识别查询条件：\n客户性别为男的客户"
    assert sheet["I2"].value == "AND"
    assert "clientSex" in sheet["J2"].value


def test_build_label_candidates_from_examples_and_enums(tmp_path: Path):
    field_definitions_path = tmp_path / "field_definitions_args.yaml"
    field_definitions_path.write_text(
        """
intents:
  - id: gender
    field: clientSex
    operator: MATCH
    value_type: enum
    enum_ref: clientSex
    examples:
      - query: 男客户
        output: {field: clientSex, operator: MATCH, value: 男}
""",
        encoding="utf-8",
    )
    field_enums_path = tmp_path / "field_enums_args.yaml"
    field_enums_path.write_text(
        """
clientSex:
  values: [男, 女]
""",
        encoding="utf-8",
    )
    value_mappings_path = tmp_path / "value_mappings_args.yaml"
    value_mappings_path.write_text("clientSex: {}\n", encoding="utf-8")

    rows = build_label_candidates(
        [{"id": "case_1", "query": "男客户"}],
        field_definitions_path=field_definitions_path,
        field_enums_path=field_enums_path,
        value_mappings_path=value_mappings_path,
    )

    assert rows[0]["agreement_status"] == "auto_approved"
    assert rows[0]["final_expected"]["conditions"][0]["field"] == "clientSex"
    assert rows[0]["candidates"][0]["source"] == "config_example"
    assert any("enum:clientSex.男" in evidence for evidence in rows[0]["candidates"][0]["evidence"])


def test_write_label_candidates_excel(tmp_path: Path):
    output_path = tmp_path / "label_candidates.xlsx"
    rows = [
        {
            "id": "case_1",
            "query": "男客户",
            "agreement_status": "auto_approved",
            "review_status": "auto_approved",
            "candidate_count": 1,
            "final_robot_text": "系统识别查询条件：\n客户性别为男的客户",
            "final_expected": {
                "query_logic": "AND",
                "conditions": [{"field": "clientSex", "operator": "MATCH", "value": "男"}],
            },
            "candidates": [
                {
                    "source": "config_example",
                    "confidence": 0.98,
                    "expected": {
                        "query_logic": "AND",
                        "conditions": [{"field": "clientSex", "operator": "MATCH", "value": "男"}],
                    },
                    "evidence": ["field_definitions.gender.examples"],
                }
            ],
        }
    ]

    write_label_candidates_excel(rows, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook["label_candidates"]
    assert sheet["A1"].value == "id"
    assert sheet["A2"].value == "case_1"
    assert sheet["C2"].value == "auto_approved"


def test_build_intent_label_candidates_adds_parser_review_candidate():
    rows = build_intent_label_candidates(
        [{"id": "case_1", "query": "45岁以上客户", "tags": ["ungraded"]}],
        [
            {
                "id": "case_1",
                "query": "45岁以上客户",
                "agreement_status": "manual_required",
                "candidates": [],
            }
        ],
        {
            "cases": [
                {
                    "id": "case_1",
                    "actual": {
                        "query_logic": "AND",
                        "conditions": [{"field": "clientAge", "operator": "GTE", "value": 45}],
                        "intent_summary": "客户年龄≥45的客户",
                        "matched_level": 2,
                    },
                    "error": None,
                }
            ]
        },
    )

    assert rows[0]["review_status"] == "parser_review"
    assert rows[0]["candidate_intent_lines"] == ["客户年龄≥45的客户"]
    assert rows[0]["candidates"][0]["source"] == "current_parse_api"


def test_build_intent_label_candidates_preserves_auto_approved_static_candidate():
    expected = {
        "query_logic": "AND",
        "conditions": [{"field": "clientSex", "operator": "MATCH", "value": "男"}],
    }

    rows = build_intent_label_candidates(
        [{"id": "case_1", "query": "男客户", "tags": ["ungraded"]}],
        [
            {
                "id": "case_1",
                "query": "男客户",
                "agreement_status": "auto_approved",
                "candidates": [
                    {
                        "source": "config_example",
                        "confidence": 0.98,
                        "expected": expected,
                        "robot_text": "客户性别为男的客户",
                        "evidence": ["field_definitions.gender.examples"],
                    }
                ],
            }
        ],
        {
            "cases": [
                {
                    "id": "case_1",
                    "actual": {
                        "query_logic": "AND",
                        "conditions": [{"field": "clientSex", "operator": "MATCH", "value": "男"}],
                        "intent_summary": "客户性别为男的客户",
                        "matched_level": 2,
                    },
                    "error": None,
                }
            ]
        },
    )

    assert rows[0]["review_status"] == "auto_approved"
    assert rows[0]["final_expected"] == expected
    assert rows[0]["final_expected_intent_lines"] == ["客户性别为男的客户"]


def test_write_intent_label_candidates_excel(tmp_path: Path):
    output_path = tmp_path / "intent_label_candidates.xlsx"
    rows = [
        {
            "id": "case_1",
            "query": "45岁以上客户",
            "review_status": "parser_review",
            "static_agreement_status": "manual_required",
            "candidate_count": 1,
            "candidate_intent_summary": "客户年龄≥45的客户",
            "candidate_intent_lines": ["客户年龄≥45的客户"],
            "final_expected_intent_lines": [],
            "final_expected": None,
            "candidates": [
                {
                    "source": "current_parse_api",
                    "confidence": 0.72,
                    "intent_summary": "客户年龄≥45的客户",
                    "expected": {
                        "query_logic": "AND",
                        "conditions": [{"field": "clientAge", "operator": "GTE", "value": 45}],
                    },
                    "evidence": ["matched_level=2"],
                }
            ],
            "error": None,
        }
    ]

    write_intent_label_candidates_excel(rows, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook["intent_label_candidates"]
    assert sheet["A1"].value == "id"
    assert sheet["A2"].value == "case_1"
    assert sheet["C2"].value == "parser_review"


def test_build_intent_gold_from_batch_excel_uses_reviewed_intent_summary(tmp_path: Path):
    input_path = tmp_path / "batch_eval_result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "cases"
    sheet.append(["id", "query", "tags", "intent_summary", "review_status", "issue_type", "final_intent_summary"])
    sheet.append(["case_1", "男客户", '["gender"]', "客户性别为男的客户", "通过", "", ""])
    sheet.append(["case_2", "身份证号110101", "", "证件类型包含身份证的客户", "错误", "value_wrong", "证件号为110101的客户"])
    sheet.append(["case_3", "未处理", "", "客户年龄≥45的客户", "错误", "missing_condition", ""])
    workbook.save(input_path)

    result = build_intent_gold_from_batch_excel(input_path)

    assert result["summary"]["generated"] == 2
    assert result["summary"]["skipped"] == 1
    assert result["rows"][0]["expected_intent"] == "客户性别为男的客户"
    assert result["rows"][1]["expected_intent"] == "证件号为110101的客户"
    assert result["rows"][1]["label_meta"]["issue_type"] == "value_wrong"


def test_write_intent_gold_jsonl(tmp_path: Path):
    output_path = tmp_path / "intent_gold.jsonl"

    write_intent_gold_jsonl(
        [
            {
                "id": "case_1",
                "query": "男客户",
                "expected_intent": "客户性别为男的客户",
                "expected_intent_lines": ["客户性别为男的客户"],
            }
        ],
        output_path,
    )

    assert "客户性别为男的客户" in output_path.read_text(encoding="utf-8")


def test_prepare_intent_review_workbook_adds_risk_columns(tmp_path: Path):
    input_path = tmp_path / "batch_eval_result.xlsx"
    output_path = tmp_path / "intent_review.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "cases"
    sheet.append(["id", "query", "actual_conditions", "intent_summary", "matched_level", "error"])
    sheet.append([
        "case_1",
        "男客户",
        '[{"field":"clientSex","operator":"MATCH","value":"男"}]',
        "客户性别为男的客户",
        2,
        "",
    ])
    sheet.append([
        "case_2",
        "手机号为13800138000",
        '[{"field":"searchClientName","operator":"MATCH","value":"13800138000"}]',
        "客户姓名为13800138000的客户",
        2,
        "",
    ])
    sheet.append([
        "case_3",
        "30多岁A2低温已婚有娃有意外伤害保险的",
        (
            '[{"field":"pCategorys","operator":"MATCH","value":"意外伤害保险"},'
            '{"field":"polNoInfo.plancodeinfo.abbrname","operator":"MATCH","value":"意外伤害"}]'
        ),
        "产品类别为意外伤害保险\n并且投保险种简称为意外伤害的客户",
        2,
        "",
    ])
    sheet.append([
        "case_4",
        "50岁B类有车险有综拓理赔有有效保单的",
        (
            '[{"field":"isBuyInsuranceCar","operator":"MATCH","value":"车险"},'
            '{"field":"assetsCondition","operator":"CONTAINS","value":["有车","有房有车"]}]'
        ),
        "车险客户\n并且客户资产状况包含有车、有房有车的客户",
        2,
        "",
    ])
    sheet.append([
        "case_5",
        "60岁以下已婚(外)孙子女有年金的",
        (
            '[{"field":"familyInfo.familyrelation","operator":"CONTAINS","value":["(外)孙子女"]},'
            '{"field":"polNoInfo.plancodeinfo.abbrname","operator":"MATCH","value":"子女"}]'
        ),
        "有(外)孙子女\n并且投保险种简称为子女的客户",
        2,
        "",
    ])
    sheet.append([
        "case_6",
        "e生保理赔过的客户",
        (
            '[{"field":"polNoInfo.claimdatainfo.claimplancodename","operator":"MATCH","value":"e生保"},'
            '{"field":"newValueLabel","operator":"MATCH","value":"e"}]'
        ),
        "理赔险种为e生保\n并且客户价值标签为e的客户",
        2,
        "",
    ])
    sheet.append([
        "case_7",
        "40岁黄金VIP寿险客户",
        (
            '[{"field":"clientAge","operator":"MATCH","value":40},'
            '{"field":"vipType","operator":"CONTAINS","value":["原黄金VIP","黄金V1","黄金V2","黄金V3"]},'
            '{"field":"planAbbrNames","operator":"EXISTS"},'
            '{"field":"polNoInfo.plancodeinfo.plantypedesc","operator":"MATCH","value":"寿险"}]'
        ),
        "客户年龄=40\n并且客户VIP等级包含原黄金VIP、黄金V1、黄金V2、黄金V3\n并且有寿险产品\n并且投保险种类别为寿险的客户",
        2,
        "",
    ])
    sheet.append([
        "case_8",
        "有车的客户",
        '[{"field":"assetsCondition","operator":"CONTAINS","value":["有车"]}]',
        "客户资产状况包含有车的客户",
        2,
        "",
    ])
    sheet.append([
        "case_9",
        "有房的客户",
        '[{"field":"assetsCondition","operator":"CONTAINS","value":["有房"]}]',
        "客户资产状况包含有房的客户",
        2,
        "",
    ])
    sheet.append([
        "case_10",
        "B类客户有万能险的",
        (
            '[{"field":"newValueLabel","operator":"MATCH","value":"B"},'
            '{"field":"insuranceType","operator":"MATCH","value":"万能型"}]'
        ),
        "客户价值标签为B\n并且产品类型为万能型的客户",
        2,
        "",
    ])
    sheet.append([
        "case_11",
        "缴费有效保单客户",
        '[{"field":"polNoInfo.polStatus","operator":"CONTAINS","value":["交费有效","自垫交清","交清","减额交清","免交","自垫有效"]}]',
        "保单状态包含交费有效、自垫交清、交清、减额交清、免交、自垫有效的客户",
        2,
        "",
    ])
    sheet.append([
        "case_12",
        "客户价值B有子女且子女年龄17岁以下",
        (
            '[{"field":"newValueLabel","operator":"MATCH","value":"B"},'
            '{"field":"familyInfo.hasChild","operator":"EXISTS"},'
            '{"field":"familyInfo.childAge","operator":"LTE","value":17}]'
        ),
        "客户价值标签为B\n并且有子女\n并且子女年龄≤17的客户",
        2,
        "",
    ])
    sheet.append([
        "case_13",
        "投保日期是今年的客户",
        "[]",
        "提示：投保日期暂不支持搜索，无法进行查询。",
        "unknown",
        "",
    ])
    sheet.append([
        "case_14",
        "男客户犹豫期时间在1月的",
        '[{"field":"clientSex","operator":"MATCH","value":"男"}]',
        "客户性别为男的客户\n提示：犹豫期时间暂不支持搜索，系统将按可支持字段搜索。",
        "unknown",
        "",
    ])
    sheet.append([
        "case_15",
        "B以上客户",
        '[{"field":"newValueLabel","operator":"CONTAINS","value":["B","A4","A3","A2","A1"]}]',
        "客户价值标签包含B、A4、A3、A2、A1的客户",
        2,
        "",
    ])
    sheet.append([
        "case_16",
        "B及以上客户",
        '[{"field":"newValueLabel","operator":"CONTAINS","value":["B","A4","A3","A2","A1"]}]',
        "客户价值标签包含B、A4、A3、A2、A1的客户",
        2,
        "",
    ])
    sheet.append([
        "case_17",
        "本科以上的客户",
        '[{"field":"education","operator":"CONTAINS","value":["硕士研究生","博士研究生","博士后"]}]',
        "客户学历包含硕士研究生、博士研究生、博士后的客户",
        2,
        "",
    ])
    sheet.append([
        "case_18",
        "45岁以上高温B类有车有重疾险有车险的",
        (
            '[{"field":"clientAge","operator":"GTE","value":45},'
            '{"field":"clientTemperature","operator":"MATCH","value":"高温"},'
            '{"field":"newValueLabel","operator":"MATCH","value":"B"},'
            '{"field":"assetsCondition","operator":"CONTAINS","value":["有车"]},'
            '{"field":"pCategorys","operator":"CONTAINS","value":["疾病保险"]},'
            '{"field":"isBuyInsuranceCar","operator":"MATCH","value":"车险"}]'
        ),
        "客户年龄≥45\n并且客户温度为高温\n并且客户价值标签为B\n并且客户资产状况包含有车\n并且产品类别包含疾病保险\n并且车险客户",
        2,
        "",
    ])
    sheet.append([
        "case_19",
        "A1客户有未成年子女的",
        (
            '[{"field":"newValueLabel","operator":"MATCH","value":"A1"},'
            '{"field":"familyInfo.hasChild","operator":"EXISTS"},'
            '{"field":"familyInfo.childAge","operator":"LTE","value":17}]'
        ),
        "客户价值标签为A1\n并且有子女\n并且子女年龄≤17的客户",
        2,
        "",
    ])
    sheet.append([
        "case_20",
        "铂金以上客户",
        '[{"field":"vipType","operator":"CONTAINS","value":["钻石VIP","金钻VIP","黑钻VIP"]}]',
        "客户VIP等级包含钻石VIP、金钻VIP、黑钻VIP的客户",
        2,
        "",
    ])
    sheet.append([
        "case_21",
        "不是本科学历的客户",
        '[{"field":"education","operator":"CONTAINS","value":["大学本科生"]}]',
        "客户学历包含大学本科生的客户",
        2,
        "",
    ])
    sheet.append([
        "case_22",
        "买过保单数量3张以上的客户",
        "[]",
        "提示：保单数量暂不支持搜索，无法进行查询。",
        "unknown",
        "",
    ])
    sheet.append([
        "case_23",
        "配置了意外伤害保险",
        '[{"field":"pCategorys","operator":"MATCH","value":"意外伤害保险"}]',
        "产品类别为意外伤害保险的客户",
        2,
        "",
    ])
    sheet.append([
        "case_24",
        "中温A类有车险且有车的",
        (
            '[{"field":"clientTemperature","operator":"MATCH","value":"中温"},'
            '{"field":"newValueLabel","operator":"MATCH","value":"A1"},'
            '{"field":"isBuyInsuranceCar","operator":"MATCH","value":"车险"},'
            '{"field":"assetsCondition","operator":"MATCH","value":"有车"}]'
        ),
        "客户温度为中温\n并且客户价值标签为A1\n并且车险客户\n并且客户资产状况为有车的客户",
        2,
        "",
    ])
    workbook.save(input_path)

    result = prepare_intent_review_workbook(input_path, output_path)

    assert result["rows"] == 24
    reviewed = load_workbook(output_path)
    review_sheet = reviewed["cases"]
    headers = [cell.value for cell in review_sheet[1]]
    assert "review_status" in headers
    assert "risk_level" in headers
    status_col = headers.index("review_status") + 1
    issue_col = headers.index("issue_type") + 1
    final_col = headers.index("final_intent_summary") + 1
    possible_col = headers.index("possible_intent_summary") + 1
    assert review_sheet.cell(2, status_col).value == "待抽查"
    assert review_sheet.cell(2, final_col).value == "客户性别为男的客户"
    assert review_sheet.cell(3, status_col).value == "待复核"
    assert "suspected_mobile_field_wrong" in review_sheet.cell(3, issue_col).value
    assert "suspected_duplicate_value_multi_fields" in review_sheet.cell(4, issue_col).value
    assert "suspected_duplicate_value_multi_fields" in review_sheet.cell(5, issue_col).value
    assert "suspected_duplicate_value_multi_fields" in review_sheet.cell(6, issue_col).value
    assert "suspected_value_truncated_or_too_short" in review_sheet.cell(7, issue_col).value
    assert "suspected_invalid_enum_value" in review_sheet.cell(7, issue_col).value
    assert review_sheet.cell(7, possible_col).value == "理赔险种为e生保的客户"
    assert "suspected_duplicate_intent_concept" in review_sheet.cell(8, issue_col).value
    assert "有寿险产品" in review_sheet.cell(8, possible_col).value
    assert "投保险种类别为寿险" not in review_sheet.cell(8, possible_col).value
    assert "suspected_special_field_value_incomplete" in review_sheet.cell(9, issue_col).value
    assert review_sheet.cell(9, possible_col).value == "客户资产状况包含有车、有房有车的客户"
    assert "suspected_special_field_value_incomplete" in review_sheet.cell(10, issue_col).value
    assert review_sheet.cell(10, possible_col).value == "客户资产状况包含有房、有房有车的客户"
    assert review_sheet.cell(11, status_col).value == "待抽查"
    assert review_sheet.cell(11, issue_col).value == "none"
    assert review_sheet.cell(11, final_col).value == "客户价值标签为B\n并且产品类型为万能型的客户"
    assert "suspected_special_field_value_incomplete" in review_sheet.cell(12, issue_col).value
    assert review_sheet.cell(12, possible_col).value == "保单状态包含交费有效的客户"
    assert review_sheet.cell(13, status_col).value == "待抽查"
    assert review_sheet.cell(13, issue_col).value == "none"
    assert review_sheet.cell(14, status_col).value == "待抽查"
    assert review_sheet.cell(14, issue_col).value == "none"
    assert review_sheet.cell(14, final_col).value == "提示：投保日期暂不支持搜索，无法进行查询。"
    assert review_sheet.cell(15, status_col).value == "待抽查"
    assert review_sheet.cell(15, issue_col).value == "none"
    assert "suspected_special_field_value_incomplete" in review_sheet.cell(16, issue_col).value
    assert review_sheet.cell(16, possible_col).value == "客户价值标签包含A4、A3、A2、A1的客户"
    assert review_sheet.cell(17, status_col).value == "待抽查"
    assert review_sheet.cell(17, issue_col).value == "none"
    assert review_sheet.cell(18, status_col).value == "待抽查"
    assert review_sheet.cell(18, issue_col).value == "none"
    assert "suspected_special_field_value_incomplete" in review_sheet.cell(19, issue_col).value
    assert "客户资产状况包含有车、有房有车" in review_sheet.cell(19, possible_col).value
    assert review_sheet.cell(20, status_col).value == "待抽查"
    assert review_sheet.cell(20, issue_col).value == "none"
    assert review_sheet.cell(21, status_col).value == "待抽查"
    assert review_sheet.cell(21, issue_col).value == "none"
    assert review_sheet.cell(22, status_col).value == "待复核"
    assert "suspected_negation_missing" in review_sheet.cell(22, issue_col).value
    assert review_sheet.cell(23, status_col).value == "待抽查"
    assert review_sheet.cell(23, issue_col).value == "none"
    assert review_sheet.cell(23, final_col).value == "提示：保单数量暂不支持搜索，无法进行查询。"
    assert review_sheet.cell(24, status_col).value == "待抽查"
    assert review_sheet.cell(24, issue_col).value == "none"
    assert "suspected_duplicate_value_multi_fields" not in review_sheet.cell(25, issue_col).value
    assert "suspected_special_field_value_incomplete" in review_sheet.cell(25, issue_col).value


def test_build_skill_eval_from_batch_excel_uses_skill_rubric(tmp_path: Path):
    input_path = tmp_path / "batch_eval_result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "cases"
    sheet.append(["id", "query", "actual_conditions", "intent_summary", "matched_level", "error"])
    sheet.append([
        "case_1",
        "男客户",
        '[{"field":"clientSex","operator":"MATCH","value":"男"}]',
        "客户性别为男的客户",
        2,
        "",
    ])
    sheet.append([
        "case_2",
        "e生保理赔过的客户",
        (
            '[{"field":"polNoInfo.claimdatainfo.claimplancodename","operator":"MATCH","value":"e生保"},'
            '{"field":"newValueLabel","operator":"MATCH","value":"e"}]'
        ),
        "理赔险种为e生保\n并且客户价值标签为e的客户",
        2,
        "",
    ])
    sheet.append([
        "case_3",
        "有车的客户",
        '[{"field":"assetsCondition","operator":"CONTAINS","value":["有车"]}]',
        "客户资产状况包含有车的客户",
        2,
        "",
    ])
    sheet.append([
        "case_4",
        "有车险客户",
        '[{"field":"isBuyInsuranceCar","operator":"MATCH","value":"车险"}]',
        "车险客户",
        2,
        "",
    ])
    sheet.append([
        "case_5",
        "B类客户有万能险的",
        (
            '[{"field":"newValueLabel","operator":"MATCH","value":"B"},'
            '{"field":"insuranceType","operator":"MATCH","value":"万能型"}]'
        ),
        "客户价值标签为B\n并且产品类型为万能型的客户",
        2,
        "",
    ])
    sheet.append([
        "case_6",
        "查找45岁以上未配置养老险的客户",
        (
            '[{"field":"clientAge","operator":"GTE","value":45},'
            '{"field":"isBuyPension","operator":"MATCH","value":"有购买"}]'
        ),
        "客户年龄≥45\n并且是否养老险客户为有购买的客户",
        2,
        "",
    ])
    sheet.append([
        "case_7",
        "查找45岁以上未配置养老险的客户",
        (
            '[{"field":"clientAge","operator":"GTE","value":45},'
            '{"field":"isBuyPension","operator":"MATCH","value":"没有购买"}]'
        ),
        "客户年龄≥45\n并且是否养老险客户为没有购买的客户",
        2,
        "",
    ])
    sheet.append([
        "case_8",
        "购买了意健险的客户",
        '[{"field":"validSinsPol","operator":"CONTAINS","value":["意健险"]}]',
        "有效综拓保单包含意健险的客户",
        2,
        "",
    ])
    workbook.save(input_path)

    result = build_skill_eval_from_batch_excel(
        input_path,
        skill_path=REPO_ROOT / "src/main/python/docs/eval_skills/client_search_intent_eval/SKILL.md",
    )

    assert result["summary"]["total"] == 8
    assert result["summary"]["verdict_counts"]["pass"] == 5
    assert result["summary"]["verdict_counts"]["fail"] == 3
    assert "value_wrong" in result["cases"][1]["error_types"]
    assert result["cases"][1]["expected_intent_summary"] == "理赔险种为e生保的客户"
    assert "value_wrong" in result["cases"][2]["error_types"]
    assert result["cases"][2]["expected_intent_summary"] == "客户资产状况包含有车、有房有车的客户"
    assert result["cases"][3]["verdict"] == "pass"
    assert result["cases"][4]["verdict"] == "pass"
    assert "value_wrong" in result["cases"][5]["error_types"]
    assert "没有购买" in result["cases"][5]["reason"]
    assert result["cases"][6]["verdict"] == "pass"
    assert result["cases"][7]["verdict"] == "pass"


def test_write_skill_eval_artifacts(tmp_path: Path):
    result = {
        "summary": {
            "total": 1,
            "pass_rate": 1.0,
            "fail_rate": 0.0,
            "uncertain_rate": 0.0,
            "error_type_counts": {},
        },
        "cases": [
            {
                "id": "case_1",
                "query": "男客户",
                "verdict": "pass",
                "confidence": 0.82,
                "severity": "low",
                "error_types": [],
                "reason": "ok",
                "expected_intent_summary": "客户性别为男的客户",
                "actual_intent_summary": "客户性别为男的客户",
                "actual_conditions": [{"field": "clientSex", "operator": "MATCH", "value": "男"}],
                "matched_level": 2,
                "deterministic_findings": [],
            }
        ],
        "candidate_gold": [
            {
                "id": "case_1",
                "query": "男客户",
                "expected_intent": "客户性别为男的客户",
                "expected_intent_lines": ["客户性别为男的客户"],
            }
        ],
    }

    artifacts = write_skill_eval_artifacts(tmp_path / "batch_eval_result.xlsx", tmp_path, result)

    assert artifacts["excel"].exists()
    assert artifacts["candidate_gold"].exists()
    workbook = load_workbook(artifacts["excel"])
    assert workbook.sheetnames == ["summary", "cases", "error_counts"]
