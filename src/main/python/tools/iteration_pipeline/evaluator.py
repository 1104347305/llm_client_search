from __future__ import annotations

import asyncio
import csv
import json
import os
import re
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

import httpx
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class EvalOptions:
    base_url: str = "http://localhost:8000"
    timeout_seconds: float = 30.0
    concurrency: int = 4


@dataclass
class LLMJudgeOptions:
    model: str
    api_key: str
    base_url: str | None = None
    timeout_seconds: float = 60.0
    concurrency: int = 4
    max_retries: int = 2


ProgressCallback = Callable[[dict[str, Any]], None]

ACCEPTED_REVIEW_STATUSES = {
    "accepted",
    "accept",
    "pass",
    "passed",
    "correct",
    "ok",
    "true",
    "1",
    "通过",
    "正确",
    "无错误",
    "可用",
}

SKILL_ERROR_TYPE_MAP = {
    "api_error": "unparsed",
    "missing_condition": "unparsed",
    "unknown_level": "unparsed",
    "suspected_mobile_field_wrong": "field_wrong",
    "suspected_id_no_field_wrong": "field_wrong",
    "suspected_policy_no_field_wrong": "field_wrong",
    "suspected_negation_missing": "operator_wrong",
    "suspected_logic_wrong": "logic_wrong",
    "suspected_operator_wrong": "operator_wrong",
    "suspected_duplicate_value_multi_fields": "duplicate_condition_across_fields",
    "suspected_duplicate_same_field_value_operator": "operator_wrong",
    "suspected_special_field_value_incomplete": "value_wrong",
    "suspected_polarity_value_wrong": "value_wrong",
    "suspected_value_truncated_or_too_short": "value_wrong",
    "suspected_duplicate_intent_concept": "extra_condition",
    "suspected_invalid_enum_value": "value_wrong",
}


def _normalize_value(value: Any) -> Any:
    if isinstance(value, list):
        return [_normalize_value(item) for item in value]
    if isinstance(value, dict):
        return {str(key): _normalize_value(value[key]) for key in sorted(value)}
    return value


def _normalize_condition(condition: dict[str, Any]) -> dict[str, Any]:
    normalized = {
        "field": condition.get("field"),
        "operator": condition.get("operator"),
    }
    if condition.get("operator") not in {"EXISTS", "NOT_EXISTS"}:
        normalized["value"] = _normalize_value(condition.get("value"))
    return normalized


def _condition_key(condition: dict[str, Any]) -> str:
    return json.dumps(_normalize_condition(condition), ensure_ascii=False, sort_keys=True)


def compare_result(expected: dict[str, Any], actual: dict[str, Any]) -> dict[str, Any]:
    expected_conditions = expected.get("conditions") or []
    actual_conditions = actual.get("conditions") or []
    expected_logic = expected.get("query_logic") or "AND"
    actual_logic = actual.get("query_logic") or "AND"

    expected_condition_keys = {_condition_key(item) for item in expected_conditions}
    actual_condition_keys = {_condition_key(item) for item in actual_conditions}

    missing = sorted(expected_condition_keys - actual_condition_keys)
    unexpected = sorted(actual_condition_keys - expected_condition_keys)
    exact_match = expected_logic == actual_logic and not missing and not unexpected

    expected_fields = {item.get("field") for item in expected_conditions}
    actual_fields = {item.get("field") for item in actual_conditions}
    expected_operators = {(item.get("field"), item.get("operator")) for item in expected_conditions}
    actual_operators = {(item.get("field"), item.get("operator")) for item in actual_conditions}

    return {
        "exact_match": exact_match,
        "query_logic_match": expected_logic == actual_logic,
        "field_match": expected_fields.issubset(actual_fields),
        "operator_match": expected_operators.issubset(actual_operators),
        "missing_conditions": missing,
        "unexpected_conditions": unexpected,
    }


def _extract_parse_result(response_json: dict[str, Any]) -> dict[str, Any]:
    data = response_json.get("data") or {}
    extra = data.get("extra_output_params") or {}
    if not isinstance(extra, dict):
        return {"query_logic": "AND", "conditions": []}
    return {
        "query_logic": extra.get("query_logic") or "AND",
        "conditions": extra.get("conditions") or [],
        "matched_level": extra.get("matched_level"),
        "cost_times": extra.get("cost_times"),
        "intent_summary": extra.get("intent_summary") or data.get("robot_text"),
        "matched_patterns": extra.get("matched_patterns"),
        "rewritten_query": extra.get("rewritten_query"),
    }


async def _call_parse_api(client: httpx.AsyncClient, base_url: str, query: str) -> tuple[dict[str, Any], float, str | None]:
    started = time.perf_counter()
    try:
        response = await client.post(
            f"{base_url.rstrip('/')}/api/v1/client_search_query_parse_no_encipher",
            json={
                "user_text": query,
                "user_id": "iteration_pipeline",
                "trace_id": f"iteration-{int(time.time() * 1000)}",
                "session_id": "iteration_pipeline",
                "source": "askbob",
            },
        )
        elapsed_ms = (time.perf_counter() - started) * 1000
        response.raise_for_status()
        return _extract_parse_result(response.json()), elapsed_ms, None
    except Exception as exc:  # noqa: BLE001 - evaluation should record every failure
        elapsed_ms = (time.perf_counter() - started) * 1000
        return {"query_logic": "AND", "conditions": []}, elapsed_ms, str(exc)


async def evaluate_cases(cases: list[dict[str, Any]], options: EvalOptions) -> dict[str, Any]:
    semaphore = asyncio.Semaphore(max(1, options.concurrency))
    timeout = httpx.Timeout(options.timeout_seconds)

    async with httpx.AsyncClient(timeout=timeout) as client:
        async def evaluate_one(case: dict[str, Any]) -> dict[str, Any]:
            async with semaphore:
                actual, elapsed_ms, error = await _call_parse_api(client, options.base_url, str(case.get("query") or ""))
                expected = case.get("expected") or {}
                comparison = compare_result(expected, actual)
                tags = case.get("tags") or []
                is_negative = "negative" in tags or not (expected.get("conditions") or [])
                return {
                    "id": case.get("id"),
                    "query": case.get("query"),
                    "tags": tags,
                    "expected": expected,
                    "actual": actual,
                    "comparison": comparison,
                    "elapsed_ms": elapsed_ms,
                    "error": error,
                    "is_negative": is_negative,
                }

        case_results = await asyncio.gather(*(evaluate_one(case) for case in cases))

    total = len(case_results)
    exact_matches = sum(1 for item in case_results if item["comparison"]["exact_match"])
    field_matches = sum(1 for item in case_results if item["comparison"]["field_match"])
    operator_matches = sum(1 for item in case_results if item["comparison"]["operator_match"])
    positive_cases = [item for item in case_results if not item["is_negative"]]
    negative_cases = [item for item in case_results if item["is_negative"]]
    empty_positive = [
        item for item in positive_cases
        if not item["actual"].get("conditions")
    ]
    false_positive = [
        item for item in negative_cases
        if item["actual"].get("conditions")
    ]
    latencies = [item["elapsed_ms"] for item in case_results]
    level_distribution: dict[str, int] = {}
    for item in case_results:
        level = str(item["actual"].get("matched_level") or "unknown")
        level_distribution[level] = level_distribution.get(level, 0) + 1

    summary = {
        "total": total,
        "total_accuracy": exact_matches / total if total else 0,
        "overall_accuracy": exact_matches / total if total else 0,
        "exact_match_rate": exact_matches / total if total else 0,
        "field_match_rate": field_matches / total if total else 0,
        "operator_match_rate": operator_matches / total if total else 0,
        "empty_rate": len(empty_positive) / len(positive_cases) if positive_cases else 0,
        "false_positive_rate": len(false_positive) / len(negative_cases) if negative_cases else 0,
        "avg_latency_ms": sum(latencies) / len(latencies) if latencies else 0,
        "p95_latency_ms": sorted(latencies)[int(len(latencies) * 0.95) - 1] if latencies else 0,
        "level_distribution": level_distribution,
        "error_count": sum(1 for item in case_results if item["error"]),
    }
    failed_cases = [item for item in case_results if not item["comparison"]["exact_match"] or item["error"]]

    return {
        "summary": summary,
        "cases": case_results,
        "failed_cases": failed_cases,
    }


def load_question_batch(path: str | Path) -> list[dict[str, Any]]:
    """Load ad-hoc evaluation questions from txt/md, jsonl, or csv.

    Supported JSONL/CSV fields:
    - id: optional case id
    - query or question: user question
    - expected: JSON object/string with expected query_logic/conditions
    - tags: JSON list or comma separated string
    """
    input_path = Path(path)
    suffix = input_path.suffix.lower()
    if suffix == ".jsonl":
        return _load_question_jsonl(input_path)
    if suffix == ".csv":
        return _load_question_csv(input_path)
    return _load_question_text(input_path)


def _load_question_jsonl(path: Path) -> list[dict[str, Any]]:
    cases: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as file:
        for line_number, line in enumerate(file, start=1):
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            try:
                item = json.loads(stripped)
            except json.JSONDecodeError as exc:
                raise ValueError(f"{path}:{line_number} invalid JSON: {exc}") from exc
            if not isinstance(item, dict):
                raise ValueError(f"{path}:{line_number} must be a JSON object")
            cases.append(_normalize_loaded_case(item, len(cases) + 1))
    return cases


def _load_question_csv(path: Path) -> list[dict[str, Any]]:
    cases: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        for row_number, row in enumerate(reader, start=2):
            query = row.get("query") or row.get("question") or row.get("问题")
            if not query:
                raise ValueError(f"{path}:{row_number} missing query/question column")
            item: dict[str, Any] = {
                "id": row.get("id") or row.get("ID"),
                "query": query,
            }
            expected = row.get("expected") or row.get("预期结果")
            if expected:
                item["expected"] = _parse_json_field(expected, f"{path}:{row_number}.expected")
            tags = row.get("tags") or row.get("标签")
            if tags:
                item["tags"] = _parse_tags(tags)
            cases.append(_normalize_loaded_case(item, len(cases) + 1))
    return cases


def _load_question_text(path: Path) -> list[dict[str, Any]]:
    cases: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as file:
        for line in file:
            query = line.strip()
            if not query or query.startswith("#"):
                continue
            if query.startswith("|") or set(query) <= {"-", "|", " "}:
                continue
            query = re.sub(r"^\s*(?:[-*]\s+|\d+[.、)]\s*)", "", query).strip()
            if query:
                cases.append({"id": f"q_{len(cases) + 1:04d}", "query": query, "tags": ["ungraded"]})
    return cases


def _normalize_loaded_case(item: dict[str, Any], index: int) -> dict[str, Any]:
    query = item.get("query") or item.get("question") or item.get("问题")
    if not query:
        raise ValueError(f"case #{index} missing query/question")
    case: dict[str, Any] = {
        "id": str(item.get("id") or f"q_{index:04d}"),
        "query": str(query),
        "tags": _parse_tags(item.get("tags") or item.get("标签") or []),
    }
    expected_marker_present = "expected" in item or "预期结果" in item
    expected = item.get("expected", item.get("预期结果"))
    if expected_marker_present and expected is not None and expected != "":
        case["expected"] = expected if isinstance(expected, dict) else _parse_json_field(str(expected), f"case #{index}.expected")
    expected_intent_lines = item.get("expected_intent_lines") or item.get("预期意图行")
    if expected_intent_lines:
        if isinstance(expected_intent_lines, str):
            parsed_lines = _parse_json_field(expected_intent_lines, f"case #{index}.expected_intent_lines")
        else:
            parsed_lines = expected_intent_lines
        if not isinstance(parsed_lines, list):
            raise ValueError(f"case #{index}.expected_intent_lines must be a JSON list")
        case["expected_intent_lines"] = [str(line) for line in parsed_lines if str(line).strip()]
    expected_intent = item.get("expected_intent") or item.get("预期意图")
    if expected_intent:
        case["expected_intent"] = str(expected_intent)
    if not case["tags"]:
        case["tags"] = ["graded"] if ("expected" in case or "expected_intent_lines" in case or "expected_intent" in case) else ["ungraded"]
    return case


def _parse_json_field(value: str, label: str) -> Any:
    try:
        return json.loads(value)
    except json.JSONDecodeError as exc:
        raise ValueError(f"{label} invalid JSON: {exc}") from exc


def _parse_tags(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return []
        if stripped.startswith("["):
            parsed = _parse_json_field(stripped, "tags")
            if isinstance(parsed, list):
                return [str(item).strip() for item in parsed if str(item).strip()]
        return [item.strip() for item in stripped.split(",") if item.strip()]
    return []


def build_expected_candidates(
    eval_result: dict[str, Any],
    *,
    include_empty_expected: bool = False,
    source: str = "current_parse_api",
    extra_label_meta: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    candidates: list[dict[str, Any]] = []
    for item in eval_result.get("cases") or []:
        actual = item.get("actual") or {}
        conditions = actual.get("conditions") or []
        tags = list(item.get("tags") or [])
        if "auto_expected" not in tags:
            tags.append("auto_expected")

        output: dict[str, Any] = {
            "id": item.get("id"),
            "query": item.get("query"),
            "tags": tags,
            "label_meta": {
                "source": source,
                "generated_at": generated_at,
                "needs_review": True,
                "matched_level": actual.get("matched_level"),
                "elapsed_ms": item.get("elapsed_ms"),
                "error": item.get("error"),
                **(extra_label_meta or {}),
            },
        }

        if item.get("error"):
            output["label_status"] = "manual_required"
            output["label_meta"]["reason"] = "api_error"
        elif conditions or include_empty_expected:
            output["label_status"] = "candidate"
            output["expected"] = {
                "query_logic": actual.get("query_logic") or "AND",
                "conditions": conditions,
            }
            output["robot_text"] = actual.get("intent_summary") or _robot_text_from_expected(output["expected"])
        else:
            output["label_status"] = "manual_required"
            output["label_meta"]["reason"] = "empty_parse_result"

        candidates.append(output)
    return candidates


def _condition_to_json(condition: Any) -> dict[str, Any]:
    if hasattr(condition, "model_dump"):
        return condition.model_dump(mode="json", exclude_none=True)
    if isinstance(condition, dict):
        return condition
    return dict(condition)


async def evaluate_question_batch_from_config(
    cases: list[dict[str, Any]],
    options: EvalOptions,
    *,
    allow_l4: bool = False,
    progress_callback: ProgressCallback | None = None,
) -> dict[str, Any]:
    from src.main.python.config.settings import settings
    from src.main.python.steps.query_router import QueryRouter

    previous_enable_l4 = settings.ENABLE_L4
    settings.ENABLE_L4 = bool(allow_l4)
    try:
        router = QueryRouter()
        semaphore = asyncio.Semaphore(max(1, options.concurrency))

        async def evaluate_one(index: int, case: dict[str, Any]) -> tuple[int, dict[str, Any]]:
            started = time.perf_counter()
            error = None
            try:
                async with semaphore:
                    parsed = await asyncio.wait_for(
                        router.route_with_peeling(str(case.get("query") or ""), trace_id="config-label"),
                        timeout=options.timeout_seconds,
                    )
                elapsed_ms = (time.perf_counter() - started) * 1000
                actual = {
                    "query_logic": parsed.query_logic.value if hasattr(parsed.query_logic, "value") else str(parsed.query_logic or "AND"),
                    "conditions": [_condition_to_json(condition) for condition in (parsed.conditions or [])],
                    "matched_level": parsed.matched_level,
                    "confidence": parsed.confidence,
                    "matched_patterns": parsed.matched_patterns,
                    "rewritten_query": parsed.rewritten_query,
                }
            except Exception as exc:  # noqa: BLE001 - labeling should keep going
                elapsed_ms = (time.perf_counter() - started) * 1000
                error = str(exc)
                actual = {"query_logic": "AND", "conditions": [], "matched_level": None}

            result = {
                "id": case.get("id"),
                "query": case.get("query"),
                "tags": case.get("tags") or [],
                "expected": case.get("expected") if "expected" in case else None,
                "actual": actual,
                "comparison": compare_result(case.get("expected") or {}, actual) if "expected" in case else None,
                "elapsed_ms": elapsed_ms,
                "error": error,
                "graded": "expected" in case,
            }
            return index, result

        tasks = [asyncio.create_task(evaluate_one(index, case)) for index, case in enumerate(cases)]
        case_results: list[dict[str, Any] | None] = [None] * len(tasks)
        completed = 0
        errors = 0
        total_elapsed_ms = 0.0
        for task in asyncio.as_completed(tasks):
            index, result = await task
            case_results[index] = result
            completed += 1
            if result["error"]:
                errors += 1
            total_elapsed_ms += float(result["elapsed_ms"] or 0)
            if progress_callback:
                progress_callback(
                    {
                        "completed": completed,
                        "total": len(tasks),
                        "errors": errors,
                        "avg_latency_ms": total_elapsed_ms / completed if completed else 0,
                        "case_id": result.get("id"),
                        "query": result.get("query"),
                        "elapsed_ms": result.get("elapsed_ms"),
                        "matched_level": result.get("actual", {}).get("matched_level"),
                        "error": result.get("error"),
                    }
                )
    finally:
        settings.ENABLE_L4 = previous_enable_l4

    finalized_case_results = [item for item in case_results if item is not None]
    total = len(finalized_case_results)
    graded_cases = [item for item in finalized_case_results if item["graded"]]
    exact_matches = sum(1 for item in graded_cases if item["comparison"] and item["comparison"]["exact_match"])
    latencies = [item["elapsed_ms"] for item in finalized_case_results]
    non_empty_cases = [item for item in finalized_case_results if item["actual"].get("conditions")]
    known_level_cases = [item for item in finalized_case_results if item["actual"].get("matched_level") is not None]
    level_distribution: dict[str, int] = {}
    for item in finalized_case_results:
        level = str(item["actual"].get("matched_level") or "unknown")
        level_distribution[level] = level_distribution.get(level, 0) + 1

    summary = {
        "total": total,
        "graded_total": len(graded_cases),
        "ungraded_total": total - len(graded_cases),
        "graded_coverage_rate": len(graded_cases) / total if total else 0,
        "api_success_rate": (total - sum(1 for item in finalized_case_results if item["error"])) / total if total else 0,
        "condition_non_empty_rate": len(non_empty_cases) / total if total else 0,
        "known_level_rate": len(known_level_cases) / total if total else 0,
        "total_accuracy": exact_matches / len(graded_cases) if graded_cases else None,
        "overall_accuracy": exact_matches / len(graded_cases) if graded_cases else None,
        "exact_match_rate": exact_matches / len(graded_cases) if graded_cases else None,
        "avg_latency_ms": sum(latencies) / len(latencies) if latencies else 0,
        "p95_latency_ms": sorted(latencies)[int(len(latencies) * 0.95) - 1] if latencies else 0,
        "level_distribution": level_distribution,
        "error_count": sum(1 for item in finalized_case_results if item["error"]),
    }
    failed_cases = [
        item for item in graded_cases
        if item["error"] or not (item["comparison"] and item["comparison"]["exact_match"])
    ]

    return {
        "summary": summary,
        "cases": finalized_case_results,
        "failed_cases": failed_cases,
    }


def write_expected_candidates(candidates: list[dict[str, Any]], output_path: str | Path) -> Path:
    path = Path(output_path).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        for item in candidates:
            file.write(json.dumps(item, ensure_ascii=False, sort_keys=True))
            file.write("\n")
    return path


def write_expected_candidates_excel(candidates: list[dict[str, Any]], output_path: str | Path) -> Path:
    path = Path(output_path).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "expected_candidates"

    headers = [
        "id",
        "query",
        "robot_text",
        "label_status",
        "needs_review",
        "source",
        "labeling_mode",
        "reason",
        "expected_query_logic",
        "expected_conditions",
        "tags",
        "matched_level",
        "elapsed_ms",
        "error",
        "label_meta",
    ]
    _write_excel_header(sheet, headers)

    for row_index, item in enumerate(candidates, start=2):
        expected = item.get("expected") or {}
        label_meta = item.get("label_meta") or {}
        values = [
            item.get("id"),
            item.get("query"),
            item.get("robot_text") or _robot_text_from_expected(expected),
            item.get("label_status"),
            label_meta.get("needs_review"),
            label_meta.get("source"),
            label_meta.get("labeling_mode"),
            label_meta.get("reason"),
            expected.get("query_logic"),
            expected.get("conditions"),
            item.get("tags"),
            label_meta.get("matched_level"),
            label_meta.get("elapsed_ms"),
            label_meta.get("error"),
            label_meta,
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column_index, _excel_scalar(value))
            if headers[column_index - 1] in {"query", "robot_text", "expected_conditions", "label_meta"}:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    _autosize_excel_columns(sheet)
    workbook.save(path)
    return path


def _write_excel_header(sheet, headers: list[str]) -> None:
    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column_index, header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _autosize_excel_columns(sheet, max_width: int = 80) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), max_width))
        sheet.column_dimensions[column_letter].width = max(10, max_length + 2)


def _excel_scalar(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def _robot_text_from_expected(expected: dict[str, Any] | None) -> str | None:
    if not expected:
        return None
    try:
        from src.main.python.models.schemas import Condition, QueryLogic
        from src.main.python.steps.intent_summary import build_intent_summary

        conditions = [Condition(**condition) for condition in (expected.get("conditions") or [])]
        query_logic = QueryLogic(expected.get("query_logic") or "AND")
        return build_intent_summary(conditions, query_logic)
    except Exception:
        return None


def extract_intent_lines(intent_summary: str | None) -> list[str]:
    """Normalize a generated intent summary into comparable intent lines.

    The operator wording itself is not duplicated here. Expected text should be
    generated by IntentSummaryService, which loads intent_summary_labels_args.yaml.
    This function only removes wrapper copy/connectors and normalizes punctuation.
    """
    if not intent_summary:
        return []

    ignored_lines = {
        "系统识别查询条件:",
        "系统识别查询条件：",
    }
    connector_lines = {"并且", "或者"}
    lines: list[str] = []
    for raw_line in str(intent_summary).splitlines():
        line = raw_line.strip()
        if not line or line in ignored_lines or line in connector_lines:
            continue
        normalized = normalize_intent_line(line)
        if normalized:
            lines.append(normalized)
    return sorted(dict.fromkeys(lines))


def normalize_intent_summary_text(intent_summary: str | None) -> str | None:
    """Normalize full summary text while preserving configured connectors/messages."""
    if intent_summary is None:
        return None
    ignored_lines = {
        "系统识别查询条件:",
        "系统识别查询条件：",
    }
    lines = [
        _normalize_intent_summary_line(line)
        for line in str(intent_summary).splitlines()
        if str(line).strip() and str(line).strip() not in ignored_lines
    ]
    return "\n".join(lines)


def _normalize_intent_summary_line(line: str) -> str:
    return (
        str(line)
        .strip()
        .replace("：", ":")
        .replace("，", ",")
        .replace("。", "")
        .replace(" ", "")
    )


def normalize_intent_line(line: str) -> str:
    normalized = _normalize_intent_summary_line(line)
    for connector in ("并且", "或者"):
        if normalized.startswith(connector) and len(normalized) > len(connector):
            return normalized[len(connector):]
    return normalized


def expected_intent_lines_from_case(case: dict[str, Any]) -> list[str]:
    if case.get("expected_intent_lines"):
        return [normalize_intent_line(line) for line in case.get("expected_intent_lines") or [] if str(line).strip()]
    if case.get("expected_intent"):
        return extract_intent_lines(str(case.get("expected_intent")))
    if case.get("expected"):
        return extract_intent_lines(_robot_text_from_expected(case.get("expected")))
    return []


def expected_intent_summary_from_case(case: dict[str, Any]) -> str | None:
    if case.get("expected_intent"):
        return str(case.get("expected_intent"))
    if case.get("expected"):
        return _robot_text_from_expected(case.get("expected"))
    if case.get("expected_intent_lines"):
        return "\n".join(str(line) for line in case.get("expected_intent_lines") or [])
    return None


def compare_intent_lines(
    expected_lines: list[str],
    actual_intent_summary: str | None,
    expected_intent_summary: str | None = None,
) -> dict[str, Any]:
    expected = {normalize_intent_line(line) for line in expected_lines if str(line).strip()}
    actual = set(extract_intent_lines(actual_intent_summary))
    matched = expected & actual
    missing = sorted(expected - actual)
    unexpected = sorted(actual - expected)
    expected_summary_normalized = normalize_intent_summary_text(expected_intent_summary)
    actual_summary_normalized = normalize_intent_summary_text(actual_intent_summary)
    summary_exact_match = (
        expected_summary_normalized == actual_summary_normalized
        if expected_summary_normalized is not None
        else None
    )
    line_exact_match = not missing and not unexpected
    exact_match = line_exact_match if summary_exact_match is None else line_exact_match and summary_exact_match
    return {
        "intent_exact_match": exact_match,
        "intent_line_exact_match": line_exact_match,
        "intent_summary_exact_match": summary_exact_match,
        "intent_line_recall": len(matched) / len(expected) if expected else None,
        "intent_line_precision": len(matched) / len(actual) if actual else None,
        "missing_intent_lines": missing,
        "unexpected_intent_lines": unexpected,
        "expected_intent_lines": sorted(expected),
        "actual_intent_lines": sorted(actual),
        "expected_intent_summary_normalized": expected_summary_normalized,
        "actual_intent_summary_normalized": actual_summary_normalized,
    }


async def evaluate_intent_cases(
    cases: list[dict[str, Any]],
    options: EvalOptions,
    progress_callback: ProgressCallback | None = None,
) -> dict[str, Any]:
    semaphore = asyncio.Semaphore(max(1, options.concurrency))
    timeout = httpx.Timeout(options.timeout_seconds)

    async with httpx.AsyncClient(timeout=timeout) as client:
        async def evaluate_one(index: int, case: dict[str, Any]) -> tuple[int, dict[str, Any]]:
            async with semaphore:
                actual, elapsed_ms, error = await _call_parse_api(client, options.base_url, str(case.get("query") or ""))
            expected_lines = expected_intent_lines_from_case(case)
            expected_summary = expected_intent_summary_from_case(case)
            comparison = compare_intent_lines(expected_lines, actual.get("intent_summary"), expected_summary)
            return index, {
                "id": case.get("id"),
                "query": case.get("query"),
                "tags": case.get("tags") or [],
                "expected": case.get("expected") if "expected" in case else None,
                "expected_intent_lines": comparison["expected_intent_lines"],
                "expected_intent_summary": expected_summary,
                "actual_intent_summary": actual.get("intent_summary"),
                "actual_intent_lines": comparison["actual_intent_lines"],
                "actual": actual,
                "intent_comparison": comparison,
                "elapsed_ms": elapsed_ms,
                "error": error,
                "graded": bool(expected_lines),
            }

        tasks = [asyncio.create_task(evaluate_one(index, case)) for index, case in enumerate(cases)]
        case_results: list[dict[str, Any] | None] = [None] * len(tasks)
        completed = 0
        errors = 0
        total_elapsed_ms = 0.0
        for task in asyncio.as_completed(tasks):
            index, result = await task
            case_results[index] = result
            completed += 1
            if result["error"]:
                errors += 1
            total_elapsed_ms += float(result["elapsed_ms"] or 0)
            if progress_callback:
                progress_callback(
                    {
                        "completed": completed,
                        "total": len(tasks),
                        "errors": errors,
                        "avg_latency_ms": total_elapsed_ms / completed if completed else 0,
                        "case_id": result.get("id"),
                        "query": result.get("query"),
                        "elapsed_ms": result.get("elapsed_ms"),
                        "matched_level": result.get("actual", {}).get("matched_level"),
                        "error": result.get("error"),
                    }
                )

    finalized_case_results = [item for item in case_results if item is not None]
    total = len(finalized_case_results)
    graded_cases = [item for item in finalized_case_results if item["graded"]]
    exact_matches = sum(
        1 for item in graded_cases
        if item["intent_comparison"]["intent_exact_match"] and not item["error"]
    )
    line_exact_matches = sum(
        1 for item in graded_cases
        if item["intent_comparison"]["intent_line_exact_match"] and not item["error"]
    )
    summary_exact_cases = [
        item for item in graded_cases
        if item["intent_comparison"]["intent_summary_exact_match"] is not None
    ]
    summary_exact_matches = sum(
        1 for item in summary_exact_cases
        if item["intent_comparison"]["intent_summary_exact_match"] and not item["error"]
    )
    recall_values = [
        item["intent_comparison"]["intent_line_recall"]
        for item in graded_cases
        if item["intent_comparison"]["intent_line_recall"] is not None
    ]
    precision_values = [
        item["intent_comparison"]["intent_line_precision"]
        for item in graded_cases
        if item["intent_comparison"]["intent_line_precision"] is not None
    ]
    latencies = [item["elapsed_ms"] for item in finalized_case_results]
    level_distribution: dict[str, int] = {}
    for item in finalized_case_results:
        level = str(item["actual"].get("matched_level") or "unknown")
        level_distribution[level] = level_distribution.get(level, 0) + 1

    summary = {
        "total": total,
        "graded_total": len(graded_cases),
        "ungraded_total": total - len(graded_cases),
        "graded_coverage_rate": len(graded_cases) / total if total else 0,
        "api_success_rate": (total - sum(1 for item in finalized_case_results if item["error"])) / total if total else 0,
        "intent_exact_match_rate": exact_matches / len(graded_cases) if graded_cases else None,
        "intent_line_exact_match_rate": line_exact_matches / len(graded_cases) if graded_cases else None,
        "intent_summary_exact_match_rate": summary_exact_matches / len(summary_exact_cases) if summary_exact_cases else None,
        "intent_line_recall": sum(recall_values) / len(recall_values) if recall_values else None,
        "intent_line_precision": sum(precision_values) / len(precision_values) if precision_values else None,
        "avg_latency_ms": sum(latencies) / len(latencies) if latencies else 0,
        "p95_latency_ms": sorted(latencies)[int(len(latencies) * 0.95) - 1] if latencies else 0,
        "level_distribution": level_distribution,
        "error_count": sum(1 for item in finalized_case_results if item["error"]),
    }
    failed_cases = [
        item for item in graded_cases
        if item["error"] or not item["intent_comparison"]["intent_exact_match"]
    ]
    return {
        "summary": summary,
        "cases": finalized_case_results,
        "failed_cases": failed_cases,
    }


def _normalize_query_key(query: Any) -> str:
    return re.sub(r"\s+", "", str(query or "").strip().replace("。", ""))


def _example_expected_from_output(output: Any) -> dict[str, Any] | None:
    if not isinstance(output, dict):
        return None
    if "conditions" in output:
        return {
            "query_logic": output.get("query_logic") or "AND",
            "conditions": output.get("conditions") or [],
        }
    if output.get("field") and output.get("operator"):
        return {
            "query_logic": "AND",
            "conditions": [output],
        }
    return None


def _load_field_definition_example_index(path: str | Path | None = None) -> dict[str, dict[str, Any]]:
    if path is None:
        from src.main.python.config.settings import settings

        path = settings.FIELD_DEFINITIONS_PATH

    definitions_path = Path(path)
    with definitions_path.open("r", encoding="utf-8") as file:
        raw = yaml.safe_load(file) or {}

    index: dict[str, dict[str, Any]] = {}
    for intent in raw.get("intents") or []:
        if not isinstance(intent, dict):
            continue
        intent_meta = {
            "intent_id": intent.get("id"),
            "field": intent.get("field"),
            "operator": intent.get("operator"),
            "config_path": str(definitions_path),
        }

        for example in intent.get("examples") or []:
            if not isinstance(example, dict) or not example.get("query"):
                continue
            expected = _example_expected_from_output(example.get("output"))
            if not expected:
                continue
            index[_normalize_query_key(example["query"])] = {
                "expected": expected,
                "label_status": "candidate",
                "label_meta": {
                    **intent_meta,
                    "example_type": "positive_example",
                },
            }

        for example in intent.get("negative_examples") or []:
            if not isinstance(example, dict) or not example.get("query"):
                continue
            index[_normalize_query_key(example["query"])] = {
                "expected": {"query_logic": "AND", "conditions": []},
                "label_status": "candidate",
                "label_meta": {
                    **intent_meta,
                    "example_type": "negative_example",
                    "reason": example.get("reason"),
                },
            }

    return index


def _load_yaml_mapping(path: str | Path | None) -> dict[str, Any]:
    if not path:
        return {}
    candidate = Path(path)
    if not candidate.exists():
        return {}
    with candidate.open("r", encoding="utf-8") as file:
        loaded = yaml.safe_load(file) or {}
    return loaded if isinstance(loaded, dict) else {}


def _load_field_definitions(path: str | Path | None = None) -> list[dict[str, Any]]:
    if path is None:
        from src.main.python.config.settings import settings

        path = settings.FIELD_DEFINITIONS_PATH
    raw = _load_yaml_mapping(path)
    return [item for item in (raw.get("intents") or []) if isinstance(item, dict)]


def _load_enum_values(path: str | Path | None = None) -> dict[str, list[str]]:
    if path is None:
        from src.main.python.config.settings import settings

        path = Path(settings.ENUMS_DIR_PATH) / "field_enums_args.yaml"
    raw = _load_yaml_mapping(path)
    values_by_field: dict[str, list[str]] = {}
    for field, spec in raw.items():
        if isinstance(spec, dict):
            values = spec.get("values") or []
        elif isinstance(spec, list):
            values = spec
        else:
            values = []
        values_by_field[str(field)] = [str(value) for value in values if str(value)]
    return values_by_field


def _load_ordered_enum_values(path: str | Path | None = None) -> dict[str, list[str]]:
    if path is None:
        from src.main.python.config.settings import settings

        path = Path(settings.ENUMS_DIR_PATH) / "field_enums_args.yaml"
    raw = _load_yaml_mapping(path)
    values_by_field: dict[str, list[str]] = {}
    for field, spec in raw.items():
        if not isinstance(spec, dict) or not spec.get("ordered"):
            continue
        values = spec.get("values") or []
        values_by_field[str(field)] = [str(value) for value in values if str(value)]
    return values_by_field


def _load_value_mappings(path: str | Path | None = None) -> dict[str, dict[str, str]]:
    if path is None:
        from src.main.python.config.settings import settings

        path = settings.VALUE_MAPPINGS_PATH
    raw = _load_yaml_mapping(path)
    result: dict[str, dict[str, str]] = {}
    for field, mappings in raw.items():
        if isinstance(mappings, dict):
            result[str(field)] = {str(alias): str(value) for alias, value in mappings.items()}
    return result


def _expected_signature(expected: dict[str, Any]) -> str:
    return json.dumps(
        {
            "query_logic": expected.get("query_logic") or "AND",
            "conditions": [_normalize_condition(condition) for condition in (expected.get("conditions") or [])],
        },
        ensure_ascii=False,
        sort_keys=True,
    )


def _candidate(
    *,
    source: str,
    confidence: float,
    expected: dict[str, Any],
    evidence: list[str],
) -> dict[str, Any]:
    return {
        "source": source,
        "confidence": confidence,
        "expected": expected,
        "robot_text": _robot_text_from_expected(expected),
        "evidence": evidence,
    }


def _build_example_candidate(case: dict[str, Any], example_index: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    match = example_index.get(_normalize_query_key(case.get("query")))
    if not match:
        return None
    meta = match.get("label_meta") or {}
    confidence = 0.98 if meta.get("example_type") == "positive_example" else 0.95
    return _candidate(
        source="config_example",
        confidence=confidence,
        expected=match["expected"],
        evidence=[f"{meta.get('config_path')}#{meta.get('intent_id')}.{meta.get('example_type')}"],
    )


def _build_value_mapping_candidates(
    query: str,
    intents: list[dict[str, Any]],
    value_mappings: dict[str, dict[str, str]],
) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    normalized_query = _normalize_query_key(query)
    for intent in intents:
        field = str(intent.get("field") or "")
        operator = str(intent.get("operator") or "")
        if not field or not operator:
            continue
        mappings = value_mappings.get(field) or {}
        for alias, standard_value in mappings.items():
            if _normalize_query_key(alias) not in normalized_query:
                continue
            expected = {
                "query_logic": "AND",
                "conditions": [{"field": field, "operator": operator, "value": standard_value}],
            }
            candidates.append(
                _candidate(
                    source="value_mapping",
                    confidence=0.72,
                    expected=expected,
                    evidence=[f"value_mappings.{field}.{alias}->{standard_value}", f"intent:{intent.get('id')}"],
                )
            )
    return candidates


def _build_enum_candidates(
    query: str,
    intents: list[dict[str, Any]],
    enum_values: dict[str, list[str]],
) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    normalized_query = _normalize_query_key(query)
    for intent in intents:
        field = str(intent.get("field") or "")
        operator = str(intent.get("operator") or "")
        value_type = str(intent.get("value_type") or "")
        enum_ref = str(intent.get("enum_ref") or intent.get("enum") or field)
        if not field or not operator or value_type not in {"enum", "infer"}:
            continue
        for enum_value in enum_values.get(enum_ref, enum_values.get(field, [])):
            if _normalize_query_key(enum_value) not in normalized_query:
                continue
            expected = {
                "query_logic": "AND",
                "conditions": [{"field": field, "operator": operator, "value": enum_value}],
            }
            candidates.append(
                _candidate(
                    source="enum_overlap",
                    confidence=0.68,
                    expected=expected,
                    evidence=[f"enum:{enum_ref}.{enum_value}", f"intent:{intent.get('id')}"],
                )
            )
    return candidates


def _dedupe_candidates(candidates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    best_by_signature: dict[str, dict[str, Any]] = {}
    for candidate in candidates:
        signature = _expected_signature(candidate["expected"])
        existing = best_by_signature.get(signature)
        if existing is None or float(candidate.get("confidence") or 0) > float(existing.get("confidence") or 0):
            best_by_signature[signature] = candidate
            continue
        if existing is not None:
            existing.setdefault("evidence", []).extend(candidate.get("evidence") or [])
    return sorted(best_by_signature.values(), key=lambda item: float(item.get("confidence") or 0), reverse=True)


def _agreement_status(candidates: list[dict[str, Any]]) -> tuple[str, dict[str, Any] | None]:
    if not candidates:
        return "manual_required", None
    signatures = {_expected_signature(candidate["expected"]) for candidate in candidates}
    top = candidates[0]
    if len(signatures) == 1 and len(candidates) >= 2:
        return "auto_approved", top
    if top.get("source") == "config_example" and float(top.get("confidence") or 0) >= 0.95:
        return "auto_approved", top
    if len(signatures) == 1:
        return "single_candidate_review", top
    return "conflict_review", None


def build_label_candidates(
    cases: list[dict[str, Any]],
    *,
    field_definitions_path: str | Path | None = None,
    field_enums_path: str | Path | None = None,
    value_mappings_path: str | Path | None = None,
) -> list[dict[str, Any]]:
    example_index = _load_field_definition_example_index(field_definitions_path)
    intents = _load_field_definitions(field_definitions_path)
    enum_values = _load_enum_values(field_enums_path)
    value_mappings = _load_value_mappings(value_mappings_path)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows: list[dict[str, Any]] = []

    for case in cases:
        raw_candidates: list[dict[str, Any]] = []
        example_candidate = _build_example_candidate(case, example_index)
        if example_candidate:
            raw_candidates.append(example_candidate)
        raw_candidates.extend(_build_value_mapping_candidates(str(case.get("query") or ""), intents, value_mappings))
        raw_candidates.extend(_build_enum_candidates(str(case.get("query") or ""), intents, enum_values))
        candidates = _dedupe_candidates(raw_candidates)
        agreement, final_candidate = _agreement_status(candidates)
        final_expected = final_candidate.get("expected") if final_candidate else None
        rows.append(
            {
                "id": case.get("id"),
                "query": case.get("query"),
                "tags": case.get("tags") or [],
                "candidates": candidates,
                "candidate_count": len(candidates),
                "agreement_status": agreement,
                "review_status": "auto_approved" if agreement == "auto_approved" else "pending",
                "final_expected": final_expected,
                "final_robot_text": _robot_text_from_expected(final_expected),
                "label_meta": {
                    "source": "multi_candidate_labeler",
                    "generated_at": generated_at,
                    "needs_review": agreement != "auto_approved",
                },
            }
        )
    return rows


def write_label_candidates_jsonl(rows: list[dict[str, Any]], output_path: str | Path) -> Path:
    path = Path(output_path).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        for row in rows:
            file.write(json.dumps(row, ensure_ascii=False, sort_keys=True))
            file.write("\n")
    return path


def write_label_candidates_excel(rows: list[dict[str, Any]], output_path: str | Path) -> Path:
    path = Path(output_path).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "label_candidates"
    headers = [
        "id",
        "query",
        "agreement_status",
        "review_status",
        "candidate_count",
        "final_robot_text",
        "final_expected",
        "candidate_1_source",
        "candidate_1_confidence",
        "candidate_1_expected",
        "candidate_1_evidence",
        "candidate_2_source",
        "candidate_2_confidence",
        "candidate_2_expected",
        "candidate_2_evidence",
        "all_candidates",
        "review_comment",
    ]
    _write_excel_header(sheet, headers)
    for row_index, row in enumerate(rows, start=2):
        candidates = row.get("candidates") or []
        first = candidates[0] if len(candidates) >= 1 else {}
        second = candidates[1] if len(candidates) >= 2 else {}
        values = [
            row.get("id"),
            row.get("query"),
            row.get("agreement_status"),
            row.get("review_status"),
            row.get("candidate_count"),
            row.get("final_robot_text"),
            row.get("final_expected"),
            first.get("source"),
            first.get("confidence"),
            first.get("expected"),
            first.get("evidence"),
            second.get("source"),
            second.get("confidence"),
            second.get("expected"),
            second.get("evidence"),
            candidates,
            "",
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column_index, _excel_scalar(value))
            if headers[column_index - 1] in {
                "query",
                "final_robot_text",
                "final_expected",
                "candidate_1_expected",
                "candidate_1_evidence",
                "candidate_2_expected",
                "candidate_2_evidence",
                "all_candidates",
                "review_comment",
            }:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    _autosize_excel_columns(sheet)
    workbook.save(path)
    return path


def build_intent_label_candidates(
    cases: list[dict[str, Any]],
    static_rows: list[dict[str, Any]],
    eval_result: dict[str, Any],
) -> list[dict[str, Any]]:
    static_by_id = {str(row.get("id")): row for row in static_rows}
    actual_by_id = {str(row.get("id")): row for row in (eval_result.get("cases") or [])}
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows: list[dict[str, Any]] = []

    for case in cases:
        case_id = str(case.get("id"))
        static_row = static_by_id.get(case_id) or {}
        actual_row = actual_by_id.get(case_id) or {}
        candidates: list[dict[str, Any]] = []

        for candidate in static_row.get("candidates") or []:
            expected = candidate.get("expected")
            intent_summary = candidate.get("robot_text") or _robot_text_from_expected(expected)
            candidates.append(
                {
                    "source": candidate.get("source"),
                    "confidence": candidate.get("confidence"),
                    "candidate_type": "config_candidate",
                    "expected": expected,
                    "intent_summary": intent_summary,
                    "intent_lines": extract_intent_lines(intent_summary),
                    "evidence": candidate.get("evidence") or [],
                }
            )

        actual = actual_row.get("actual") or {}
        actual_conditions = actual.get("conditions") or []
        if actual_row and not actual_row.get("error"):
            parser_expected = {
                "query_logic": actual.get("query_logic") or "AND",
                "conditions": actual_conditions,
            }
            parser_summary = actual.get("intent_summary") or _robot_text_from_expected(parser_expected)
            candidates.append(
                {
                    "source": "current_parse_api",
                    "confidence": 0.72 if actual_conditions else 0.35,
                    "candidate_type": "parser_observation",
                    "expected": parser_expected,
                    "intent_summary": parser_summary,
                    "intent_lines": extract_intent_lines(parser_summary),
                    "evidence": [
                        f"matched_level={actual.get('matched_level') or 'unknown'}",
                        f"condition_count={len(actual_conditions)}",
                    ],
                }
            )

        candidates = _dedupe_intent_candidates(candidates)
        static_agreement = str(static_row.get("agreement_status") or "")
        final_candidate = _choose_intent_candidate(candidates, prefer_static=static_agreement == "auto_approved")
        review_status = _intent_review_status(static_agreement, final_candidate, actual_row)

        rows.append(
            {
                "id": case.get("id"),
                "query": case.get("query"),
                "tags": case.get("tags") or [],
                "review_status": review_status,
                "static_agreement_status": static_agreement or "manual_required",
                "candidate_count": len(candidates),
                "final_expected": final_candidate.get("expected") if final_candidate else None,
                "candidate_intent_summary": final_candidate.get("intent_summary") if final_candidate else None,
                "candidate_intent_lines": final_candidate.get("intent_lines") if final_candidate else [],
                "final_expected_intent_lines": final_candidate.get("intent_lines") if review_status == "auto_approved" and final_candidate else [],
                "candidates": candidates,
                "actual": actual,
                "error": actual_row.get("error"),
                "label_meta": {
                    "source": "intent_label_candidates",
                    "generated_at": generated_at,
                    "needs_review": review_status != "auto_approved",
                },
            }
        )
    return rows


def _dedupe_intent_candidates(candidates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    best_by_signature: dict[str, dict[str, Any]] = {}
    for candidate in candidates:
        signature = json.dumps(
            {
                "summary": candidate.get("intent_summary"),
                "expected": candidate.get("expected"),
            },
            ensure_ascii=False,
            sort_keys=True,
        )
        existing = best_by_signature.get(signature)
        if not existing or float(candidate.get("confidence") or 0) > float(existing.get("confidence") or 0):
            best_by_signature[signature] = candidate
    return sorted(
        best_by_signature.values(),
        key=lambda item: float(item.get("confidence") or 0),
        reverse=True,
    )


def _choose_intent_candidate(candidates: list[dict[str, Any]], *, prefer_static: bool) -> dict[str, Any] | None:
    if not candidates:
        return None
    if prefer_static:
        for candidate in candidates:
            if candidate.get("candidate_type") == "config_candidate":
                return candidate
    return candidates[0]


def _intent_review_status(
    static_agreement: str,
    final_candidate: dict[str, Any] | None,
    actual_row: dict[str, Any],
) -> str:
    if static_agreement == "auto_approved" and final_candidate:
        return "auto_approved"
    if not final_candidate:
        return "manual_required"
    if actual_row.get("error"):
        return "api_error"
    if static_agreement in {"single_candidate_review", "conflict_review"}:
        return static_agreement
    if final_candidate.get("candidate_type") == "parser_observation":
        return "parser_review"
    return "pending"


def write_intent_label_candidates_jsonl(rows: list[dict[str, Any]], output_path: str | Path) -> Path:
    path = Path(output_path).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        for row in rows:
            file.write(json.dumps(row, ensure_ascii=False, sort_keys=True))
            file.write("\n")
    return path


def write_intent_label_candidates_excel(rows: list[dict[str, Any]], output_path: str | Path) -> Path:
    path = Path(output_path).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "intent_label_candidates"
    headers = [
        "id",
        "query",
        "review_status",
        "static_agreement_status",
        "candidate_count",
        "candidate_intent_summary",
        "candidate_intent_lines",
        "final_expected_intent_lines",
        "final_expected",
        "candidate_1_source",
        "candidate_1_confidence",
        "candidate_1_intent_summary",
        "candidate_1_expected",
        "candidate_1_evidence",
        "candidate_2_source",
        "candidate_2_confidence",
        "candidate_2_intent_summary",
        "candidate_2_expected",
        "candidate_2_evidence",
        "all_candidates",
        "error",
        "review_comment",
    ]
    _write_excel_header(sheet, headers)
    for row_index, row in enumerate(rows, start=2):
        candidates = row.get("candidates") or []
        first = candidates[0] if len(candidates) >= 1 else {}
        second = candidates[1] if len(candidates) >= 2 else {}
        values = [
            row.get("id"),
            row.get("query"),
            row.get("review_status"),
            row.get("static_agreement_status"),
            row.get("candidate_count"),
            row.get("candidate_intent_summary"),
            row.get("candidate_intent_lines"),
            row.get("final_expected_intent_lines"),
            row.get("final_expected"),
            first.get("source"),
            first.get("confidence"),
            first.get("intent_summary"),
            first.get("expected"),
            first.get("evidence"),
            second.get("source"),
            second.get("confidence"),
            second.get("intent_summary"),
            second.get("expected"),
            second.get("evidence"),
            candidates,
            row.get("error"),
            "",
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column_index, _excel_scalar(value))
            if headers[column_index - 1] in {
                "query",
                "candidate_intent_summary",
                "candidate_intent_lines",
                "final_expected_intent_lines",
                "final_expected",
                "candidate_1_intent_summary",
                "candidate_1_expected",
                "candidate_1_evidence",
                "candidate_2_intent_summary",
                "candidate_2_expected",
                "candidate_2_evidence",
                "all_candidates",
                "review_comment",
            }:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    _autosize_excel_columns(sheet, max_width=100)
    workbook.save(path)
    return path


def build_intent_gold_from_batch_excel(
    input_path: str | Path,
    *,
    sheet_name: str = "cases",
    accept_unreviewed: bool = False,
) -> dict[str, Any]:
    path = Path(input_path).resolve()
    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"{path} missing sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    header_index = {header: index for index, header in enumerate(headers) if header}
    rows: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        record = {header: values[index] for header, index in header_index.items() if index < len(values)}
        if not any(value not in (None, "") for value in record.values()):
            continue

        case_id = _excel_first(record, "id", "ID", "case_id")
        query = _excel_first(record, "query", "question", "问题")
        if not query:
            skipped.append({"row": row_number, "id": case_id, "reason": "missing_query"})
            continue

        skip_marker = str(_excel_first(record, "skip", "exclude", "是否跳过") or "").strip().lower()
        if skip_marker in {"1", "true", "yes", "y", "skip", "跳过", "是"}:
            skipped.append({"row": row_number, "id": case_id, "query": query, "reason": "skip_marker"})
            continue

        review_status = str(_excel_first(record, "review_status", "人工结论", "结论") or "").strip()
        issue_type = str(_excel_first(record, "issue_type", "error_type", "错误类型", "问题类型") or "").strip()
        final_summary = _excel_first(
            record,
            "final_intent_summary",
            "expected_intent",
            "standard_intent_summary",
            "标准意图",
            "修正意图",
        )
        final_lines = _excel_first(
            record,
            "final_intent_lines",
            "expected_intent_lines",
            "standard_intent_lines",
            "标准意图行",
            "修正意图行",
        )
        actual_summary = _excel_first(record, "intent_summary", "actual_intent_summary", "candidate_intent_summary")

        expected_intent = str(final_summary).strip() if final_summary not in (None, "") else None
        expected_lines = _parse_excel_intent_lines(final_lines)
        accepted = _is_accepted_review_status(review_status)
        if not expected_intent and not expected_lines and (accepted or (accept_unreviewed and not review_status)):
            expected_intent = str(actual_summary).strip() if actual_summary not in (None, "") else None

        if not expected_intent and not expected_lines:
            skipped.append(
                {
                    "row": row_number,
                    "id": case_id,
                    "query": query,
                    "review_status": review_status,
                    "issue_type": issue_type,
                    "reason": "missing_final_intent",
                }
            )
            continue

        output: dict[str, Any] = {
            "id": str(case_id or f"excel_{len(rows) + 1:04d}"),
            "query": str(query),
            "tags": _parse_tags(_excel_first(record, "tags", "标签") or []) + ["intent_gold_from_excel"],
            "label_meta": {
                "source": "batch_eval_result_excel",
                "source_file": str(path),
                "source_sheet": sheet_name,
                "source_row": row_number,
                "review_status": review_status,
                "issue_type": issue_type,
            },
        }
        if expected_intent:
            output["expected_intent"] = expected_intent
            output["expected_intent_lines"] = extract_intent_lines(expected_intent)
        else:
            output["expected_intent_lines"] = expected_lines
        rows.append(output)

    return {
        "rows": rows,
        "skipped": skipped,
        "summary": {
            "input": str(path),
            "sheet": sheet_name,
            "total_data_rows": max(sheet.max_row - 1, 0),
            "generated": len(rows),
            "skipped": len(skipped),
            "accept_unreviewed": accept_unreviewed,
        },
    }


def write_intent_gold_jsonl(rows: list[dict[str, Any]], output_path: str | Path) -> Path:
    path = Path(output_path).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        for row in rows:
            file.write(json.dumps(row, ensure_ascii=False, sort_keys=True))
            file.write("\n")
    return path


def prepare_intent_review_workbook(
    input_path: str | Path,
    output_path: str | Path,
    *,
    sheet_name: str = "cases",
) -> dict[str, Any]:
    source = Path(input_path).resolve()
    output = Path(output_path).resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(source)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"{source} missing sheet: {sheet_name}")
    sheet = workbook[sheet_name]

    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    header_index = {header: index + 1 for index, header in enumerate(headers) if header}
    review_headers = [
        "review_status",
        "issue_type",
        "risk_level",
        "auto_suggestion",
        "possible_intent_summary",
        "final_intent_summary",
        "final_intent_lines",
        "skip",
        "review_comment",
    ]
    for header in review_headers:
        if header not in header_index:
            column = sheet.max_column + 1
            sheet.cell(1, column, header)
            header_index[header] = column

    status_counts: dict[str, int] = {}
    for row_number in range(2, sheet.max_row + 1):
        record = {
            header: sheet.cell(row_number, column).value
            for header, column in header_index.items()
        }
        if not any(value not in (None, "") for value in record.values()):
            continue
        analysis = _analyze_intent_review_row(record)
        status_counts[analysis["review_status"]] = status_counts.get(analysis["review_status"], 0) + 1

        for key in ("review_status", "issue_type", "risk_level", "auto_suggestion"):
            cell = sheet.cell(row_number, header_index[key])
            if cell.value in (None, ""):
                cell.value = analysis[key]
        if analysis.get("prefill_final_intent_summary") and sheet.cell(row_number, header_index["final_intent_summary"]).value in (None, ""):
            sheet.cell(row_number, header_index["final_intent_summary"]).value = analysis["prefill_final_intent_summary"]
        if analysis.get("possible_intent_summary") and sheet.cell(row_number, header_index["possible_intent_summary"]).value in (None, ""):
            sheet.cell(row_number, header_index["possible_intent_summary"]).value = analysis["possible_intent_summary"]

    _format_intent_review_sheet(sheet, header_index, review_headers)
    workbook.save(output)
    return {
        "input": str(source),
        "output": str(output),
        "sheet": sheet_name,
        "rows": max(sheet.max_row - 1, 0),
        "review_status_counts": status_counts,
    }


def build_skill_eval_from_batch_excel(
    input_path: str | Path,
    *,
    sheet_name: str = "cases",
    skill_path: str | Path | None = None,
) -> dict[str, Any]:
    path = Path(input_path).resolve()
    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"{path} missing sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    header_index = {header: index for index, header in enumerate(headers) if header}
    skill_text = ""
    if skill_path:
        skill_text = Path(skill_path).read_text(encoding="utf-8")

    cases: list[dict[str, Any]] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        record = {header: values[index] for header, index in header_index.items() if index < len(values)}
        if not any(value not in (None, "") for value in record.values()):
            continue
        cases.append(_skill_eval_case_from_record(record, source_row=row_number))

    summary = _summarize_skill_eval_cases(cases)
    return {
        "summary": {
            **summary,
            "input": str(path),
            "sheet": sheet_name,
            "skill_path": str(Path(skill_path).resolve()) if skill_path else None,
            "skill_loaded": bool(skill_text),
        },
        "cases": cases,
        "candidate_gold": _candidate_gold_from_skill_eval_cases(cases),
    }


async def build_skill_eval_from_batch_excel_with_llm_judge(
    input_path: str | Path,
    *,
    sheet_name: str = "cases",
    skill_path: str | Path,
    judge_options: LLMJudgeOptions,
) -> dict[str, Any]:
    path = Path(input_path)
    workbook = load_workbook(path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"sheet not found: {sheet_name}")
    sheet = workbook[sheet_name]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    header_index = {header: index for index, header in enumerate(headers) if header}
    skill_text = Path(skill_path).read_text(encoding="utf-8")

    records: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        record = {header: values[index] for header, index in header_index.items() if index < len(values)}
        if not any(value not in (None, "") for value in record.values()):
            continue
        records.append((row_number, record))

    semaphore = asyncio.Semaphore(max(1, judge_options.concurrency))

    async def judge_one(source_row: int, record: dict[str, Any]) -> dict[str, Any]:
        async with semaphore:
            return await _llm_judge_case_from_record(
                record,
                source_row=source_row,
                skill_text=skill_text,
                judge_options=judge_options,
            )

    cases = await asyncio.gather(*(judge_one(row_number, record) for row_number, record in records))
    summary = _summarize_skill_eval_cases(cases)
    return {
        "summary": {
            **summary,
            "input": str(path),
            "sheet": sheet_name,
            "skill_path": str(Path(skill_path).resolve()),
            "skill_loaded": True,
            "judge": "llm",
            "judge_model": judge_options.model,
        },
        "cases": cases,
        "candidate_gold": _candidate_gold_from_skill_eval_cases(cases),
    }


def _skill_eval_case_from_record(record: dict[str, Any], *, source_row: int) -> dict[str, Any]:
    query = str(_excel_first(record, "query", "question", "问题") or "")
    actual_conditions = _parse_excel_json_value(_excel_first(record, "actual_conditions", "conditions")) or []
    if not isinstance(actual_conditions, list):
        actual_conditions = []
    actual_intent_summary = str(_excel_first(record, "intent_summary", "actual_intent_summary", "candidate_intent_summary") or "")
    judgment = _judge_case_with_skill_rubric(record, query, actual_conditions, actual_intent_summary)
    return {
        "id": _excel_first(record, "id", "ID", "case_id"),
        "query": query,
        "verdict": judgment["verdict"],
        "confidence": judgment["confidence"],
        "severity": judgment["severity"],
        "error_types": judgment["error_types"],
        "raw_issue_types": judgment["raw_issue_types"],
        "reason": judgment["reason"],
        "expected_intent_summary": judgment["expected_intent_summary"],
        "expected_intent_lines": extract_intent_lines(judgment["expected_intent_summary"]),
        "actual_intent_summary": actual_intent_summary,
        "actual_conditions": actual_conditions,
        "matched_level": _excel_first(record, "matched_level", "level"),
        "source_row": source_row,
        "deterministic_findings": judgment["skill_findings"],
    }


async def _llm_judge_case_from_record(
    record: dict[str, Any],
    *,
    source_row: int,
    skill_text: str,
    judge_options: LLMJudgeOptions,
) -> dict[str, Any]:
    try:
        from openai import AsyncOpenAI
    except ImportError as exc:
        raise RuntimeError("openai package is required for --llm-judge") from exc

    query = str(_excel_first(record, "query", "question", "问题") or "")
    actual_conditions = _parse_excel_json_value(_excel_first(record, "actual_conditions", "conditions")) or []
    if not isinstance(actual_conditions, list):
        actual_conditions = []
    actual_intent_summary = str(_excel_first(record, "intent_summary", "actual_intent_summary", "candidate_intent_summary") or "")
    matched_level = _excel_first(record, "matched_level", "level")
    local_judgment = _judge_case_with_skill_rubric(record, query, actual_conditions, actual_intent_summary)

    client_kwargs: dict[str, Any] = {"api_key": judge_options.api_key}
    if judge_options.base_url:
        client_kwargs["base_url"] = judge_options.base_url
    client = AsyncOpenAI(**client_kwargs)
    prompt = _build_llm_judge_prompt(
        skill_text=skill_text,
        query=query,
        actual_conditions=actual_conditions,
        actual_intent_summary=actual_intent_summary,
        matched_level=matched_level,
        local_judgment=local_judgment,
    )
    try:
        content = await _call_llm_judge_with_retries(client, prompt, judge_options)
    except Exception as exc:  # noqa: BLE001 - one failed judge request should not abort the batch
        return {
            "id": _excel_first(record, "id", "ID", "case_id"),
            "query": query,
            "verdict": "uncertain",
            "confidence": 0.0,
            "severity": "medium",
            "error_types": [],
            "raw_issue_types": [],
            "reason": f"LLM Judge 调用失败，需复核：{type(exc).__name__}: {exc}",
            "expected_intent_summary": actual_intent_summary,
            "expected_intent_lines": extract_intent_lines(actual_intent_summary),
            "actual_intent_summary": actual_intent_summary,
            "actual_conditions": actual_conditions,
            "matched_level": matched_level,
            "source_row": source_row,
            "deterministic_findings": [
                {
                    "error_type": finding.get("error_type"),
                    "raw_issue_type": finding.get("raw_issue_type"),
                    "reason": finding.get("reason"),
                }
                for finding in (local_judgment.get("skill_findings") or [])
            ],
        }
    judgment = _normalize_llm_judge_response(content, actual_intent_summary)
    return {
        "id": _excel_first(record, "id", "ID", "case_id"),
        "query": query,
        "verdict": judgment["verdict"],
        "confidence": judgment["confidence"],
        "severity": judgment["severity"],
        "error_types": judgment["error_types"],
        "raw_issue_types": [],
        "reason": judgment["reason"],
        "expected_intent_summary": judgment["expected_intent_summary"],
        "expected_intent_lines": extract_intent_lines(judgment["expected_intent_summary"]),
        "actual_intent_summary": actual_intent_summary,
        "actual_conditions": actual_conditions,
        "matched_level": matched_level,
        "source_row": source_row,
        "deterministic_findings": [
            {
                "error_type": finding.get("error_type"),
                "raw_issue_type": finding.get("raw_issue_type"),
                "reason": finding.get("reason"),
            }
            for finding in (local_judgment.get("skill_findings") or [])
        ],
    }


async def _call_llm_judge_with_retries(client: Any, prompt: str, judge_options: LLMJudgeOptions) -> str:
    last_error: Exception | None = None
    attempts = max(1, judge_options.max_retries + 1)
    for attempt in range(attempts):
        try:
            response = await client.chat.completions.create(
                model=judge_options.model,
                messages=[
                    {
                        "role": "system",
                        "content": "你是客户搜索意图评估专家。必须严格按用户提供的 SKILL.md 评估，并只输出 JSON。",
                    },
                    {"role": "user", "content": prompt},
                ],
                temperature=0,
                timeout=judge_options.timeout_seconds,
                response_format={"type": "json_object"},
            )
            return response.choices[0].message.content or "{}"
        except Exception as exc:  # noqa: BLE001 - retry OpenAI-compatible transient failures
            last_error = exc
            if attempt >= attempts - 1:
                break
            await asyncio.sleep(min(2 ** attempt, 8))
    assert last_error is not None
    raise last_error


def _build_llm_judge_prompt(
    *,
    skill_text: str,
    query: str,
    actual_conditions: list[Any],
    actual_intent_summary: str,
    matched_level: Any,
    local_judgment: dict[str, Any],
) -> str:
    return json.dumps(
        {
            "task": "按 SKILL.md 判断客户搜索解析结果是否正确。deterministic_findings 只是参考，最终以 SKILL.md 和语义等价为准。",
            "output_schema": {
                "verdict": "pass | fail | uncertain",
                "confidence": "0到1之间的数字",
                "severity": "low | medium | high",
                "error_types": ["unparsed/missing_condition/extra_condition/duplicate_condition_across_fields/field_wrong/operator_wrong/value_wrong/logic_wrong/unsupported_handling_wrong"],
                "reason": "简洁中文原因",
                "expected_intent_summary": "若 pass 则通常等于 actual_intent_summary；若 fail 则给出纠正后的意图文本，无法确定则为空",
            },
            "skill_md": skill_text,
            "case": {
                "query": query,
                "actual_conditions": actual_conditions,
                "actual_intent_summary": actual_intent_summary,
                "matched_level": matched_level,
            },
            "local_rule_reference": {
                "verdict": local_judgment.get("verdict"),
                "error_types": local_judgment.get("error_types"),
                "reason": local_judgment.get("reason"),
                "expected_intent_summary": local_judgment.get("expected_intent_summary"),
                "deterministic_findings": local_judgment.get("skill_findings"),
            },
        },
        ensure_ascii=False,
    )


def _normalize_llm_judge_response(content: str, actual_intent_summary: str) -> dict[str, Any]:
    try:
        parsed = json.loads(content)
    except json.JSONDecodeError:
        parsed = {}
    verdict = str(parsed.get("verdict") or "uncertain").lower()
    if verdict not in {"pass", "fail", "uncertain"}:
        verdict = "uncertain"
    error_types = parsed.get("error_types") if isinstance(parsed.get("error_types"), list) else []
    normalized_error_types = [str(item) for item in error_types if str(item).strip()]
    expected = str(parsed.get("expected_intent_summary") or "").strip()
    if verdict == "pass" and not expected:
        expected = actual_intent_summary
    try:
        confidence = float(parsed.get("confidence", 0.0))
    except (TypeError, ValueError):
        confidence = 0.0
    confidence = min(1.0, max(0.0, confidence))
    severity = str(parsed.get("severity") or ("low" if verdict == "pass" else "medium")).lower()
    if severity not in {"low", "medium", "high"}:
        severity = "medium"
    return {
        "verdict": verdict,
        "confidence": confidence,
        "severity": severity,
        "error_types": normalized_error_types,
        "reason": str(parsed.get("reason") or "").strip(),
        "expected_intent_summary": expected,
    }


def _judge_case_with_skill_rubric(
    record: dict[str, Any],
    query: str,
    actual_conditions: list[Any],
    actual_intent_summary: str,
) -> dict[str, Any]:
    error = str(_excel_first(record, "error") or "").strip()
    if error:
        return _skill_judgment(
            "fail",
            ["unparsed"],
            ["api_error"],
            "接口调用失败，无法依据 skill 判断解析结果。",
            "",
            "high",
        )

    has_unsupported = _has_unsupported_intent_summary(actual_intent_summary)
    if has_unsupported and not actual_conditions:
        return _skill_judgment(
            "pass",
            [],
            [],
            "依据 SKILL.md：unsupported-only 场景字段提示正确时，无需展示 value/operator，判定通过。",
            actual_intent_summary,
            "low",
        )

    if not actual_intent_summary or ("未识别到明确查询条件" in actual_intent_summary and not has_unsupported):
        return _skill_judgment(
            "fail",
            ["unparsed"],
            ["missing_condition"],
            "依据 SKILL.md：query 有可评估内容但未生成有效意图。",
            "",
            "high",
        )

    fields = {
        str(condition.get("field") or "")
        for condition in actual_conditions
        if isinstance(condition, dict)
    }
    operators = {
        str(condition.get("operator") or "")
        for condition in actual_conditions
        if isinstance(condition, dict)
    }
    invalid_enum_values = _invalid_enum_condition_values(actual_conditions)
    duplicate_same_field_values = _duplicate_same_field_value_operator_conditions(actual_conditions, query)
    special_field_findings = _special_field_usage_findings(actual_conditions, query)
    duplicate_value_groups = _duplicate_condition_values_across_fields(actual_conditions, query)
    polarity_value_findings = _polarity_value_findings(actual_conditions, query)
    raw_issue_types: list[str] = []

    if _expects_negative_operator_from_config(query) and not (
        {"NOT_CONTAINS", "NOT_EXISTS"} & operators
        or _intent_summary_has_negative_operator_text(actual_intent_summary)
    ):
        raw_issue_types.append("suspected_negation_missing")
    if invalid_enum_values:
        raw_issue_types.append("suspected_invalid_enum_value")
    if duplicate_same_field_values:
        raw_issue_types.append("suspected_duplicate_same_field_value_operator")
    if special_field_findings:
        raw_issue_types.append("suspected_special_field_value_incomplete")
    if duplicate_value_groups:
        raw_issue_types.append("suspected_duplicate_value_multi_fields")
    if polarity_value_findings:
        raw_issue_types.append("suspected_polarity_value_wrong")

    if not raw_issue_types:
        return _skill_judgment(
            "pass",
            [],
            [],
            "依据 SKILL.md：actual_intent_summary 与 query 语义等价，字段、operator、value 未命中高置信错误规则。",
            actual_intent_summary,
            "low",
        )

    error_types = sorted({SKILL_ERROR_TYPE_MAP.get(issue, issue) for issue in raw_issue_types})
    possible_intent_summary = _suggest_possible_intent_summary(
        actual_conditions,
        query,
        invalid_enum_values=invalid_enum_values,
        short_value_groups=[],
        duplicate_value_groups=duplicate_value_groups,
        duplicate_intent_concepts=[],
        duplicate_same_field_values=duplicate_same_field_values,
        special_field_findings=special_field_findings,
    )
    reason = _skill_rubric_fail_reason(
        raw_issue_types,
        invalid_enum_values=invalid_enum_values,
        duplicate_same_field_values=duplicate_same_field_values,
        special_field_findings=special_field_findings,
        duplicate_value_groups=duplicate_value_groups,
        polarity_value_findings=polarity_value_findings,
    )
    return _skill_judgment(
        "fail",
        error_types,
        raw_issue_types,
        reason,
        possible_intent_summary,
        "medium",
    )


def _skill_judgment(
    verdict: str,
    error_types: list[str],
    raw_issue_types: list[str],
    reason: str,
    expected_intent_summary: str,
    severity: str,
) -> dict[str, Any]:
    return {
        "verdict": verdict,
        "confidence": 0.84 if verdict == "pass" else 0.82,
        "severity": severity,
        "error_types": error_types,
        "raw_issue_types": raw_issue_types,
        "reason": reason,
        "expected_intent_summary": expected_intent_summary,
        "skill_findings": [
            {
                "error_type": SKILL_ERROR_TYPE_MAP.get(issue, issue),
                "raw_issue_type": issue,
                "reason": _finding_reason(issue),
            }
            for issue in raw_issue_types
        ],
    }


def _skill_rubric_fail_reason(
    raw_issue_types: list[str],
    *,
    invalid_enum_values: list[dict[str, str]],
    duplicate_same_field_values: list[dict[str, Any]],
    special_field_findings: list[dict[str, Any]],
    duplicate_value_groups: list[dict[str, Any]],
    polarity_value_findings: list[dict[str, Any]],
) -> str:
    if "suspected_invalid_enum_value" in raw_issue_types:
        details = "；".join(f"{item['field']}={item['value']}" for item in invalid_enum_values[:3])
        return f"依据 SKILL.md：字段枚举值不合法，{details}。"
    if "suspected_special_field_value_incomplete" in raw_issue_types:
        details = "；".join(
            f"{item['field']}={','.join(item['actual_values'])} -> {','.join(item['expected_values'])}"
            for item in special_field_findings[:3]
        )
        return f"依据 SKILL.md：特殊字段业务语义未完整展开，{details}。"
    if "suspected_duplicate_same_field_value_operator" in raw_issue_types:
        details = "；".join(
            f"{item['field']}={item['value']}({','.join(item['operators'])})"
            for item in duplicate_same_field_values[:3]
        )
        return f"依据 SKILL.md：同一字段同一值存在多个 operator，{details}。"
    if "suspected_duplicate_value_multi_fields" in raw_issue_types:
        details = "；".join(
            f"{item['value']} -> {','.join(item['fields'])}"
            for item in duplicate_value_groups[:3]
        )
        return f"依据 SKILL.md：同一 query 片段或 value 疑似被解析到多个字段，{details}。"
    if "suspected_polarity_value_wrong" in raw_issue_types:
        details = "；".join(
            f"{item['field']}={item['actual_value']}，应为{item['expected_value']}"
            for item in polarity_value_findings[:3]
        )
        return f"依据 SKILL.md：购买/配置极性语义与 value 不一致，{details}。"
    if "suspected_negation_missing" in raw_issue_types:
        return "依据 SKILL.md：query 命中字段定义中的 NOT_CONTAINS/NOT_EXISTS 负向用法，但结果未体现负向 operator 或文案。"
    return "依据 SKILL.md：命中高置信 skill 规则错误。"


def _skill_eval_severity(issue_types: list[str]) -> str:
    high = {"api_error", "missing_condition", "unknown_level"}
    if high & set(issue_types):
        return "high"
    if issue_types:
        return "medium"
    return "low"


def _skill_eval_confidence(verdict: str, severity: str, has_expected: bool) -> float:
    if verdict == "pass":
        return 0.82
    if severity == "high":
        return 0.9
    return 0.84 if has_expected else 0.76


def _skill_eval_reason(verdict: str, analysis: dict[str, Any], error_types: list[str]) -> str:
    if verdict == "pass":
        return "未命中确定性错误规则，当前意图可作为低风险候选，建议抽查。"
    suggestion = str(analysis.get("auto_suggestion") or "").strip()
    if suggestion:
        return suggestion
    return "命中确定性错误规则：" + ",".join(error_types)


def _finding_reason(issue: str) -> str:
    reasons = {
        "api_error": "接口调用失败，无法判断解析结果。",
        "missing_condition": "query 有可评估内容但解析为空或摘要为未识别。",
        "unknown_level": "解析层级未知，需复核。",
        "suspected_mobile_field_wrong": "query 包含手机号但未命中客户手机号字段。",
        "suspected_id_no_field_wrong": "query 包含身份证号但未命中证件号字段。",
        "suspected_policy_no_field_wrong": "query 包含保单号但未命中保单号字段。",
        "suspected_negation_missing": "query 含否定语义但结果未体现否定 operator 或否定文案。",
        "suspected_logic_wrong": "query 含 OR 语义但摘要未体现“或者”。",
        "suspected_operator_wrong": "query 含范围/比较语义但摘要未体现比较或区间。",
        "suspected_duplicate_value_multi_fields": "同一查询值或片段疑似命中多个字段。",
        "suspected_duplicate_same_field_value_operator": "同一字段同一值被多个 operator 重复解析。",
        "suspected_special_field_value_incomplete": "字段特殊业务语义未完整展开为标准枚举值。",
        "suspected_polarity_value_wrong": "购买/配置类极性 value 与 query 语义不一致。",
        "suspected_value_truncated_or_too_short": "value 疑似从更长词中截断。",
        "suspected_duplicate_intent_concept": "同一业务概念在多条意图中重复出现。",
        "suspected_invalid_enum_value": "value 不在该字段枚举值中。",
    }
    return reasons.get(issue, issue)


def _summarize_skill_eval_cases(cases: list[dict[str, Any]]) -> dict[str, Any]:
    total = len(cases)
    verdict_counts: dict[str, int] = {}
    error_counts: dict[str, int] = {}
    severity_counts: dict[str, int] = {}
    for case in cases:
        verdict = str(case.get("verdict") or "unknown")
        verdict_counts[verdict] = verdict_counts.get(verdict, 0) + 1
        severity = str(case.get("severity") or "unknown")
        severity_counts[severity] = severity_counts.get(severity, 0) + 1
        for error_type in case.get("error_types") or []:
            error_counts[error_type] = error_counts.get(error_type, 0) + 1
    return {
        "total": total,
        "pass_rate": verdict_counts.get("pass", 0) / total if total else 0,
        "fail_rate": verdict_counts.get("fail", 0) / total if total else 0,
        "uncertain_rate": verdict_counts.get("uncertain", 0) / total if total else 0,
        "verdict_counts": verdict_counts,
        "severity_counts": severity_counts,
        "error_type_counts": error_counts,
    }


def _candidate_gold_from_skill_eval_cases(cases: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for case in cases:
        expected = str(case.get("expected_intent_summary") or "").strip()
        if not expected:
            continue
        rows.append(
            {
                "id": case.get("id"),
                "query": case.get("query"),
                "expected_intent": expected,
                "expected_intent_lines": case.get("expected_intent_lines") or [],
                "tags": ["skill_eval_candidate_gold"],
                "label_meta": {
                    "source": "skill_eval",
                    "verdict": case.get("verdict"),
                    "confidence": case.get("confidence"),
                    "severity": case.get("severity"),
                    "needs_review": case.get("verdict") != "pass",
                    "error_types": case.get("error_types") or [],
                },
            }
        )
    return rows


def _analyze_intent_review_row(record: dict[str, Any]) -> dict[str, Any]:
    query = str(_excel_first(record, "query", "question", "问题") or "")
    intent_summary = str(_excel_first(record, "intent_summary", "actual_intent_summary", "candidate_intent_summary") or "")
    matched_level = str(_excel_first(record, "matched_level", "level") or "")
    error = str(_excel_first(record, "error") or "").strip()
    conditions = _parse_excel_json_value(_excel_first(record, "actual_conditions", "conditions")) or []
    if not isinstance(conditions, list):
        conditions = []
    fields = {str(condition.get("field") or "") for condition in conditions if isinstance(condition, dict)}
    operators = {str(condition.get("operator") or "") for condition in conditions if isinstance(condition, dict)}
    has_unsupported_intent_summary = _has_unsupported_intent_summary(intent_summary)
    duplicate_value_groups = _duplicate_condition_values_across_fields(conditions, query)
    short_value_groups = _suspicious_short_condition_values(conditions, query)
    duplicate_intent_concepts = _duplicate_intent_concepts(intent_summary, query)
    invalid_enum_values = _invalid_enum_condition_values(conditions)
    duplicate_same_field_values = _duplicate_same_field_value_operator_conditions(conditions, query)
    special_field_findings = _special_field_usage_findings(conditions, query)
    possible_intent_summary = ""

    issue_types: list[str] = []
    risk_level = "low"
    suggestion = "低风险：可抽查后填“通过”"
    review_status = "待抽查"
    prefill_summary = ""

    if error:
        issue_types.append("api_error")
        risk_level = "high"
        suggestion = "接口错误：先排查请求或服务"
        review_status = "待复核"
    elif has_unsupported_intent_summary and not conditions:
        prefill_summary = intent_summary
    elif not has_unsupported_intent_summary and (
        not intent_summary or "未识别到明确查询条件" in intent_summary or not conditions
    ):
        issue_types.append("missing_condition")
        risk_level = "high"
        suggestion = "未识别或无条件：需要补 final_intent_summary / final_intent_lines，或 skip"
        review_status = "待复核"

    if not has_unsupported_intent_summary and matched_level.lower() in {"unknown", "none", "null", ""}:
        issue_types.append("unknown_level")
        risk_level = "high"
        review_status = "待复核"

    if has_unsupported_intent_summary and not conditions:
        return {
            "review_status": review_status,
            "issue_type": "none",
            "risk_level": risk_level,
            "auto_suggestion": suggestion,
            "prefill_final_intent_summary": prefill_summary,
            "possible_intent_summary": possible_intent_summary,
        }

    if re.search(r"1[3-9]\d{9}", query) and "clientMobile" not in fields:
        issue_types.append("suspected_mobile_field_wrong")
    if re.search(r"\d{17}[\dXx]", query) and "idNo" not in fields:
        issue_types.append("suspected_id_no_field_wrong")
    if re.search(r"(保单号|保单).*?[A-Za-z0-9]{6,}", query) and "polNo" not in fields:
        issue_types.append("suspected_policy_no_field_wrong")
    if _expects_negative_operator_from_config(query) and not (
        {"NOT_CONTAINS", "NOT_EXISTS"} & operators
        or _intent_summary_has_negative_operator_text(intent_summary)
    ):
        issue_types.append("suspected_negation_missing")
    if re.search(r"(或者|或|任一|之一)", query) and "或者" not in intent_summary:
        issue_types.append("suspected_logic_wrong")
    if (
        re.search(r"(以上|以下|大于|小于|不低于|不高于|至少|最多|区间|之间|到)", query)
        and not _has_ordered_enum_comparison(conditions, query)
        and not re.search(r"(≥|≤|>|<|在|~|之间)", intent_summary)
    ):
        issue_types.append("suspected_operator_wrong")
    if duplicate_value_groups:
        issue_types.append("suspected_duplicate_value_multi_fields")
    if duplicate_same_field_values:
        issue_types.append("suspected_duplicate_same_field_value_operator")
    if special_field_findings:
        issue_types.append("suspected_special_field_value_incomplete")
    if short_value_groups:
        issue_types.append("suspected_value_truncated_or_too_short")
    if duplicate_intent_concepts:
        issue_types.append("suspected_duplicate_intent_concept")
    if invalid_enum_values:
        issue_types.append("suspected_invalid_enum_value")

    if issue_types and risk_level != "high":
        risk_level = "medium"
        possible_intent_summary = _suggest_possible_intent_summary(
            conditions,
            query,
            invalid_enum_values=invalid_enum_values,
            short_value_groups=short_value_groups,
            duplicate_value_groups=duplicate_value_groups,
            duplicate_intent_concepts=duplicate_intent_concepts,
            duplicate_same_field_values=duplicate_same_field_values,
            special_field_findings=special_field_findings,
        )
        if invalid_enum_values:
            group_text = "；".join(
                f"{item['field']}={item['value']}"
                for item in invalid_enum_values[:3]
            )
            suggestion = f"疑似枚举值非法：{group_text}。请按字段枚举修正标准意图或删除误解析条件"
        elif short_value_groups:
            group_text = "；".join(
                f"{item['value']} -> {item['field']}"
                for item in short_value_groups[:3]
            )
            suggestion = f"疑似 value 截断或过短：{group_text}。请确认是否把长词误拆成了单字/短值"
        elif duplicate_same_field_values:
            group_text = "；".join(
                f"{item['field']}={item['value']}({','.join(item['operators'])})"
                for item in duplicate_same_field_values[:3]
            )
            suggestion = f"疑似同字段同值被多个 operator 解析：{group_text}。请保留符合 query 语义的 operator"
        elif special_field_findings:
            group_text = "；".join(
                f"{item['field']}={','.join(item['actual_values'])} -> {','.join(item['expected_values'])}"
                for item in special_field_findings[:3]
            )
            suggestion = f"疑似字段特殊用法未展开：{group_text}。请按字段业务语义修正标准意图"
        elif duplicate_value_groups:
            group_text = "；".join(
                f"{item['value']} -> {','.join(item['fields'])}"
                for item in duplicate_value_groups[:3]
            )
            suggestion = f"疑似同一查询条件命中多个字段：{group_text}。请确认是否需要保留多字段 OR，或修正为单一标准意图"
        elif duplicate_intent_concepts:
            group_text = "；".join(
                f"{item['concept']} -> 第{','.join(str(line) for line in item['line_numbers'])}条意图"
                for item in duplicate_intent_concepts[:3]
            )
            suggestion = f"疑似同一业务概念重复出现：{group_text}。请确认是否多解析了一个条件"
        else:
            suggestion = "疑似问题：请确认 intent_summary，错误则填写修正意图"
        review_status = "待复核"
    elif not issue_types:
        prefill_summary = intent_summary

    return {
        "review_status": review_status,
        "issue_type": ",".join(dict.fromkeys(issue_types)) or "none",
        "risk_level": risk_level,
        "auto_suggestion": suggestion,
        "prefill_final_intent_summary": prefill_summary,
        "possible_intent_summary": possible_intent_summary,
    }


def _suggest_possible_intent_summary(
    conditions: list[Any],
    query: str,
    *,
    invalid_enum_values: list[dict[str, str]],
    short_value_groups: list[dict[str, str]],
    duplicate_value_groups: list[dict[str, Any]],
    duplicate_intent_concepts: list[dict[str, Any]],
    duplicate_same_field_values: list[dict[str, Any]],
    special_field_findings: list[dict[str, Any]],
) -> str:
    removable = {
        (item["field"], item["value"])
        for item in invalid_enum_values + short_value_groups
    }
    removable.update(_removable_duplicate_value_conditions(conditions, duplicate_value_groups))
    removable.update(_removable_duplicate_concept_conditions(conditions, query, duplicate_intent_concepts))
    removable.update(_removable_duplicate_same_field_value_operator_conditions(conditions, duplicate_same_field_values, query))

    adjusted_conditions = _apply_special_field_replacements(conditions, special_field_findings)
    filtered: list[dict[str, Any]] = []
    removed_any = False
    replaced_any = bool(special_field_findings)
    for condition in adjusted_conditions:
        if not isinstance(condition, dict):
            continue
        if _condition_matches_removable(condition, removable):
            removed_any = True
            continue
        filtered.append(condition)
    if not removed_any and not replaced_any:
        return ""
    return _robot_text_from_expected({"query_logic": "AND", "conditions": filtered}) or ""


def _apply_special_field_replacements(
    conditions: list[Any],
    special_field_findings: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if not special_field_findings:
        return [condition for condition in conditions if isinstance(condition, dict)]
    replacements = {
        str(item.get("field") or ""): item.get("expected_values") or []
        for item in special_field_findings
    }
    adjusted: list[dict[str, Any]] = []
    for condition in conditions:
        if not isinstance(condition, dict):
            continue
        copied = dict(condition)
        field = str(copied.get("field") or "")
        if field in replacements:
            copied["operator"] = "CONTAINS"
            copied["value"] = replacements[field]
        adjusted.append(copied)
    return adjusted


def _condition_matches_removable(condition: dict[str, Any], removable: set[tuple[str, str]]) -> bool:
    field = str(condition.get("field") or "")
    operator = str(condition.get("operator") or "")
    for value in _condition_scalar_values(condition.get("value")):
        if (field, str(value)) in removable or (field, _normalize_raw_value_for_review(value)) in removable:
            return True
        if (f"{field}::{operator}", str(value)) in removable or (f"{field}::{operator}", _normalize_raw_value_for_review(value)) in removable:
            return True
    return False


def _has_unsupported_intent_summary(intent_summary: str) -> bool:
    return bool(
        re.search(
            r"提示[:：].+?暂不支持搜索，(?:无法进行查询|系统将按可支持字段搜索)。?",
            str(intent_summary or ""),
        )
    )


def _expects_negative_operator_from_config(query: str) -> bool:
    normalized_query = _normalize_raw_value_for_review(query)
    if not normalized_query or not _phrase_has_negative_operator_intent(normalized_query):
        return False
    for definition in _load_field_definitions():
        if str(definition.get("operator") or "") not in {"NOT_CONTAINS", "NOT_EXISTS"}:
            continue
        if _negative_definition_matches_query_text(definition, normalized_query):
            return True
    return False


def _negative_definition_matches_query_text(definition: dict[str, Any], normalized_query: str) -> bool:
    for phrase in _definition_match_phrases(definition):
        normalized_phrase = _normalize_raw_value_for_review(phrase)
        if len(normalized_phrase) < 2 or not _phrase_has_negative_operator_intent(normalized_phrase):
            continue
        if normalized_phrase in normalized_query:
            return True
    return False


def _phrase_has_negative_operator_intent(normalized_phrase: str) -> bool:
    return bool(re.search(r"(未|没有|没|无|不含|不包含|不是|不属于|非|缺少)", normalized_phrase))


def _definition_matches_query_text(definition: dict[str, Any], normalized_query: str) -> bool:
    for phrase in _definition_match_phrases(definition):
        normalized_phrase = _normalize_raw_value_for_review(phrase)
        if len(normalized_phrase) < 2:
            continue
        if normalized_phrase in normalized_query or normalized_query in normalized_phrase:
            return True
    return False


def _definition_match_phrases(definition: dict[str, Any]) -> list[str]:
    phrases: list[str] = []
    retrieval_text = str(definition.get("retrieval_text") or "")
    phrases.extend(re.split(r"[\s,，、;；]+", retrieval_text))
    for example in definition.get("examples") or []:
        if isinstance(example, dict) and example.get("query"):
            phrases.append(str(example.get("query")))
    return [phrase for phrase in phrases if phrase]


def _intent_summary_has_negative_operator_text(intent_summary: str) -> bool:
    return bool(re.search(r"(没有|不包含|不是|无|未配置|未购买|缺少)", str(intent_summary or "")))


def _polarity_value_findings(conditions: list[Any], query: str) -> list[dict[str, Any]]:
    expected_value = _expected_purchase_polarity_value(query)
    if not expected_value:
        return []
    opposite = "没有购买" if expected_value == "有购买" else "有购买"
    findings: list[dict[str, Any]] = []
    for condition in conditions:
        if not isinstance(condition, dict):
            continue
        field = str(condition.get("field") or "")
        operator = str(condition.get("operator") or "")
        if operator != "MATCH":
            continue
        for value in _condition_scalar_values(condition.get("value")):
            normalized = _normalize_raw_value_for_review(value)
            if normalized == opposite:
                findings.append(
                    {
                        "field": field,
                        "actual_value": str(value),
                        "expected_value": expected_value,
                    }
                )
    return findings


def _expected_purchase_polarity_value(query: str) -> str:
    normalized = _normalize_raw_value_for_review(query)
    if not normalized:
        return ""
    if _phrase_has_negative_operator_intent(normalized) and re.search(r"(配置|购买|买|投保|持有|养老险|健康险|产险|保险)", normalized):
        return "没有购买"
    if re.search(r"(已配置|配置了|配置有|配有|购买了|买了|买过|投保了|投保过|持有|已有|有了|有过|有买|有购买|有养老险|有健康险|有产险)", normalized):
        return "有购买"
    return ""


def _special_field_usage_findings(conditions: list[Any], query: str) -> list[dict[str, Any]]:
    findings: list[dict[str, Any]] = []
    expected_assets_values = _expected_assets_condition_values_from_query(query)
    if expected_assets_values:
        for condition in conditions:
            if not isinstance(condition, dict):
                continue
            if str(condition.get("field") or "") != "assetsCondition":
                continue
            actual_values = _condition_scalar_values(condition.get("value"))
            if set(map(_normalize_raw_value_for_review, actual_values)) != set(map(_normalize_raw_value_for_review, expected_assets_values)):
                findings.append(
                    {
                        "field": "assetsCondition",
                        "actual_values": actual_values,
                        "expected_values": expected_assets_values,
                        "reason": "assetsCondition 特殊枚举：有车/有房需要覆盖组合资产枚举。",
                    }
                )
    expected_policy_status_values = _expected_policy_status_values_from_query(query)
    if expected_policy_status_values:
        for condition in conditions:
            if not isinstance(condition, dict):
                continue
            if str(condition.get("field") or "") != "polNoInfo.polStatus":
                continue
            actual_values = _condition_scalar_values(condition.get("value"))
            if set(map(_normalize_raw_value_for_review, actual_values)) != set(map(_normalize_raw_value_for_review, expected_policy_status_values)):
                findings.append(
                    {
                        "field": "polNoInfo.polStatus",
                        "actual_values": actual_values,
                        "expected_values": expected_policy_status_values,
                        "reason": "保单状态特殊口径：缴费有效只等于交费有效，保单有效才展开为有效状态集合。",
                    }
                )
    findings.extend(_ordered_enum_usage_findings(conditions, query))
    return findings


def _expected_assets_condition_values_from_query(query: str) -> list[str]:
    normalized = _normalize_raw_value_for_review(query)
    if not normalized:
        return []
    if re.search(r"(有房有车|有车有房)", normalized):
        return ["有房有车"]
    if re.search(r"(无房无车|无车无房)", normalized):
        return ["无房无车"]
    if re.search(r"(有车无房|有车没房|有车没有房|只有车)", normalized):
        return ["有车"]
    if re.search(r"(无车有房|没车有房|没有车有房|只有房)", normalized):
        return ["有房"]
    if re.search(r"有车(?!险)", normalized):
        return ["有车", "有房有车"]
    if re.search(r"有房", normalized):
        return ["有房", "有房有车"]
    return []


def _expected_policy_status_values_from_query(query: str) -> list[str]:
    normalized = _normalize_raw_value_for_review(query)
    if not normalized:
        return []
    if re.search(r"(缴费有效|交费有效)", normalized):
        return ["交费有效"]
    if re.search(r"(保单有效|有效保单|生效保单|保单生效|保单状态有效)", normalized):
        return ["交费有效", "自垫交清", "交清", "减额交清", "免交", "自垫有效"]
    return []


def _ordered_enum_usage_findings(conditions: list[Any], query: str) -> list[dict[str, Any]]:
    ordered_values_by_field = _load_ordered_enum_values()
    if not ordered_values_by_field:
        return []
    findings: list[dict[str, Any]] = []
    for condition in conditions:
        if not isinstance(condition, dict):
            continue
        field = str(condition.get("field") or "")
        ordered_values = ordered_values_by_field.get(field)
        if not ordered_values:
            continue
        expected_values = _expected_ordered_enum_values_from_query(query, ordered_values, field)
        if not expected_values:
            continue
        actual_values = _condition_scalar_values(condition.get("value"))
        if set(map(_normalize_raw_value_for_review, actual_values)) == set(map(_normalize_raw_value_for_review, expected_values)):
            continue
        findings.append(
            {
                "field": field,
                "actual_values": actual_values,
                "expected_values": expected_values,
                "reason": "有序枚举比较应按 field_enums_args.yaml 中 ordered: true 的低到高顺序展开。",
            }
        )
    return findings


def _has_ordered_enum_comparison(conditions: list[Any], query: str) -> bool:
    ordered_values_by_field = _load_ordered_enum_values()
    if not ordered_values_by_field:
        return False
    fields = {
        str(condition.get("field") or "")
        for condition in conditions
        if isinstance(condition, dict)
    }
    for field in fields:
        ordered_values = ordered_values_by_field.get(field)
        if ordered_values and _expected_ordered_enum_values_from_query(query, ordered_values, field):
            return True
    return False


def _expected_ordered_enum_values_from_query(query: str, ordered_values: list[str], field: str) -> list[str]:
    normalized = _normalize_raw_value_for_review(query)
    group_values = _expected_ordered_enum_group_values_from_query(normalized, ordered_values, field)
    if group_values:
        return group_values
    value_aliases = _ordered_enum_value_aliases(field, ordered_values)
    for value, aliases in sorted(value_aliases.items(), key=lambda item: max(len(alias) for alias in item[1]), reverse=True):
        suffix = r"(?:类|等级|档|档位|客户)?"
        index = ordered_values.index(value)
        for alias in sorted(aliases, key=len, reverse=True):
            escaped = re.escape(_normalize_raw_value_for_review(alias))
            if re.search(rf"{escaped}{suffix}及以上", normalized):
                return ordered_values[index:]
            if re.search(rf"{escaped}{suffix}以上", normalized):
                return ordered_values[index + 1:]
            if re.search(rf"{escaped}{suffix}及以下", normalized):
                return ordered_values[:index + 1]
            if re.search(rf"{escaped}{suffix}以下", normalized):
                return ordered_values[:index]
    return []


def _expected_ordered_enum_group_values_from_query(normalized_query: str, ordered_values: list[str], field: str) -> list[str]:
    groups = _ordered_enum_group_aliases(field)
    for alias, group_members in sorted(groups.items(), key=lambda item: len(item[0]), reverse=True):
        indexes = [ordered_values.index(value) for value in group_members if value in ordered_values]
        if not indexes:
            continue
        start = min(indexes)
        end = max(indexes)
        escaped = re.escape(_normalize_raw_value_for_review(alias))
        suffix = r"(?:VIP|客户|会员|等级|档|档位)?"
        if re.search(rf"{escaped}{suffix}及以上", normalized_query):
            return ordered_values[start:]
        if re.search(rf"{escaped}{suffix}(?:以上|更高)", normalized_query):
            return ordered_values[end + 1:]
        if re.search(rf"{escaped}{suffix}及以下", normalized_query):
            return ordered_values[:end + 1]
        if re.search(rf"{escaped}{suffix}以下", normalized_query):
            return ordered_values[:start]
    return []


def _ordered_enum_group_aliases(field: str) -> dict[str, list[str]]:
    if field == "vipType":
        return {
            "白银": ["白银1", "白银2", "白银3"],
            "黄金": ["黄金V1", "黄金V2", "黄金V3", "原黄金VIP"],
            "铂金": ["铂金V1", "铂金V2", "原铂金VIP"],
        }
    return {}


def _ordered_enum_value_aliases(field: str, ordered_values: list[str]) -> dict[str, set[str]]:
    aliases = {value: {value} for value in ordered_values}
    mappings = _load_value_mappings().get(field) or {}
    for alias, canonical in mappings.items():
        canonical_text = str(canonical)
        if canonical_text in aliases:
            aliases[canonical_text].add(str(alias))
    return aliases


def _duplicate_same_field_value_operator_conditions(conditions: list[Any], query: str) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str], set[str]] = {}
    for condition in conditions:
        if not isinstance(condition, dict):
            continue
        field = str(condition.get("field") or "")
        operator = str(condition.get("operator") or "")
        if not field or not operator:
            continue
        for value in _condition_scalar_values(condition.get("value")):
            normalized = _normalize_raw_value_for_review(value)
            if normalized:
                groups.setdefault((field, normalized), set()).add(operator)
    return [
        {"field": field, "value": value, "operators": sorted(operators)}
        for (field, value), operators in sorted(groups.items())
        if len(operators) > 1
    ]


def _removable_duplicate_same_field_value_operator_conditions(
    conditions: list[Any],
    duplicate_same_field_values: list[dict[str, Any]],
    query: str,
) -> set[tuple[str, str]]:
    removable: set[tuple[str, str]] = set()
    if not duplicate_same_field_values:
        return removable
    preferred = _preferred_operator_from_query(query)
    for group in duplicate_same_field_values:
        field = str(group.get("field") or "")
        value = str(group.get("value") or "")
        operators = set(group.get("operators") or [])
        keep_operator = preferred if preferred in operators else _default_operator_to_keep(operators)
        for operator in operators:
            if operator == keep_operator:
                continue
            removable.add((f"{field}::{operator}", value))
    return removable


def _preferred_operator_from_query(query: str) -> str | None:
    if re.search(r"(及以上|以上|大于等于|不低于|至少)", query):
        return "GTE"
    if re.search(r"(及以下|以下|小于等于|不高于|最多)", query):
        return "LTE"
    if re.search(r"(大于|超过|高于)", query):
        return "GT"
    if re.search(r"(小于|低于)", query):
        return "LT"
    return None


def _default_operator_to_keep(operators: set[str]) -> str:
    for operator in ("RANGE", "GTE", "LTE", "GT", "LT", "MATCH", "CONTAINS"):
        if operator in operators:
            return operator
    return sorted(operators)[0]


def _removable_duplicate_value_conditions(
    conditions: list[Any],
    duplicate_value_groups: list[dict[str, Any]],
) -> set[tuple[str, str]]:
    removable: set[tuple[str, str]] = set()
    if not duplicate_value_groups:
        return removable
    for group in duplicate_value_groups:
        fields = set(group.get("fields") or [])
        candidates = [
            condition for condition in conditions
            if isinstance(condition, dict) and str(condition.get("field") or "") in fields
        ]
        keep = _choose_duplicate_condition_to_keep(candidates)
        for condition in candidates:
            if condition is keep:
                continue
            field = str(condition.get("field") or "")
            for value in _condition_scalar_values(condition.get("value")):
                removable.add((field, str(value)))
                removable.add((field, _normalize_raw_value_for_review(value)))
    return removable


def _choose_duplicate_condition_to_keep(candidates: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not candidates:
        return None
    priority = {
        "pCategorys": 90,
        "isBuyInsuranceCar": 90,
        "familyInfo.familyrelation": 90,
        "polNoInfo.plancodeinfo.plantypedesc": 80,
        "planAbbrNames": 80,
        "polNoInfo.plancodeinfo.abbrname": 40,
        "assetsCondition": 30,
    }
    return max(
        candidates,
        key=lambda condition: (
            priority.get(str(condition.get("field") or ""), 50),
            max((len(_normalize_raw_value_for_review(value)) for value in _condition_scalar_values(condition.get("value"))), default=0),
        ),
    )


def _removable_duplicate_concept_conditions(
    conditions: list[Any],
    query: str,
    duplicate_intent_concepts: list[dict[str, Any]],
) -> set[tuple[str, str]]:
    removable: set[tuple[str, str]] = set()
    if not duplicate_intent_concepts:
        return removable
    concepts = {str(item.get("concept") or "") for item in duplicate_intent_concepts}
    if "寿险" in concepts:
        for condition in conditions:
            if not isinstance(condition, dict):
                continue
            field = str(condition.get("field") or "")
            if field == "polNoInfo.plancodeinfo.plantypedesc":
                for value in _condition_scalar_values(condition.get("value")):
                    if _normalize_raw_value_for_review(value) == "寿险":
                        removable.add((field, str(value)))
    return removable


def _invalid_enum_condition_values(conditions: list[Any]) -> list[dict[str, str]]:
    enum_values = _load_enum_values()
    invalid: list[dict[str, str]] = []
    for condition in conditions:
        if not isinstance(condition, dict):
            continue
        field = str(condition.get("field") or "")
        allowed = set(enum_values.get(field) or [])
        if not allowed:
            continue
        operator = str(condition.get("operator") or "")
        if operator in {"EXISTS", "NOT_EXISTS"}:
            continue
        for value in _condition_scalar_values(condition.get("value")):
            value_text = str(value)
            if value_text not in allowed:
                invalid.append({"field": field, "value": value_text})
    return invalid


def _duplicate_condition_values_across_fields(conditions: list[Any], query: str = "") -> list[dict[str, Any]]:
    values: dict[str, set[str]] = {}
    value_items: list[dict[str, str]] = []
    for condition in conditions:
        if not isinstance(condition, dict):
            continue
        operator = str(condition.get("operator") or "")
        if operator not in {"MATCH", "CONTAINS", "NOT_CONTAINS"}:
            continue
        field = str(condition.get("field") or "")
        if not field:
            continue
        for value in _condition_scalar_values(condition.get("value")):
            normalized = _normalize_duplicate_value(value)
            if not normalized:
                continue
            values.setdefault(normalized, set()).add(field)
            value_items.append({"value": normalized, "field": field})

    duplicate_groups = [
        {"value": value, "fields": sorted(fields)}
        for value, fields in sorted(values.items())
        if len(fields) > 1
    ]
    seen = {
        (group["value"], tuple(group["fields"]))
        for group in duplicate_groups
    }
    query_normalized = _normalize_duplicate_value(query)
    for index, left in enumerate(value_items):
        for right in value_items[index + 1:]:
            if left["field"] == right["field"]:
                continue
            overlap_value = _overlapping_condition_value(left["value"], right["value"], query_normalized)
            if not overlap_value:
                continue
            fields = tuple(sorted({left["field"], right["field"]}))
            key = (overlap_value, fields)
            if key in seen:
                continue
            duplicate_groups.append({"value": overlap_value, "fields": list(fields)})
            seen.add(key)
    return duplicate_groups


def _suspicious_short_condition_values(conditions: list[Any], query: str) -> list[dict[str, str]]:
    query_normalized = _normalize_raw_value_for_review(query)
    suspicious: list[dict[str, str]] = []
    for condition in conditions:
        if not isinstance(condition, dict):
            continue
        operator = str(condition.get("operator") or "")
        if operator not in {"MATCH", "CONTAINS", "NOT_CONTAINS"}:
            continue
        field = str(condition.get("field") or "")
        for value in _condition_scalar_values(condition.get("value")):
            normalized = _normalize_raw_value_for_review(value)
            if not normalized:
                continue
            if _is_valid_short_enum_phrase(field, normalized, query_normalized):
                continue
            if (
                len(normalized) == 1
                and re.match(r"[A-Za-z0-9]", normalized)
                and re.search(rf"{re.escape(normalized)}[\u4e00-\u9fffA-Za-z0-9]{{1,}}", query_normalized)
            ):
                suspicious.append({"field": field, "value": normalized})
    return suspicious


def _is_valid_short_enum_phrase(field: str, value: str, query_normalized: str) -> bool:
    if field != "newValueLabel":
        return False
    allowed = set(_load_enum_values().get(field) or [])
    if value not in allowed:
        return False
    return bool(re.search(rf"(?:客户价值|价值标签)?{re.escape(value)}(?:类|类客户|客户)?", query_normalized))


def _duplicate_intent_concepts(intent_summary: str, query: str) -> list[dict[str, Any]]:
    query_normalized = _normalize_duplicate_value(query)
    lines = extract_intent_lines(intent_summary)
    concepts_by_line: dict[str, set[int]] = {}
    for line_number, line in enumerate(lines, start=1):
        for concept in _intent_concepts_from_line(line):
            if concept not in {"寿险"}:
                continue
            if concept not in query_normalized:
                continue
            concepts_by_line.setdefault(concept, set()).add(line_number)
    return [
        {"concept": concept, "line_numbers": sorted(line_numbers)}
        for concept, line_numbers in sorted(concepts_by_line.items(), key=lambda item: (-len(item[0]), item[0]))
        if len(line_numbers) > 1
    ]


def _intent_concepts_from_line(line: str) -> set[str]:
    normalized = _normalize_duplicate_value(line)
    stopwords = {
        "客户",
        "包含",
        "投保",
        "险种",
        "简称",
        "类别",
        "产品",
        "状态",
        "价值",
        "标签",
        "等级",
        "年龄",
        "婚姻",
        "状况",
    }
    concepts: set[str] = set()
    for token in re.findall(r"[\u4e00-\u9fffA-Za-z0-9]+", normalized):
        if len(token) < 2:
            continue
        for size in range(2, min(6, len(token)) + 1):
            for start in range(0, len(token) - size + 1):
                concept = token[start:start + size]
                if concept in stopwords:
                    continue
                concepts.add(concept)
    return concepts


def _overlapping_condition_value(left: str, right: str, query: str) -> str:
    if left == right:
        return left
    if len(left) >= 2 and len(right) >= 2 and (left in right or right in left):
        return right if left in right else left
    if query:
        left_spans = _find_all_spans(query, left)
        right_spans = _find_all_spans(query, right)
        if _has_non_overlapping_span_pair(left_spans, right_spans):
            return ""
        for left_span in left_spans:
            for right_span in right_spans:
                if max(left_span[0], right_span[0]) < min(left_span[1], right_span[1]):
                    start = min(left_span[0], right_span[0])
                    end = max(left_span[1], right_span[1])
                    return query[start:end]
    return ""


def _has_non_overlapping_span_pair(
    left_spans: list[tuple[int, int]],
    right_spans: list[tuple[int, int]],
) -> bool:
    for left_span in left_spans:
        for right_span in right_spans:
            if left_span[1] <= right_span[0] or right_span[1] <= left_span[0]:
                return True
    return False


def _find_all_spans(text: str, needle: str) -> list[tuple[int, int]]:
    if not needle:
        return []
    spans: list[tuple[int, int]] = []
    start = 0
    while True:
        index = text.find(needle, start)
        if index < 0:
            return spans
        spans.append((index, index + len(needle)))
        start = index + 1


def _find_span(text: str, needle: str) -> tuple[int, int] | None:
    if not needle:
        return None
    index = text.find(needle)
    if index < 0:
        return None
    return index, index + len(needle)


def _condition_scalar_values(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item) for item in value if str(item).strip()]
    if isinstance(value, dict):
        return []
    return [str(value)]


def _normalize_duplicate_value(value: str) -> str:
    normalized = _normalize_raw_value_for_review(value)
    if not normalized or normalized in {"客户", "有购买", "没有购买", "是", "否"}:
        return ""
    if len(normalized) < 2 and not normalized.isdigit():
        return ""
    return normalized


def _normalize_raw_value_for_review(value: Any) -> str:
    return re.sub(r"[\s()（）]+", "", str(value).strip())


def _parse_excel_json_value(value: Any) -> Any:
    if value in (None, ""):
        return None
    if isinstance(value, (dict, list)):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return None


def _format_intent_review_sheet(sheet: Any, header_index: dict[str, int], review_headers: list[str]) -> None:
    header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for header in review_headers:
        cell = sheet.cell(1, header_index[header])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in range(2, sheet.max_row + 1):
        risk = str(sheet.cell(row, header_index["risk_level"]).value or "")
        fill = None
        if risk == "high":
            fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        elif risk == "medium":
            fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        elif risk == "low":
            fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        if fill:
            for header in ("review_status", "issue_type", "risk_level", "auto_suggestion"):
                sheet.cell(row, header_index[header]).fill = fill
        for header in review_headers:
            sheet.cell(row, header_index[header]).alignment = Alignment(wrap_text=True, vertical="top")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}1"
    wide_headers = {"auto_suggestion", "possible_intent_summary", "final_intent_summary", "final_intent_lines"}
    for header in review_headers:
        column_letter = get_column_letter(header_index[header])
        sheet.column_dimensions[column_letter].width = 52 if header in wide_headers else 28


def _excel_first(record: dict[str, Any], *names: str) -> Any:
    for name in names:
        if name in record and record[name] not in (None, ""):
            return record[name]
    return None


def _is_accepted_review_status(value: str) -> bool:
    return value.strip().lower() in ACCEPTED_REVIEW_STATUSES


def _parse_excel_intent_lines(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return [normalize_intent_line(line) for line in value if str(line).strip()]
    text = str(value).strip()
    if not text:
        return []
    if text.startswith("["):
        parsed = _parse_json_field(text, "final_intent_lines")
        if not isinstance(parsed, list):
            raise ValueError("final_intent_lines must be a JSON list")
        return [normalize_intent_line(line) for line in parsed if str(line).strip()]
    return [normalize_intent_line(line) for line in re.split(r"[\n；;]+", text) if line.strip()]


def build_expected_candidates_from_config_examples(
    cases: list[dict[str, Any]],
    *,
    field_definitions_path: str | Path | None = None,
) -> list[dict[str, Any]]:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    example_index = _load_field_definition_example_index(field_definitions_path)
    candidates: list[dict[str, Any]] = []

    for case in cases:
        tags = list(case.get("tags") or [])
        if "auto_expected" not in tags:
            tags.append("auto_expected")
        if "config_example" not in tags:
            tags.append("config_example")

        match = example_index.get(_normalize_query_key(case.get("query")))
        label_meta = {
            "source": "field_definitions_examples",
            "generated_at": generated_at,
            "needs_review": False,
            "labeling_mode": "config_examples_exact_match",
        }
        output: dict[str, Any] = {
            "id": case.get("id"),
            "query": case.get("query"),
            "tags": tags,
            "label_meta": label_meta,
        }
        if match:
            output["label_status"] = match["label_status"]
            output["expected"] = match["expected"]
            output["robot_text"] = _robot_text_from_expected(output["expected"])
            output["label_meta"].update(match.get("label_meta") or {})
        else:
            output["label_status"] = "manual_required"
            output["label_meta"]["needs_review"] = True
            output["label_meta"]["reason"] = "no_exact_config_example"
        candidates.append(output)

    return candidates


async def evaluate_question_batch(
    cases: list[dict[str, Any]],
    options: EvalOptions,
    progress_callback: ProgressCallback | None = None,
) -> dict[str, Any]:
    semaphore = asyncio.Semaphore(max(1, options.concurrency))
    timeout = httpx.Timeout(options.timeout_seconds)

    async with httpx.AsyncClient(timeout=timeout) as client:
        async def evaluate_one(index: int, case: dict[str, Any]) -> tuple[int, dict[str, Any]]:
            async with semaphore:
                actual, elapsed_ms, error = await _call_parse_api(client, options.base_url, str(case.get("query") or ""))
                has_expected = "expected" in case
                expected = case.get("expected") or {}
                comparison = compare_result(expected, actual) if has_expected else None
                return index, {
                    "id": case.get("id"),
                    "query": case.get("query"),
                    "tags": case.get("tags") or [],
                    "expected": expected if has_expected else None,
                    "actual": actual,
                    "comparison": comparison,
                    "elapsed_ms": elapsed_ms,
                    "error": error,
                    "graded": has_expected,
                }

        tasks = [asyncio.create_task(evaluate_one(index, case)) for index, case in enumerate(cases)]
        case_results: list[dict[str, Any] | None] = [None] * len(tasks)
        completed = 0
        errors = 0
        total_elapsed_ms = 0.0
        for task in asyncio.as_completed(tasks):
            index, result = await task
            case_results[index] = result
            completed += 1
            if result["error"]:
                errors += 1
            total_elapsed_ms += float(result["elapsed_ms"] or 0)
            if progress_callback:
                progress_callback(
                    {
                        "completed": completed,
                        "total": len(tasks),
                        "errors": errors,
                        "avg_latency_ms": total_elapsed_ms / completed if completed else 0,
                        "case_id": result.get("id"),
                        "query": result.get("query"),
                        "elapsed_ms": result.get("elapsed_ms"),
                        "matched_level": result.get("actual", {}).get("matched_level"),
                        "error": result.get("error"),
                    }
                )

    finalized_case_results = [item for item in case_results if item is not None]
    total = len(finalized_case_results)
    graded_cases = [item for item in finalized_case_results if item["graded"]]
    exact_matches = sum(1 for item in graded_cases if item["comparison"] and item["comparison"]["exact_match"])
    field_matches = sum(1 for item in graded_cases if item["comparison"] and item["comparison"]["field_match"])
    operator_matches = sum(1 for item in graded_cases if item["comparison"] and item["comparison"]["operator_match"])
    negative_cases = [
        item for item in graded_cases
        if not ((item.get("expected") or {}).get("conditions") or [])
    ]
    positive_cases = [item for item in graded_cases if item not in negative_cases]
    empty_positive = [
        item for item in positive_cases
        if not item["actual"].get("conditions")
    ]
    false_positive = [
        item for item in negative_cases
        if item["actual"].get("conditions")
    ]
    latencies = [item["elapsed_ms"] for item in finalized_case_results]
    non_empty_cases = [
        item for item in finalized_case_results
        if item["actual"].get("conditions")
    ]
    known_level_cases = [
        item for item in finalized_case_results
        if item["actual"].get("matched_level") is not None
    ]
    level_distribution: dict[str, int] = {}
    for item in finalized_case_results:
        level = str(item["actual"].get("matched_level") or "unknown")
        level_distribution[level] = level_distribution.get(level, 0) + 1

    summary = {
        "total": total,
        "graded_total": len(graded_cases),
        "ungraded_total": total - len(graded_cases),
        "graded_coverage_rate": len(graded_cases) / total if total else 0,
        "api_success_rate": (total - sum(1 for item in finalized_case_results if item["error"])) / total if total else 0,
        "condition_non_empty_rate": len(non_empty_cases) / total if total else 0,
        "known_level_rate": len(known_level_cases) / total if total else 0,
        "total_accuracy": exact_matches / len(graded_cases) if graded_cases else None,
        "overall_accuracy": exact_matches / len(graded_cases) if graded_cases else None,
        "exact_match_rate": exact_matches / len(graded_cases) if graded_cases else None,
        "field_match_rate": field_matches / len(graded_cases) if graded_cases else None,
        "operator_match_rate": operator_matches / len(graded_cases) if graded_cases else None,
        "empty_rate": len(empty_positive) / len(positive_cases) if positive_cases else None,
        "false_positive_rate": len(false_positive) / len(negative_cases) if negative_cases else None,
        "avg_latency_ms": sum(latencies) / len(latencies) if latencies else 0,
        "p95_latency_ms": sorted(latencies)[int(len(latencies) * 0.95) - 1] if latencies else 0,
        "level_distribution": level_distribution,
        "error_count": sum(1 for item in case_results if item["error"]),
    }
    failed_cases = [
        item for item in graded_cases
        if item["error"] or not (item["comparison"] and item["comparison"]["exact_match"])
    ]

    return {
        "summary": summary,
        "cases": finalized_case_results,
        "failed_cases": failed_cases,
    }


def check_acceptance(summary: dict[str, Any], acceptance: dict[str, Any]) -> list[str]:
    failures: list[str] = []
    checks = [
        ("min_overall_accuracy", "overall_accuracy", ">="),
        ("min_exact_match_rate", "exact_match_rate", ">="),
        ("max_empty_rate", "empty_rate", "<="),
        ("max_false_positive_rate", "false_positive_rate", "<="),
        ("max_avg_latency_ms", "avg_latency_ms", "<="),
        ("max_p95_latency_ms", "p95_latency_ms", "<="),
    ]
    for acceptance_key, summary_key, operator in checks:
        if acceptance_key not in acceptance:
            continue
        expected = float(acceptance[acceptance_key])
        actual = float(summary.get(summary_key) or 0)
        if operator == ">=" and actual < expected:
            failures.append(f"{summary_key}={actual:.4f} below {acceptance_key}={expected:.4f}")
        if operator == "<=" and actual > expected:
            failures.append(f"{summary_key}={actual:.4f} above {acceptance_key}={expected:.4f}")
    return failures
