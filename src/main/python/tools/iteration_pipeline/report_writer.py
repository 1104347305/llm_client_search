from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.main.python.tools.iteration_pipeline.change_set import ChangeSet


def write_eval_artifacts(change_set: ChangeSet, eval_result: dict[str, Any], acceptance_failures: list[str]) -> dict[str, Path]:
    output_dir = change_set.iteration_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    eval_path = output_dir / "eval_result.json"
    failed_path = output_dir / "failed_cases.jsonl"
    report_path = output_dir / "report.md"

    eval_path.write_text(
        json.dumps(eval_result, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    with failed_path.open("w", encoding="utf-8") as file:
        for case in eval_result.get("failed_cases") or []:
            file.write(json.dumps(case, ensure_ascii=False, sort_keys=True))
            file.write("\n")

    report_path.write_text(
        render_report(change_set, eval_result, acceptance_failures),
        encoding="utf-8",
    )
    return {
        "eval_result": eval_path,
        "failed_cases": failed_path,
        "report": report_path,
    }


def write_batch_eval_artifacts(
    input_path: Path,
    output_dir: Path,
    eval_result: dict[str, Any],
) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)

    eval_path = output_dir / "batch_eval_result.json"
    failed_path = output_dir / "batch_failed_cases.jsonl"
    report_path = output_dir / "batch_report.md"
    excel_path = output_dir / "batch_eval_result.xlsx"

    eval_path.write_text(
        json.dumps(eval_result, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    with failed_path.open("w", encoding="utf-8") as file:
        for case in eval_result.get("failed_cases") or []:
            file.write(json.dumps(case, ensure_ascii=False, sort_keys=True))
            file.write("\n")

    report_path.write_text(
        render_batch_report(input_path, eval_result),
        encoding="utf-8",
    )
    write_batch_eval_excel(input_path, eval_result, excel_path)
    return {
        "eval_result": eval_path,
        "failed_cases": failed_path,
        "report": report_path,
        "excel": excel_path,
    }


def write_intent_eval_artifacts(
    input_path: Path,
    output_dir: Path,
    eval_result: dict[str, Any],
) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)

    eval_path = output_dir / "intent_eval_result.json"
    failed_path = output_dir / "intent_failed_cases.jsonl"
    report_path = output_dir / "intent_report.md"
    excel_path = output_dir / "intent_eval_result.xlsx"

    eval_path.write_text(
        json.dumps(eval_result, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    with failed_path.open("w", encoding="utf-8") as file:
        for case in eval_result.get("failed_cases") or []:
            file.write(json.dumps(case, ensure_ascii=False, sort_keys=True))
            file.write("\n")

    report_path.write_text(
        render_intent_report(input_path, eval_result),
        encoding="utf-8",
    )
    write_intent_eval_excel(input_path, eval_result, excel_path)
    return {
        "eval_result": eval_path,
        "failed_cases": failed_path,
        "report": report_path,
        "excel": excel_path,
    }


def write_skill_eval_artifacts(
    input_path: Path,
    output_dir: Path,
    eval_result: dict[str, Any],
) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)

    eval_path = output_dir / "skill_eval_result.json"
    report_path = output_dir / "skill_eval_report.md"
    excel_path = output_dir / "skill_eval_result.xlsx"
    gold_path = output_dir / "candidate_intent_gold.jsonl"

    eval_path.write_text(
        json.dumps(eval_result, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    report_path.write_text(render_skill_eval_report(input_path, eval_result), encoding="utf-8")
    write_skill_eval_excel(input_path, eval_result, excel_path)
    with gold_path.open("w", encoding="utf-8") as file:
        for row in eval_result.get("candidate_gold") or []:
            file.write(json.dumps(row, ensure_ascii=False, sort_keys=True))
            file.write("\n")

    return {
        "eval_result": eval_path,
        "report": report_path,
        "excel": excel_path,
        "candidate_gold": gold_path,
    }


def write_batch_eval_excel(input_path: Path, eval_result: dict[str, Any], output_path: Path) -> Path:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    cases_sheet = workbook.create_sheet("cases")
    failed_sheet = workbook.create_sheet("failed_cases")

    _write_summary_sheet(summary_sheet, input_path, eval_result.get("summary") or {})
    _write_cases_sheet(cases_sheet, eval_result.get("cases") or [])
    _write_cases_sheet(failed_sheet, eval_result.get("failed_cases") or [])

    workbook.save(output_path)
    return output_path


def write_intent_eval_excel(input_path: Path, eval_result: dict[str, Any], output_path: Path) -> Path:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    cases_sheet = workbook.create_sheet("cases")
    failed_sheet = workbook.create_sheet("failed_cases")

    _write_summary_sheet(summary_sheet, input_path, eval_result.get("summary") or {})
    _write_intent_cases_sheet(cases_sheet, eval_result.get("cases") or [])
    _write_intent_cases_sheet(failed_sheet, eval_result.get("failed_cases") or [])

    workbook.save(output_path)
    return output_path


def write_skill_eval_excel(input_path: Path, eval_result: dict[str, Any], output_path: Path) -> Path:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    cases_sheet = workbook.create_sheet("cases")
    errors_sheet = workbook.create_sheet("error_counts")

    _write_summary_sheet(summary_sheet, input_path, eval_result.get("summary") or {})
    _write_skill_cases_sheet(cases_sheet, eval_result.get("cases") or [])
    _write_skill_error_counts_sheet(errors_sheet, eval_result.get("summary", {}).get("error_type_counts") or {})

    workbook.save(output_path)
    return output_path


def _write_summary_sheet(sheet, input_path: Path, summary: dict[str, Any]) -> None:
    rows = [
        ("generated_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("input_path", str(input_path)),
    ]
    rows.extend((key, value) for key, value in sorted(summary.items()))

    _write_header(sheet, ["metric", "value"])
    for row_index, (key, value) in enumerate(rows, start=2):
        sheet.cell(row_index, 1, key)
        cell = sheet.cell(row_index, 2, _excel_scalar(value))
        if _is_rate_metric(str(key)) and isinstance(value, (int, float)):
            cell.number_format = "0.00%"

    _autosize_columns(sheet)


def _write_cases_sheet(sheet, cases: list[dict[str, Any]]) -> None:
    headers = [
        "id",
        "query",
        "graded",
        "exact_match",
        "field_match",
        "operator_match",
        "query_logic_match",
        "elapsed_ms",
        "error",
        "tags",
        "expected_query_logic",
        "expected_conditions",
        "actual_query_logic",
        "actual_conditions",
        "matched_level",
        "intent_summary",
        "rewritten_query",
        "missing_conditions",
        "unexpected_conditions",
    ]
    _write_header(sheet, headers)

    for row_index, case in enumerate(cases, start=2):
        comparison = case.get("comparison") or {}
        expected = case.get("expected") or {}
        actual = case.get("actual") or {}
        values = [
            case.get("id"),
            case.get("query"),
            case.get("graded"),
            comparison.get("exact_match"),
            comparison.get("field_match"),
            comparison.get("operator_match"),
            comparison.get("query_logic_match"),
            case.get("elapsed_ms"),
            case.get("error"),
            case.get("tags"),
            expected.get("query_logic"),
            expected.get("conditions"),
            actual.get("query_logic"),
            actual.get("conditions"),
            actual.get("matched_level"),
            actual.get("intent_summary"),
            actual.get("rewritten_query"),
            comparison.get("missing_conditions"),
            comparison.get("unexpected_conditions"),
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column_index, _excel_scalar(value))
            if headers[column_index - 1] in {"query", "expected_conditions", "actual_conditions", "intent_summary"}:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    _autosize_columns(sheet, max_width=80)


def _write_intent_cases_sheet(sheet, cases: list[dict[str, Any]]) -> None:
    headers = [
        "id",
        "query",
        "graded",
        "intent_exact_match",
        "intent_line_exact_match",
        "intent_summary_exact_match",
        "intent_line_recall",
        "intent_line_precision",
        "elapsed_ms",
        "error",
        "tags",
        "expected_intent_summary",
        "expected_intent_lines",
        "actual_intent_summary",
        "actual_intent_lines",
        "missing_intent_lines",
        "unexpected_intent_lines",
        "matched_level",
        "actual_conditions",
    ]
    _write_header(sheet, headers)

    for row_index, case in enumerate(cases, start=2):
        comparison = case.get("intent_comparison") or {}
        actual = case.get("actual") or {}
        values = [
            case.get("id"),
            case.get("query"),
            case.get("graded"),
            comparison.get("intent_exact_match"),
            comparison.get("intent_line_exact_match"),
            comparison.get("intent_summary_exact_match"),
            comparison.get("intent_line_recall"),
            comparison.get("intent_line_precision"),
            case.get("elapsed_ms"),
            case.get("error"),
            case.get("tags"),
            case.get("expected_intent_summary"),
            case.get("expected_intent_lines"),
            case.get("actual_intent_summary"),
            case.get("actual_intent_lines"),
            comparison.get("missing_intent_lines"),
            comparison.get("unexpected_intent_lines"),
            actual.get("matched_level"),
            actual.get("conditions"),
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column_index, _excel_scalar(value))
            if headers[column_index - 1] in {
                "query",
                "expected_intent_summary",
                "expected_intent_lines",
                "actual_intent_summary",
                "actual_intent_lines",
                "missing_intent_lines",
                "unexpected_intent_lines",
                "actual_conditions",
            }:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    _autosize_columns(sheet, max_width=90)


def _write_skill_cases_sheet(sheet, cases: list[dict[str, Any]]) -> None:
    headers = [
        "id",
        "query",
        "verdict",
        "confidence",
        "severity",
        "error_types",
        "reason",
        "expected_intent_summary",
        "actual_intent_summary",
        "actual_conditions",
        "matched_level",
        "deterministic_findings",
        "review_status",
        "review_comment",
    ]
    _write_header(sheet, headers)
    for row_index, case in enumerate(cases, start=2):
        values = [
            case.get("id"),
            case.get("query"),
            case.get("verdict"),
            case.get("confidence"),
            case.get("severity"),
            case.get("error_types"),
            case.get("reason"),
            case.get("expected_intent_summary"),
            case.get("actual_intent_summary"),
            case.get("actual_conditions"),
            case.get("matched_level"),
            case.get("deterministic_findings"),
            "pending" if case.get("verdict") != "pass" else "spot_check",
            "",
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column_index, _excel_scalar(value))
            if headers[column_index - 1] in {
                "query",
                "error_types",
                "reason",
                "expected_intent_summary",
                "actual_intent_summary",
                "actual_conditions",
                "deterministic_findings",
                "review_comment",
            }:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    _autosize_columns(sheet, max_width=90)


def _write_skill_error_counts_sheet(sheet, error_counts: dict[str, int]) -> None:
    _write_header(sheet, ["error_type", "count"])
    for row_index, (error_type, count) in enumerate(sorted(error_counts.items()), start=2):
        sheet.cell(row_index, 1, error_type)
        sheet.cell(row_index, 2, count)
    _autosize_columns(sheet)


def _write_header(sheet, headers: list[str]) -> None:
    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column_index, header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _autosize_columns(sheet, max_width: int = 60) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), max_width))
        sheet.column_dimensions[column_letter].width = max(10, max_length + 2)


def _excel_scalar(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def _is_rate_metric(metric: str) -> bool:
    return metric.endswith("_rate") or metric in {"total_accuracy", "overall_accuracy"}


def _format_percent(value: float) -> str:
    return f"{value * 100:.2f}%"


def _format_optional_percent(value: Any) -> str:
    if value is None:
        return "N/A"
    return _format_percent(float(value))


def render_batch_report(input_path: Path, eval_result: dict[str, Any]) -> str:
    summary = eval_result.get("summary") or {}
    failed_cases = eval_result.get("failed_cases") or []
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines = [
        "# 批量问题自动化评估报告",
        "",
        f"生成时间：{generated_at}",
        f"输入文件：{input_path}",
        "",
        "## 概览",
        "",
        f"- 样本总数：{summary.get('total', 0)}",
        f"- 已标注样本：{summary.get('graded_total', 0)}",
        f"- 未标注样本：{summary.get('ungraded_total', 0)}",
        f"- graded_coverage_rate：{_format_percent(float(summary.get('graded_coverage_rate') or 0))}",
        f"- api_success_rate：{_format_percent(float(summary.get('api_success_rate') or 0))}",
        f"- condition_non_empty_rate：{_format_percent(float(summary.get('condition_non_empty_rate') or 0))}",
        f"- known_level_rate：{_format_percent(float(summary.get('known_level_rate') or 0))}",
        f"- total_accuracy：{_format_optional_percent(summary.get('total_accuracy'))}",
        f"- exact_match_rate：{_format_optional_percent(summary.get('exact_match_rate'))}",
        f"- field_match_rate：{_format_optional_percent(summary.get('field_match_rate'))}",
        f"- operator_match_rate：{_format_optional_percent(summary.get('operator_match_rate'))}",
        f"- empty_rate：{_format_optional_percent(summary.get('empty_rate'))}",
        f"- false_positive_rate：{_format_optional_percent(summary.get('false_positive_rate'))}",
        f"- avg_latency_ms：{float(summary.get('avg_latency_ms') or 0):.2f}",
        f"- p95_latency_ms：{float(summary.get('p95_latency_ms') or 0):.2f}",
        f"- error_count：{summary.get('error_count', 0)}",
        "",
        "## 层级分布",
        "",
        "| matched_level | 数量 |",
        "| --- | --- |",
    ]

    for level, count in sorted((summary.get("level_distribution") or {}).items()):
        lines.append(f"| {level} | {count} |")

    lines.extend(
        [
            "",
            "## 失败样本",
            "",
            "| id | query | 归因 |",
            "| --- | --- | --- |",
        ]
    )
    if failed_cases:
        for item in failed_cases[:100]:
            reason = _guess_failure_reason(item)
            query = str(item.get("query") or "").replace("|", "\\|")
            lines.append(f"| {item.get('id')} | {query} | {reason} |")
    else:
        lines.append("| - | - | 无 |")

    if not summary.get("graded_total"):
        lines.extend(
            [
                "",
                "## 标注建议",
                "",
                "- 当前输入没有 expected 标准答案，本次报告只统计解析层级、耗时和接口错误。",
                "- 后续可以把问题改成 JSONL/CSV，并补 expected.conditions，即可自动计算准确率。",
            ]
        )

    return "\n".join(lines) + "\n"


def render_intent_report(input_path: Path, eval_result: dict[str, Any]) -> str:
    summary = eval_result.get("summary") or {}
    failed_cases = eval_result.get("failed_cases") or []
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines = [
        "# 意图文本评估报告",
        "",
        f"生成时间：{generated_at}",
        f"输入文件：{input_path}",
        "",
        "## 概览",
        "",
        f"- 样本总数：{summary.get('total', 0)}",
        f"- 已标注样本：{summary.get('graded_total', 0)}",
        f"- 未标注样本：{summary.get('ungraded_total', 0)}",
        f"- graded_coverage_rate：{_format_percent(float(summary.get('graded_coverage_rate') or 0))}",
        f"- api_success_rate：{_format_percent(float(summary.get('api_success_rate') or 0))}",
        f"- intent_exact_match_rate：{_format_optional_percent(summary.get('intent_exact_match_rate'))}",
        f"- intent_line_exact_match_rate：{_format_optional_percent(summary.get('intent_line_exact_match_rate'))}",
        f"- intent_summary_exact_match_rate：{_format_optional_percent(summary.get('intent_summary_exact_match_rate'))}",
        f"- intent_line_recall：{_format_optional_percent(summary.get('intent_line_recall'))}",
        f"- intent_line_precision：{_format_optional_percent(summary.get('intent_line_precision'))}",
        f"- avg_latency_ms：{float(summary.get('avg_latency_ms') or 0):.2f}",
        f"- p95_latency_ms：{float(summary.get('p95_latency_ms') or 0):.2f}",
        f"- error_count：{summary.get('error_count', 0)}",
        "",
        "## 层级分布",
        "",
        "| matched_level | 数量 |",
        "| --- | --- |",
    ]

    for level, count in sorted((summary.get("level_distribution") or {}).items()):
        lines.append(f"| {level} | {count} |")

    lines.extend(
        [
            "",
            "## 失败样本",
            "",
            "| id | query | 缺失意图 | 多余意图 |",
            "| --- | --- | --- | --- |",
        ]
    )
    if failed_cases:
        for item in failed_cases[:100]:
            comparison = item.get("intent_comparison") or {}
            query = str(item.get("query") or "").replace("|", "\\|")
            missing = "；".join(comparison.get("missing_intent_lines") or []) or "-"
            unexpected = "；".join(comparison.get("unexpected_intent_lines") or []) or "-"
            lines.append(f"| {item.get('id')} | {query} | {missing} | {unexpected} |")
    else:
        lines.append("| - | - | 无 | 无 |")

    lines.extend(
        [
            "",
            "## 说明",
            "",
            "- `expected.conditions` 会通过现有 IntentSummaryService 生成标准意图文本。",
            "- 字段名、operator 文案、连接词和特殊模板来自 `intent_summary_labels_args.yaml`。",
            "- 主指标比较意图文本行集合；`actual_conditions` 保留用于失败定位。",
        ]
    )

    return "\n".join(lines) + "\n"


def render_skill_eval_report(input_path: Path, eval_result: dict[str, Any]) -> str:
    summary = eval_result.get("summary") or {}
    cases = eval_result.get("cases") or []
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        "# Skill 评估报告",
        "",
        f"生成时间：{generated_at}",
        f"输入文件：{input_path}",
        "",
        "## 概览",
        "",
        f"- 样本总数：{summary.get('total', 0)}",
        f"- pass_rate：{_format_percent(float(summary.get('pass_rate') or 0))}",
        f"- fail_rate：{_format_percent(float(summary.get('fail_rate') or 0))}",
        f"- uncertain_rate：{_format_percent(float(summary.get('uncertain_rate') or 0))}",
        "",
        "## 错误类型分布",
        "",
        "| error_type | 数量 |",
        "| --- | --- |",
    ]
    for error_type, count in sorted((summary.get("error_type_counts") or {}).items()):
        lines.append(f"| {error_type} | {count} |")
    lines.extend(
        [
            "",
            "## 高优先级样本",
            "",
            "| id | query | error_types | reason |",
            "| --- | --- | --- | --- |",
        ]
    )
    high_cases = [case for case in cases if case.get("verdict") != "pass"]
    if high_cases:
        for case in high_cases[:100]:
            query = str(case.get("query") or "").replace("|", "\\|")
            reason = str(case.get("reason") or "").replace("|", "\\|")
            lines.append(f"| {case.get('id')} | {query} | {','.join(case.get('error_types') or [])} | {reason} |")
    else:
        lines.append("| - | - | 无 | 无 |")
    return "\n".join(lines) + "\n"


def render_report(change_set: ChangeSet, eval_result: dict[str, Any], acceptance_failures: list[str]) -> str:
    summary = eval_result.get("summary") or {}
    failed_cases = eval_result.get("failed_cases") or []
    passed = not acceptance_failures
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines = [
        f"# {change_set.id} 自动化迭代报告",
        "",
        f"生成时间：{generated_at}",
        "",
        "## 结论",
        "",
        f"- 是否通过验收：{'是' if passed else '否'}",
        f"- 样本总数：{summary.get('total', 0)}",
        f"- 总准确率：{_format_percent(float(summary.get('total_accuracy', summary.get('overall_accuracy', summary.get('exact_match_rate'))) or 0))}",
        f"- 整体准确率：{_format_percent(float(summary.get('overall_accuracy', summary.get('exact_match_rate')) or 0))}",
        f"- exact_match_rate：{_format_percent(float(summary.get('exact_match_rate') or 0))}",
        f"- field_match_rate：{_format_percent(float(summary.get('field_match_rate') or 0))}",
        f"- operator_match_rate：{_format_percent(float(summary.get('operator_match_rate') or 0))}",
        f"- empty_rate：{_format_percent(float(summary.get('empty_rate') or 0))}",
        f"- false_positive_rate：{_format_percent(float(summary.get('false_positive_rate') or 0))}",
        f"- avg_latency_ms：{float(summary.get('avg_latency_ms') or 0):.2f}",
        f"- p95_latency_ms：{float(summary.get('p95_latency_ms') or 0):.2f}",
        f"- error_count：{summary.get('error_count', 0)}",
        "",
        "## 本次变更",
        "",
        f"- 标题：{change_set.title}",
        f"- owner：{change_set.owner or ''}",
        f"- 字段数：{len(change_set.fields)}",
        f"- 枚举数：{len(change_set.enums)}",
        f"- value mapping 字段数：{len(change_set.value_mappings)}",
        f"- L2 规则数：{len(change_set.l2_rules)}",
        "",
        "## 层级分布",
        "",
        "| matched_level | 数量 |",
        "| --- | --- |",
    ]

    for level, count in sorted((summary.get("level_distribution") or {}).items()):
        lines.append(f"| {level} | {count} |")

    lines.extend(["", "## 验收失败项", ""])
    if acceptance_failures:
        lines.extend(f"- {item}" for item in acceptance_failures)
    else:
        lines.append("- 无")

    lines.extend(
        [
            "",
            "## 失败样本",
            "",
            "| id | query | 归因 |",
            "| --- | --- | --- |",
        ]
    )
    if failed_cases:
        for item in failed_cases[:50]:
            reason = _guess_failure_reason(item)
            query = str(item.get("query") or "").replace("|", "\\|")
            lines.append(f"| {item.get('id')} | {query} | {reason} |")
    else:
        lines.append("| - | - | 无 |")

    lines.extend(
        [
            "",
            "## 下一轮行动",
            "",
            "- [ ] 查看 failed_cases.jsonl 中的失败样本。",
            "- [ ] 对 field_not_recalled 补 retrieval_text/examples 后重建字段索引。",
            "- [ ] 对 l2_false_positive 收紧 enhanced_rules 正则上下文。",
            "- [ ] 对 enum_not_normalized 补 value_mappings 或枚举值。",
        ]
    )
    return "\n".join(lines) + "\n"


def _guess_failure_reason(case: dict[str, Any]) -> str:
    if case.get("error"):
        return "api_error"
    comparison = case.get("comparison") or {}
    actual = case.get("actual") or {}
    expected = case.get("expected") or {}
    if expected.get("conditions") and not actual.get("conditions"):
        return "empty_result_or_field_not_recalled"
    if not expected.get("conditions") and actual.get("conditions"):
        return "false_positive"
    if comparison.get("missing_conditions"):
        return "condition_missing_or_value_wrong"
    if comparison.get("unexpected_conditions"):
        return "unexpected_condition"
    if not comparison.get("query_logic_match"):
        return "logic_wrong"
    return "unknown"
