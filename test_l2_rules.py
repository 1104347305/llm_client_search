"""
L2规则引擎测试脚本

功能：
1. 直接测试 Level2EnhancedMatcher 的匹配效果
2. 不依赖API服务，可以离线测试
3. 输出详细的匹配结果和统计信息
4. 支持导出到Excel

使用方法：
python test_l2_rules.py
"""
import asyncio
import sys
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any
import json

# 添加项目路径
sys.path.insert(0, str(Path(__file__).parent))

from app.core.level2_enhanced_matcher import Level2EnhancedMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


class L2RuleTester:
    """L2规则引擎测试器"""

    def __init__(self, config_path: str = "config/enhanced_rules.yaml"):
        """初始化测试器"""
        self.matcher = Level2EnhancedMatcher(config_path)
        self.results = []

    async def test_query(self, query: str) -> Dict[str, Any]:
        """测试单个查询"""
        try:
            conditions, remaining_text, has_residual = await self.matcher.match(query)

            return {
                "success": True,
                "query": query,
                "matched": len(conditions) > 0,
                "conditions_count": len(conditions),
                "conditions": [
                    {
                        "field": cond.field,
                        "operator": cond.operator.value,
                        "value": str(cond.value) if cond.value is not None else None
                    }
                    for cond in conditions
                ],
                "remaining_text": remaining_text,
                "has_residual": has_residual
            }
        except Exception as e:
            return {
                "success": False,
                "query": query,
                "matched": False,
                "conditions_count": 0,
                "conditions": [],
                "remaining_text": "",
                "has_residual": False,
                "error": str(e)
            }

    async def run_batch_test(self, queries: List[str]):
        """批量测试"""
        print(f"开始测试 {len(queries)} 个查询...")
        print(f"规则数量: {self.matcher.get_rules_count()}")
        print(f"枚举字段: {list(self.matcher.enum_values.keys())}")
        print("-" * 80)

        for i, query in enumerate(queries, 1):
            print(f"\n[{i}/{len(queries)}] 测试: {query}")
            result = await self.test_query(query)
            self.results.append(result)

            if result["success"]:
                if result["matched"]:
                    print(f"  ✓ 匹配成功: {result['conditions_count']} 个条件")
                    for cond in result["conditions"]:
                        print(f"    - {cond['field']} {cond['operator']} {cond['value']}")
                    if result["has_residual"]:
                        print(f"  ⚠ 剩余文本: {result['remaining_text']}")
                else:
                    print(f"  ✗ 未匹配")
            else:
                print(f"  ✗ 错误: {result.get('error')}")

        print("\n" + "=" * 80)
        print("测试完成!")
        self._print_statistics()

    def _print_statistics(self):
        """打印统计信息"""
        total = len(self.results)
        matched = sum(1 for r in self.results if r["matched"])
        unmatched = total - matched
        has_residual = sum(1 for r in self.results if r["has_residual"])

        print(f"\n统计信息:")
        print(f"  总查询数: {total}")
        print(f"  匹配成功: {matched} ({matched/total*100:.1f}%)")
        print(f"  未匹配: {unmatched} ({unmatched/total*100:.1f}%)")
        print(f"  有剩余文本: {has_residual} ({has_residual/total*100:.1f}%)")

        # 按条件数统计
        cond_stats = {}
        for r in self.results:
            count = r["conditions_count"]
            cond_stats[count] = cond_stats.get(count, 0) + 1

        print(f"\n条件数分布:")
        for count in sorted(cond_stats.keys()):
            print(f"  {count} 个条件: {cond_stats[count]} 个查询")

        # 按字段统计
        field_stats = {}
        for r in self.results:
            for cond in r["conditions"]:
                field = cond["field"]
                field_stats[field] = field_stats.get(field, 0) + 1

        print(f"\n字段使用频率 (Top 10):")
        sorted_fields = sorted(field_stats.items(), key=lambda x: x[1], reverse=True)[:10]
        for field, count in sorted_fields:
            print(f"  {field}: {count} 次")

    def export_to_excel(self, output_file: str):
        """导出结果到Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "L2规则测试结果"

        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # 表头
        headers = ["序号", "查询问题", "是否匹配", "条件数", "匹配条件", "剩余文本", "备注"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 填充数据
        for i, result in enumerate(self.results, 2):
            ws.cell(i, 1, i - 1)  # 序号
            ws.cell(i, 2, result["query"])  # 查询问题
            ws.cell(i, 3, "是" if result["matched"] else "否")  # 是否匹配
            ws.cell(i, 4, result["conditions_count"])  # 条件数

            # 匹配条件（JSON格式）
            if result["conditions"]:
                conditions_str = json.dumps(result["conditions"], ensure_ascii=False, indent=2)
                ws.cell(i, 5, conditions_str)
            else:
                ws.cell(i, 5, "无")

            # 剩余文本
            ws.cell(i, 6, result.get("remaining_text", ""))

            # 备注
            if not result["success"]:
                ws.cell(i, 7, f"错误: {result.get('error')}")
            elif result["has_residual"]:
                ws.cell(i, 7, "有剩余文本")
            else:
                ws.cell(i, 7, "")

        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 20

        # 保存文件
        wb.save(output_file)
        print(f"\n结果已导出到: {output_file}")


def load_test_queries(file_path: str = None) -> List[str]:
    """加载测试查询"""
    # 如果提供了文件路径，从文件读取
    if file_path and Path(file_path).exists():
        with open(file_path, 'r', encoding='utf-8') as f:
            queries = []
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    queries.append(line)
            return queries

    # 否则使用内置测试用例
    return [
        # 姓名相关
        '姓张的客户',
        '名字带伟的客户',
        '张珊',
        '被保人名叫李四',
        '投保人姓王',
        '受益人叫张三',

        # 手机号相关
        '15817760299',
        '手机号158开头的客户',
        '手机号段581776的客户',
        '手机尾号0299的客户',

        # 身份证相关
        '身份证号510101196109291482',
        '510101开头的身份证客户',
        '身份证尾号9291482的客户',

        # 保单号相关
        '保单号为：P644037341678127',
        '保单号P644037的客户',
        '保单号78127结尾的客户',

        # 客户号相关
        '客户号为C335906420260306的客户',
        'C335906的客户',
        '420260306',

        # 生日相关
        '本月生日的客户',
        '下个月生日的客户',
        '未来一周生日的客户',

        # 年龄相关
        '1953年出生的客户',
        '20至30岁的客户',
        '18-40岁客户',
        '61岁以上客户',
        '二十多岁的客户',

        # 学历相关
        '查找本科学历以上的客户',
        '学历为本科的客户',
        '大专学历的客户',

        # 客温相关
        '中温客户',
        '客温为高温的客户',
        '低温和中温的客户',
        '中温及以上的客户',

        # 标签相关
        '邻退小康',
        '有邻退小康标签的客户',
        '有哪些中年焦虑的客户',

        # 客户等级相关
        '黄金V1',
        '铂金以上客户',
        '原黄金VIP客户',

        # 客户角色相关
        '仅仅是投保人的客户',

        # 保单状态相关
        '存续单客户',
        '在职有效单的客户',

        # 产品相关
        '买了金瑞人生20，但是没有配置盛世金越的客户',
        '购买了学平险产品的客户',
        '未购买学平险产品的客户',
        '购买了意健险的客户',
        '未购买意健险的客户',
        '购买过e生保，并且生效中的客户',
        '购买过两全产品的客户',
        '购买过守护重疾26的客户',
        '平安福客户',
        '金越司庆版客户',

        # 理赔相关
        '有过综拓产品理赔报案的客户',
        '去年理赔的客户',
        'e生保理赔过的客户',
        '平安福理赔客户',

        # 居家相关
        '居家潜客',
        '居家客户',
        '居家等级V1的客户',

        # 会员相关
        '预达标康养客户',
        '逸享会员客户',
        '安有护国际版客户',
        '安有护客户',
        '预达标臻享家医客户',
        '臻享家医客户',

        # 子女相关
        '未成年子女',
        '子女在小学阶段的客户',
        '子女3-5周岁的客户',
        '有小朋友的客户',
        '子女在上初中/高中的客户',

        # 保单到期/缴费相关
        '10天内到期的短期保单',
        '有等待续保保单的客户',
        '有应缴日在下周的客户',
        '下个月需要缴费的客户',

        # 条款相关
        '有除责条款的客户',
        '有降档条款的客户',
        '有减费的客户',

        # 生存金相关
        '未领取生存金的客户',
        '有生存金利息没领的客户',
        '生存金账户余额超过5万的客户',

        # 婚姻相关
        '已婚客户',
        '刚结婚的二十多岁青年家庭',
        '二十多岁刚结婚的青年客户',
        '20-30岁刚结婚的青年客户',
        '未婚客户',

        # 资产相关
        '有车的客户',
        '有房的客户',
        '家庭年收入50万以上的客户',
        '有房产且未配置家财险的客户',

        # 复杂组合条件
        '45岁以上未配置养老险的客户',
        '家里有小朋友但没买教育金的客户',
        '已婚、有车、没买百万医疗的客户',
        '35岁有小朋友还没配置重疾险的客户',
        '已婚、有车、没有配置百万医疗保险的客户',
        '35岁已婚、有子女、未配置重疾险的客户',
        '未配置医疗险客户',
        '未购买百万医疗保险的客户',
        '青年客户没有e生保的客户',
        '中年重疾保额低的客户',
        '没有附加险的客户',
        '5岁以上、已婚、有子女、有经济基础、未配置养老险和意外险的客户',
        '5岁以上、已婚、有子女、有经济基础、未配置养老险但配置了寿险的客户',
        '有哪些客户买了医疗险，但是没有买养老险',
        '查找买了医疗险、寿险但是没有买年金险的客户',
        '大连金州区、30-40岁、年交保费10万以上、有万能险的客户',

        # 其他
        '我离职员工的孤儿单客户',
        '身份证过期的客户',
        '最近1年内承保的客户',
        '2018年7月投保的客户',
        '客户年交保费20万以上',
        '有重疾险客户',
        '所有万能险客户名单',
        '有意外医疗客户',
    ]


async def main():
    """主函数"""
    # 配置
    queries_file = "test_queries_v1.txt" if len(sys.argv) < 2 else sys.argv[1]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"l2_test_results_{timestamp}.xlsx"

    # 加载测试查询
    if Path(queries_file).exists():
        print(f"从文件加载测试查询: {queries_file}")
        queries = load_test_queries(queries_file)
    else:
        print("使用内置测试查询")
        queries = load_test_queries()

    # 创建测试器
    tester = L2RuleTester()

    # 运行批量测试
    await tester.run_batch_test(queries)

    # 导出结果
    tester.export_to_excel(output_file)


if __name__ == "__main__":
    asyncio.run(main())
