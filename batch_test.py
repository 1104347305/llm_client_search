"""
批量测试脚本 - 测试自然语言搜索接口并导出结果到Excel

功能：
1. 从 test_queries.txt 读取测试问题
2. 调用自然语言搜索接口
3. 将查询条件和客户列表保存到 Excel 文件

依赖：
pip install openpyxl httpx
"""
import asyncio
import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any
import json


class BatchTester:
    """批量测试器"""

    def __init__(self, base_url: str = "http://localhost:8080"):
        self.base_url = base_url
        self.results = []

    async def test_query(self, query: str, agent_id: str = "A000000") -> Dict[str, Any]:
        """测试单个查询"""
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                response = await client.post(
                    f"{self.base_url}/api/v1/search/natural",
                    json={
                        "query": query,
                        "agent_id": agent_id,
                        "page": 1,
                        "size": 100  # 获取更多结果
                    }
                )
                response.raise_for_status()
                return response.json()
        except Exception as e:
            return {
                "success": False,
                "message": str(e),
                "data": {},
                "matched_level": 0,
                "confidence": 0.0,
                "conditions": [],
                "query_logic": "AND"
            }

    async def run_batch_test(self, queries_file: str):
        """批量测试"""
        # 读取测试问题
        queries = []
        with open(queries_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    queries.append(line)

        print(f"共加载 {len(queries)} 个测试问题")

        # 逐个测试
        for i, query in enumerate(queries, 1):
            print(f"\n[{i}/{len(queries)}] 测试: {query}")
            result = await self.test_query(query)

            # 保存结果
            self.results.append({
                "query": query,
                "result": result
            })

            # 打印简要信息
            if result.get("success"):
                print(f"  ✓ Level {result.get('matched_level')}, "
                      f"置信度 {result.get('confidence')}, "
                      f"条件数 {result.get('conditions', [])}, "
                      f"结果数 {result.get('data', {}).get('data', {}).get('total', 0)}")
            else:
                print(f"  ✗ 失败: {result.get('message')}")

            # 避免请求过快
            await asyncio.sleep(0.1)

        print(f"\n测试完成！共 {len(self.results)} 个查询")

    def export_to_excel(self, output_file: str):
        """导出结果到Excel"""
        wb = Workbook()

        # 创建汇总表
        ws_summary = wb.active
        ws_summary.title = "测试结果"

        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # 表头：序号、查询问题、Level、查询条件、客户列表
        summary_headers = ["序号", "查询问题", "Level", "查询条件（检索接口入参）", "客户列表（最多10条）"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 填充数据
        for i, item in enumerate(self.results, 2):
            query = item["query"]
            result = item["result"]

            # 序号
            ws_summary.cell(i, 1, i - 1)

            # 查询问题
            ws_summary.cell(i, 2, query)

            # Level
            ws_summary.cell(i, 3, result.get("matched_level", 0))

            # 查询条件（单行JSON格式）
            conditions = result.get("conditions") or []
            query_logic = result.get("query_logic", "AND")
            conditions_json = {
                "query_logic": query_logic,
                "conditions": [
                    {
                        "field": cond.get("field", ""),
                        "operator": cond.get("operator", ""),
                        "value": cond.get("value", "")
                    }
                    for cond in conditions
                ]
            }
            ws_summary.cell(i, 4, json.dumps(conditions_json, ensure_ascii=False))

            # 客户列表（最多10条，逗号分隔）
            customers = result.get("data", {}).get("data", {}).get("list", [])[:10]
            if customers:
                customer_list = []
                for customer in customers:
                    customer_info = f"{customer.get('name', '')}({customer.get('customer_id', '')})-{customer.get('mobile_phone', '')}"
                    customer_list.append(customer_info)
                ws_summary.cell(i, 5, ", ".join(customer_list))
            else:
                ws_summary.cell(i, 5, "无结果")

        # 调整列宽
        ws_summary.column_dimensions['A'].width = 8
        ws_summary.column_dimensions['B'].width = 50
        ws_summary.column_dimensions['C'].width = 10
        ws_summary.column_dimensions['D'].width = 60
        ws_summary.column_dimensions['E'].width = 60

        # 保存文件
        wb.save(output_file)
        print(f"\n结果已导出到: {output_file}")


async def main():
    """主函数"""
    # 配置
    queries_file = "test_queries_v1.txt"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"test_results_{timestamp}.xlsx"

    # 检查文件是否存在
    if not Path(queries_file).exists():
        print(f"错误: 找不到测试问题文件 {queries_file}")
        return

    # 创建测试器
    tester = BatchTester(base_url="http://localhost:8080")

    # 运行批量测试
    await tester.run_batch_test(queries_file)

    # 导出结果
    tester.export_to_excel(output_file)

    print("\n测试统计:")
    total = len(tester.results)
    success = sum(1 for r in tester.results if r["result"].get("success"))
    print(f"  总数: {total}")
    print(f"  成功: {success}")
    print(f"  失败: {total - success}")

    # 按层级统计
    level_stats = {}
    for r in tester.results:
        level = r["result"].get("matched_level", 0)
        level_stats[level] = level_stats.get(level, 0) + 1

    print("\n层级分布:")
    for level in sorted(level_stats.keys()):
        print(f"  Level {level}: {level_stats[level]} 个")


if __name__ == "__main__":
    asyncio.run(main())
